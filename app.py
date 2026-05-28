# app.py
import math
import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import multiphase_engine as engine
import validation_cases as val_cases
import report_generator
import fanno_engine as fanno
import ro_engine as ro
import psv_engine as psv
import cv_engine as cv
import hashlib
from physics.friction import churchill_f
from workflows.pipeline_case import compute_pipeline_case
from workflows.pump_case import compute_pump_case
from standards.piping import sum_le_fit
import json
from fluids.two_phase import (Taitel_Dukler_regime as _TD_regime,
                               Mandhane_Gregory_Aziz_regime as _MGA_regime)

st.set_page_config(
    page_title="FlowBench — General Flow Workbench",
    page_icon="⚙️",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
/* ── Metric cards ─────────────────────────────────────────── */
[data-testid="stMetricValue"] {
    font-size: 1.45rem !important;
    font-weight: 600 !important;
    color: #0F172A !important;
    letter-spacing: -0.01em !important;
}
[data-testid="stMetricLabel"] {
    font-size: 0.82rem !important;
    font-weight: 500 !important;
    color: #475569 !important;
    text-transform: none !important;
    letter-spacing: normal !important;
}
[data-testid="stMetricDelta"] { font-size: 0.80rem !important; }

/* ── Section labels inside containers (**Bold** pattern) ─── */
[data-testid="stMarkdownContainer"] p strong {
    font-size: 0.78rem !important;
    font-weight: 700 !important;
    letter-spacing: 0.06em !important;
    text-transform: uppercase !important;
    color: #64748B !important;
}

/* ── Body text and captions ───────────────────────────────── */
[data-testid="stMarkdownContainer"] p {
    font-size: 0.92rem;
    line-height: 1.55;
    color: #1E293B;
}
[data-testid="stCaptionContainer"] p {
    font-size: 0.82rem !important;
    color: #64748B !important;
    line-height: 1.5 !important;
}

/* ── Subheaders ───────────────────────────────────────────── */
h3 {
    font-size: 1.05rem !important;
    font-weight: 600 !important;
    color: #1E293B !important;
    letter-spacing: -0.01em !important;
}

/* ── Dividers ─────────────────────────────────────────────── */
hr { margin: 0.75rem 0 !important; }
</style>
""", unsafe_allow_html=True)

st.title("FlowBench")

# ============================================================================
# SAVE / LOAD  (defined here so sidebar can call them on first render)
# ============================================================================
def _collect_save_state() -> dict:
    s = st.session_state
    data: dict = {
        "version": 1,
        "label_a": s.get("label_a", "Case A"),
        "label_b": s.get("label_b", "Case B"),
    }
    for cid in ("a", "b"):
        gas_flows = {
            k[len(f"{cid}_gflow_"):]: v
            for k, v in s.items()
            if k.startswith(f"{cid}_gflow_")
        }
        data[cid] = {
            "segments":           s.get(f"{cid}_segments", []),
            "P_bara":             s.get(f"{cid}_P_bara", 20.0),
            "T_C":                s.get(f"{cid}_T_C", 60.0),
            "gas_species_widget": s.get(f"{cid}_gas_species_widget", ["H₂"]),
            "gas_flows":          gas_flows,
            "correlation":        s.get(f"{cid}_correlation", "Beggs-Brill"),
            "voidage_method":     s.get(f"{cid}_voidage_method", "Homogeneous"),
            "cg_mw":              s.get(f"{cid}_cg_mw"),
            "cg_mu":              s.get(f"{cid}_cg_mu"),
            "cl_rho":             s.get(f"{cid}_cl_rho"),
            "cl_mu":              s.get(f"{cid}_cl_mu"),
            "cl_sigma":           s.get(f"{cid}_cl_sigma"),
        }
    for cid in ("c", "d"):
        n_left  = int(s.get(f"{cid}_n_left",  3))
        n_right = int(s.get(f"{cid}_n_right", 3))
        data[cid] = {
            "P_target_sep":    s.get(f"{cid}_P_target_sep", 16.5),
            "T_C":             s.get(f"{cid}_T_C", 60.0),
            "hdr_dn":          s.get(f"{cid}_hdr_dn", "DN100"),
            "hdr_pn":          s.get(f"{cid}_hdr_pn", "PN40"),
            "hdr_mat":         s.get(f"{cid}_hdr_mat", "SS316L"),
            "hdr_lined":       s.get(f"{cid}_hdr_lined", False),
            "hdr_lmat":        s.get(f"{cid}_hdr_lmat", "FEP"),
            "hdr_lthk":        s.get(f"{cid}_hdr_lthk", 1.0),
            "hdr_fits":        s.get(f"{cid}_hdr_fits", []),
            "n_left":          n_left,
            "n_right":         n_right,
            "left_positions":  [s.get(f"{cid}_pos_left_{i}",  float((i+1)*2.5)) for i in range(n_left)],
            "right_positions": [s.get(f"{cid}_pos_right_{i}", float((i+1)*2.5)) for i in range(n_right)],
            "t_dn":            s.get(f"{cid}_t_dn", "DN150"),
            "t_pn":            s.get(f"{cid}_t_pn", "PN40"),
            "t_mat":           s.get(f"{cid}_t_mat", "SS316L"),
            "t_len":           s.get(f"{cid}_t_len", 1.0),
            "correlation":     s.get(f"{cid}_correlation", "Beggs-Brill"),
            "voidage_method":  s.get(f"{cid}_voidage_method", "Homogeneous"),
        }
    return data


def _apply_save_state(data: dict) -> None:
    import copy
    s = st.session_state
    s["label_a"] = data.get("label_a", "Case A")
    s["label_b"] = data.get("label_b", "Case B")

    for cid in ("a", "b"):
        cd = data.get(cid, {})
        if not cd:
            continue
        if "segments" in cd:
            s[f"{cid}_segments"] = copy.deepcopy(cd["segments"])
        for key in ("P_bara", "T_C", "gas_species_widget", "correlation", "voidage_method",
                    "cg_mw", "cg_mu", "cl_rho", "cl_mu", "cl_sigma"):
            if cd.get(key) is not None:
                s[f"{cid}_{key}"] = cd[key]
        for sp, flow in cd.get("gas_flows", {}).items():
            s[f"{cid}_gflow_{sp}"] = flow

    for cid in ("c", "d"):
        cd = data.get(cid, {})
        if not cd:
            continue
        for key in ("P_target_sep", "T_C", "hdr_dn", "hdr_pn", "hdr_mat",
                    "hdr_lined", "hdr_lmat", "hdr_lthk", "hdr_fits",
                    "n_left", "n_right", "t_dn", "t_pn", "t_mat", "t_len",
                    "correlation", "voidage_method"):
            if cd.get(key) is not None:
                s[f"{cid}_{key}"] = cd[key]
        for i, pos in enumerate(cd.get("left_positions", [])):
            s[f"{cid}_pos_left_{i}"] = float(pos)
        for i, pos in enumerate(cd.get("right_positions", [])):
            s[f"{cid}_pos_right_{i}"] = float(pos)


# ============================================================================
# SIDEBAR
# ============================================================================
with st.sidebar:
    st.header("FlowBench")
    with st.expander("About & Capabilities", expanded=False):
        st.markdown("""
        Engineering workbench for steady-state flow calculations — pipe hydraulics,
        equipment sizing, and fluid thermodynamics in a single browser-based tool.

        ---
        **Pipeline Hydraulics** (tabs A / B, Header A / B, Compare, Goal Seek)

        Segment-by-segment pressure marching; fluid properties re-evaluated at each
        segment from the current pressure.

        | Flow mode | Approach |
        |---|---|
        | Single-phase liquid | Darcy-Weisbach, Churchill friction, CoolProp ρ/μ/σ; KOH empirical model (20–40 wt%) |
        | Single-phase gas | Isothermal compressible Darcy-Weisbach; ideal-gas density marched with pressure |
        | Gas + liquid (two-phase) | 6 ΔP correlations × 2 void-fraction models; ΔP split into frictional, gravitational, accelerational |
        | Saturated / VLE | Single-component pure fluid; quality evolves isenthalpically; CoolProp saturation tables |

        Two-phase correlations: Beggs-Brill, Friedel, Lockhart-Martinelli,
        Müller-Steinhagen & Heck, Chisholm, Kim-Mudawar.
        Void-fraction: Homogeneous, Rouhani-1 drift-flux.

        Outputs: segment ΔP table · pipeline schematic · pressure profile ·
        horizontal and vertical flow regime maps with operating points ·
        phase distribution (VLE) · method sensitivity sweep · API RP 14E erosion
        check · valve Kv sizing · export to Word or Excel.

        **Header cases** compute the governing-arm pressure drop and goal-seek to a
        target separator pressure. **Compare** overlays Cases A and B across all
        12 method combinations. **Goal Seek** back-calculates inlet or separator
        pressure for both systems simultaneously.

        **Pipe library** — DN20–DN250, PN20/25/40, 5 materials (SS316L, Duplex,
        Carbon Steel, Hastelloy C-276, Titanium Gr. 2), optional fluoropolymer liner
        (PTFE, FEP, PFA, PVDF). 17 fitting types (Crane TP-410). Inline components:
        control valves (Kv or ΔP mode), heat exchangers.

        ---
        **Engineering Calculators**

        | Tool | Description |
        |---|---|
        | **Fanno Flow** | Adiabatic compressible duct flow — inlet Mach from conditions, friction → exit Mach, static and stagnation properties, choking margin |
        | **RO (Restriction Orifice)** | ISO 5167 orifice sizing for gas and liquid; single-stage and multistage arrays; Reader-Harris/Gallagher Cd |
        | **PSV (Pressure Safety Valve)** | API 520 / 526 area sizing for gas, steam, liquid relief; back-pressure correction; standard orifice letter selection |
        | **Control Valve** | IEC 60534 Kv/Cv for liquid (cavitation and flashing checks), gas, and steam; suggested body size at target opening |
        | **Dissolved Gas Flash** | Henry's law dissolution (H₂, O₂, CO₂, N₂, CH₄) in water and KOH; flash calculation on depressurisation |
        | **Pump** | H-Q curve fit (3-point or tabular), system curve, operating point, speed-scaling (affinity laws), NPSH check, shaft and motor power; PD pump design pressure |
        | **Line Size** | DN selection meeting velocity and ΔP/100 m criteria; service presets for process liquid, pump suction/discharge, gas, steam, slurry |
        """)
    with st.expander("Model details — Pipeline Hydraulics", expanded=False):
        st.markdown("""
        **Assumptions**
        1. Gas density: ideal-gas law (ρ = PM/RT); viscosity from CoolProp
        2. Liquid: incompressible at local T and P; properties from CoolProp
        3. KOH solution: empirical fits — ρ (ICT/Perry's), μ (Vogel equation),
           σ (water baseline + KOH correction); valid 10–90 °C, 0–40 wt%
        4. VLE mode: isenthalpic flash — inlet total enthalpy fixed at (P_in, x_in);
           quality re-derived at each segment from CoolProp saturation tables
        5. Bore = f(DN, PN) only — ANSI B36.10/19, material-independent wall thickness
        6. Lined pipe: effective ID = metal bore − 2 × liner thickness
        7. Pressure marching: all properties re-evaluated at each segment inlet pressure
        8. Steady-state only — no transient, surge, or waterhammer
        9. Two-phase void fraction: homogeneous α = (x/ρg) / (x/ρg + (1−x)/ρl),
           or Rouhani-1 drift-flux model
        10. Single-phase gas pipeline: isothermal compressible Darcy-Weisbach
            (for adiabatic compressible duct flow use the Fanno Flow calculator)

        **Flow regime classification (two-phase)**
        - Horizontal (|θ| ≤ 15°): Taitel-Dukler (1976) + Mandhane-Gregory-Aziz (1974)
          → Stratified / Intermittent (Slug) / Annular-Dispersed / Bubbly
        - Vertical upflow (|θ| ≥ 75°): Wallis annular-onset criterion + void-fraction
          thresholds → Bubbly / Slug / Churn / Annular
        - Vertical downflow: Falling Film / Annular

        **Liquid species**
        - CoolProp: Water (IAPWS-IF97), organics, refrigerants (DIPPR / REFPROP)
        - Mixtures: ρ and σ mass-weighted, μ log-mean
        - KOH solution: built-in empirical model, concentration 20–40 wt%

        **References** — Beggs & Brill (1973) SPE-4007-PA · Friedel (1979) ·
        Lockhart & Martinelli (1949) · Müller-Steinhagen & Heck (1986) ·
        Chisholm (1973) · Kim & Mudawar (2012) · CoolProp (Bell et al. 2014) ·
        fluids library · Crane TP-410 (2013) · API RP 14E (2007)
        """)
    with st.expander("Validation", expanded=False):
        st.markdown("Run a built-in reference case and compare the calculated ΔP against the expected value.")
        _sb_cases     = val_cases.list_validation_cases()
        _sb_case_opts = [item[0] for item in _sb_cases]
        _sb_sel       = st.selectbox("Reference case", options=_sb_case_opts,
                                     format_func=lambda x: val_cases.get_validation_case(x)["name"],
                                     key="sb_val_sel")
        if _sb_sel:
            _sb_case = val_cases.get_validation_case(_sb_sel)
            st.info(val_cases.get_case_info(_sb_sel))
            if st.button("Run", key="sb_run_val", width='stretch'):
                _sb_vi   = _sb_case["inputs"]
                _sb_mode = _sb_case.get("mode", "gas_liquid")
                _sb_D    = engine.PIPE_DATABASE[_sb_vi["pipe_dn"]][_sb_vi["pipe_pn"]]
                _sb_r    = engine.MATERIAL_ROUGHNESS["SS316L"]
                _sb_dp   = 0.0
                _sb_P_Pa = _sb_vi["P_bara"] * 1e5
                try:
                    for _sb_seg in _sb_vi["segments"]:
                        _sb_ang = {"Horizontal": 0.0, "Vertical Upflow": np.pi/2,
                                   "Vertical Downflow": -np.pi/2}[_sb_seg["type"]]
                        _sb_le  = 0.0
                        if _sb_seg.get("fittings", "None") in engine.FITTING_Le_over_D:
                            _sb_le = (engine.FITTING_Le_over_D[_sb_seg["fittings"]]
                                      * _sb_D * _sb_seg.get("fitting_count", 0))
                        if _sb_mode == "vle":
                            _sb_vp = engine.calculate_vle_properties(
                                _sb_vi["vle_fluid"], _sb_P_Pa / 1e5,
                                _sb_vi["vle_x_mass"], _sb_vi["vle_m_total_kgs"])
                        else:
                            _sb_vp = engine.calculate_two_phase_properties(
                                _sb_P_Pa / 1e5, _sb_vi["T_C"],
                                _sb_vi.get("gas_flows_kgh", {}),
                                "Custom", 0.0,
                                liquid_flows_kgh=_sb_vi.get("liquid_flows_kgh"))
                        _sb_res = engine.calculate_segment_pressure_drop(
                            _sb_vp, _sb_D, _sb_r, _sb_seg["length"] + _sb_le, _sb_ang)
                        _sb_dp  += _sb_res["dP_Pa"]
                        _sb_P_Pa = max(1e4, _sb_P_Pa - _sb_res["dP_Pa"])
                    _sb_calc = _sb_dp / 1000.0
                    _sb_exp  = _sb_case["expected_total_dp_kpa"]
                    _sb_tol  = _sb_case.get("tolerance_pct", 5.0)
                    r1, r2, r3 = st.columns(3)
                    r1.metric("Calculated", f"{_sb_calc:.3f} kPa")
                    if _sb_exp is None:
                        r2.metric("Expected", "—  (auto-calib.)")
                        r3.metric("Deviation", "—")
                        st.info(f"Auto-calibrated case: {_sb_calc:.3f} kPa (no reference to compare).")
                    else:
                        _sb_err  = abs(_sb_calc - _sb_exp) / abs(_sb_exp) * 100.0
                        _sb_pass = _sb_err <= _sb_tol
                        r2.metric("Expected",  f"{_sb_exp:.3f} kPa")
                        r3.metric("Deviation", f"{_sb_err:.2f} %",
                                  delta=f"Pass (≤{_sb_tol:.0f}%)" if _sb_pass else f"Fail (>{_sb_tol:.0f}%)",
                                  delta_color="normal" if _sb_pass else "inverse")
                        if _sb_pass:
                            st.success(f"Passed — {_sb_err:.2f}% within ±{_sb_tol:.0f}%")
                        else:
                            st.warning(f"Failed — {_sb_err:.2f}% exceeds ±{_sb_tol:.0f}%")
                except Exception as _sb_err_ex:
                    st.error(f"Validation run failed: {_sb_err_ex}")

    st.divider()
    st.header("Session")

    # ── Save ──────────────────────────────────────────────────────────────────
    _save_json = json.dumps(_collect_save_state(), indent=2, ensure_ascii=False)
    st.download_button(
        "Save session (.json)",
        data=_save_json,
        file_name="hydraulic_session.json",
        mime="application/json",
        width='stretch',
        help="Download all current inputs as a JSON file you can reload later.",
    )

    # ── Load ──────────────────────────────────────────────────────────────────
    _uploaded = st.file_uploader(
        "Load session (.json)",
        type="json",
        label_visibility="collapsed",
        help="Upload a previously saved session JSON to restore all inputs.",
        key="session_uploader",
    )
    if _uploaded is not None:
        try:
            _loaded = json.loads(_uploaded.read())
            if _loaded.get("version") != 1:
                st.warning("Unrecognised file format — not loaded.")
            else:
                _apply_save_state(_loaded)
                st.success("Session loaded.")
                st.rerun()
        except Exception as _e:
            st.error(f"Could not load session: {_e}")

    st.divider()
    st.caption(
        "**Disclaimer** — FlowBench is provided for general reference only. "
        "No warranty is given for accuracy, completeness, or fitness for any "
        "particular purpose. The authors accept no liability for errors, "
        "omissions, bugs, or misuse of this tool. "
        "Validate all results independently before use in any design, "
        "procurement, or safety-critical application."
    )

# ============================================================================
# PRESETS  (shared across both cases)
# ============================================================================

_DEFAULT_SEGMENTS = [
    {"type": "Horizontal",      "dn": "DN50", "pn": "PN40", "material": "SS316L",
     "length": 20.0,
     "fittings_list": [{"type": "90° Standard Elbow", "qty": 2}],
     "lined": False, "liner_material": "FEP", "liner_thickness_mm": 1.0},
    {"type": "Vertical Upflow", "dn": "DN50", "pn": "PN40", "material": "SS316L",
     "length": 10.0,
     "fittings_list": [],
     "lined": False, "liner_material": "FEP", "liner_thickness_mm": 1.0},
]

_VALID_MATS   = set(engine.MATERIAL_ROUGHNESS.keys())
_VALID_LINERS = set(engine.LINER_ROUGHNESS.keys())


# ============================================================================
# REGIME COLOUR HELPERS
# Keyword-based matching so any regime string from the engine is coloured
# correctly, including compound strings like "intermittent / slug".
# Order matters: more specific keywords are checked first.
# ============================================================================
_REGIME_LINE_KW = [          # (keyword, hex colour)  — schematic line colour
    ("churn",        "#DC2626"),   # red
    ("falling",      "#0891B2"),   # cyan  (before "annular" to catch falling film/annular)
    ("bubb",         "#7C3AED"),   # purple (catches both "bubble" and "bubbly")
    ("mist",         "#059669"),   # green
    ("annular",      "#059669"),   # green
    ("slug",         "#D97706"),   # amber
    ("intermittent", "#D97706"),   # amber
    ("elongated",    "#D97706"),   # amber
    ("wave",         "#3B82F6"),   # blue
    ("stratified",   "#3B82F6"),   # blue
    ("dispersed",    "#059669"),   # green
]
_REGIME_FILL_KW = [          # (keyword, rgba fill)   — pressure profile band
    ("churn",        "rgba(254,226,226,0.70)"),
    ("falling",      "rgba(207,250,254,0.70)"),
    ("bubb",         "rgba(237,233,254,0.70)"),
    ("mist",         "rgba(209,250,229,0.70)"),
    ("annular",      "rgba(209,250,229,0.70)"),
    ("slug",         "rgba(254,243,199,0.70)"),
    ("intermittent", "rgba(254,243,199,0.70)"),
    ("elongated",    "rgba(254,243,199,0.70)"),
    ("wave",         "rgba(219,234,254,0.70)"),
    ("stratified",   "rgba(219,234,254,0.70)"),
    ("dispersed",    "rgba(209,250,229,0.70)"),
]
_REGIME_BORDER_KW = [        # (keyword, border colour) — pressure profile band
    ("churn",        "#FCA5A5"),
    ("falling",      "#67E8F9"),
    ("bubb",         "#C4B5FD"),
    ("mist",         "#6EE7B7"),
    ("annular",      "#6EE7B7"),
    ("slug",         "#FCD34D"),
    ("intermittent", "#FCD34D"),
    ("elongated",    "#FCD34D"),
    ("wave",         "#93C5FD"),
    ("stratified",   "#93C5FD"),
    ("dispersed",    "#6EE7B7"),
]

def _regime_color(regime_str, kw_list, default):
    r = regime_str.lower()
    return next((col for kw, col in kw_list if kw in r), default)


@st.cache_data(show_spinner="Computing regime map…")
def _compute_regime_grid(
    rhol: float, rhog: float, mul: float, mug: float,
    sigma: float, D: float, roughness: float, use_horiz: bool,
) -> tuple:
    """
    Sweep a 50×50 log-log V_sl × V_sg grid and classify each cell using
    exactly the same correlations as the engine (Taitel-Dukler 1976 +
    Mandhane-Gregory-Aziz 1974 for horizontal; Wallis + void-fraction
    thresholds for vertical upflow).

    Returns (td_grid, full_grid, vsl_list, vsg_list) — all plain Python
    lists so Streamlit can cache and serialise them cleanly.
    """
    N = 80
    _g = 9.80665
    vsl_arr = np.logspace(-3, 1, N)   # V_sl: 0.001 → 10 m/s
    vsg_arr = np.logspace(-3, 2, N)   # V_sg: 0.001 → 100 m/s
    A = np.pi / 4 * D ** 2

    td_grid   = [[""] * N for _ in range(N)]   # coloring key (TD regime or vert label)
    full_grid = [[""] * N for _ in range(N)]   # hover label

    if use_horiz:
        for i, vsg in enumerate(vsg_arr):
            for j, vsl in enumerate(vsl_arr):
                ml = vsl * rhol * A
                mg = vsg * rhog * A
                m  = ml + mg
                x  = mg / m if m > 0 else 0.0
                try:
                    td  = _TD_regime(m=m, x=x, rhol=rhol, rhog=rhog,
                                     mul=mul, mug=mug, D=D,
                                     angle=0.0, roughness=roughness)[0]
                    mga = _MGA_regime(m=m, x=x, rhol=rhol, rhog=rhog,
                                      mul=mul, mug=mug, sigma=sigma, D=D)[0]
                    td_grid[i][j]   = td
                    full_grid[i][j] = f"{td} / {mga}"
                except Exception:
                    td_grid[i][j]   = "intermittent"
                    full_grid[i][j] = "intermittent / slug"
    else:
        # Vertical upflow: Wallis annular onset + homogeneous void fraction
        # (mirrors the engine's _classify_regime logic for |θ|≥75°)
        try:
            V_ann = 3.1 * (_g * sigma * (rhol - rhog) / rhog ** 2) ** 0.25
        except Exception:
            V_ann = 1e9
        for i, vsg in enumerate(vsg_arr):
            for j, vsl in enumerate(vsl_arr):
                mg = vsg * rhog * A
                ml = vsl * rhol * A
                m  = ml + mg
                x  = mg / m if m > 0 else 0.0
                alpha_h = vsg / (vsg + vsl)   # homogeneous void fraction
                if x > 0.90:
                    reg = "mist / annular"
                elif vsg >= V_ann:
                    reg = "annular"
                elif alpha_h >= 0.52:
                    reg = "churn"
                elif alpha_h >= 0.25:
                    reg = "slug"
                else:
                    reg = "bubble"
                td_grid[i][j]   = reg
                full_grid[i][j] = reg

    return td_grid, full_grid, vsl_arr.tolist(), vsg_arr.tolist()


def _build_regime_fig(
    td_grid, full_grid, vsl_arr, vsg_arr,
    op_recs: list,
    title: str,
) -> "go.Figure":
    """Build a single flow-regime map figure.

    Args:
        td_grid / full_grid : 2-D lists from _compute_regime_grid
        vsl_arr / vsg_arr   : velocity grids (linear values)
        op_recs             : segment records whose operating points to overlay
        title               : chart title string
    """
    _log_vsl = np.log10(vsl_arr)
    _log_vsg = np.log10(vsg_arr)

    _all_regs   = sorted(set(r for row in td_grid for r in row if r))
    _reg_to_idx = {r: i for i, r in enumerate(_all_regs)}
    _idx_to_col = [_regime_color(r, _REGIME_LINE_KW, "#94A3B8") for r in _all_regs]
    _n_reg      = len(_all_regs)

    _z = [[_reg_to_idx.get(td_grid[i][j], 0) for j in range(len(vsl_arr))]
          for i in range(len(vsg_arr))]

    _cs = []
    for _ci, _cc in enumerate(_idx_to_col):
        _cs.extend([[_ci / _n_reg, _cc], [(_ci + 1) / _n_reg, _cc]])

    fig = go.Figure()
    fig.add_trace(go.Heatmap(
        x=list(_log_vsl), y=list(_log_vsg), z=_z,
        text=full_grid,
        colorscale=_cs,
        zmin=0, zmax=_n_reg,
        showscale=False,
        opacity=0.30,
        hovertemplate="Regime: %{text}<extra></extra>",
    ))

    # Zone labels at each regime's log-space centroid
    _zone_acc: dict = {}
    for _gi, _row in enumerate(td_grid):
        for _gj, _reg in enumerate(_row):
            if not _reg:
                continue
            if _reg not in _zone_acc:
                _zone_acc[_reg] = [0.0, 0.0, 0]
            _zone_acc[_reg][0] += float(_log_vsl[_gj])
            _zone_acc[_reg][1] += float(_log_vsg[_gi])
            _zone_acc[_reg][2] += 1

    for _zreg, (_svsl, _svsg, _cnt) in _zone_acc.items():
        if _cnt == 0:
            continue
        fig.add_annotation(
            x=_svsl / _cnt, y=_svsg / _cnt,
            xref="x", yref="y",
            text=f"<b>{_zreg}</b>",
            showarrow=False,
            font=dict(size=9, color="#1E293B"),
            bgcolor="rgba(255,255,255,0.55)",
            borderpad=2,
        )

    # Operating points — cluster identical (V_sl, V_sg) positions
    _op_clusters: dict = {}
    for _r in op_recs:
        _vsg_r = max(_r["V_sg (m/s)"], 1e-4)
        _vsl_r = max(_r["V_sl (m/s)"], 1e-4)
        _ck = (round(np.log10(_vsl_r), 2), round(np.log10(_vsg_r), 2))
        _op_clusters.setdefault(_ck, []).append(_r)

    _seen_reg_map: set = set()
    for _cgroup in _op_clusters.values():
        _vsl_c = float(np.mean([max(r["V_sl (m/s)"], 1e-4) for r in _cgroup]))
        _vsg_c = float(np.mean([max(r["V_sg (m/s)"], 1e-4) for r in _cgroup]))
        _reg_list = [r["Regime"] for r in _cgroup]
        _reg  = max(set(_reg_list), key=_reg_list.count)
        _col  = _regime_color(_reg, _REGIME_LINE_KW, "#64748B")
        _types = {r.get("Type", "Horizontal") for r in _cgroup}
        _sym  = ("circle"   if _types == {"Horizontal"} else
                 "diamond"  if "Horizontal" not in _types else
                 "diamond-wide")
        _n_c  = len(_cgroup)
        _lbl  = ", ".join(r["Seg"] for r in _cgroup)
        _fsz  = max(6, 9 - (_n_c - 1) * 2)
        _msz  = 16 + (_n_c - 1) * 4
        _show_leg = _reg not in _seen_reg_map
        _seen_reg_map.add(_reg)
        _hover_lines = "<br>".join(
            f"<b>{r['Seg']}</b> {r['Pipe']}  {r.get('Type','—')}"
            f"  |  {r['Regime']}"
            f"  |  ΔP {r['ΔP (kPa)']:.3f} kPa"
            for r in _cgroup
        )
        fig.add_trace(go.Scatter(
            x=[float(np.log10(_vsl_c))], y=[float(np.log10(_vsg_c))],
            mode="markers+text",
            marker=dict(size=_msz, color=_col, symbol=_sym,
                        line=dict(color="white", width=1.5)),
            text=[_lbl], textposition="middle center",
            textfont=dict(size=_fsz, color="white"),
            name=_reg, legendgroup=_reg, showlegend=_show_leg,
            hovertemplate=(
                f"{_hover_lines}<br>"
                f"V_sl = {_vsl_c:.4f} m/s<br>"
                f"V_sg = {_vsg_c:.4f} m/s"
                "<extra></extra>"
            ),
        ))

    _xtv = [-3, -2, -1, 0, 1]
    _xtx = ["0.001", "0.01", "0.1", "1", "10"]
    _ytv = [-3, -2, -1, 0, 1, 2]
    _ytx = ["0.001", "0.01", "0.1", "1", "10", "100"]
    fig.update_layout(
        template="plotly_white", height=470,
        margin=dict(l=70, r=20, t=40, b=70),
        xaxis=dict(title="V_sl  superficial liquid velocity (m/s)",
                   type="linear", range=[-3, 1],
                   tickvals=_xtv, ticktext=_xtx,
                   gridcolor="#F1F5F9", linecolor="#E2E8F0"),
        yaxis=dict(title="V_sg  superficial gas velocity (m/s)",
                   type="linear", range=[-3, 2],
                   tickvals=_ytv, ticktext=_ytx,
                   gridcolor="#F1F5F9", linecolor="#E2E8F0"),
        legend=dict(title="Computed regime", bgcolor="rgba(255,255,255,0.9)",
                    bordercolor="#E2E8F0", borderwidth=1, font=dict(size=11)),
        font=dict(size=12, color="#374151"),
        title=dict(text=title, font=dict(size=13), x=0),
    )
    return fig


# ============================================================================
# CASE RUNNER  — renders one full case and returns results for Compare tab
# ============================================================================
def run_case(cid: str, accent: str, default_segments=None) -> dict:
    """
    Render inputs + outputs for one case.

    cid              : "a", "b", or "c"
    accent           : hex colour used for the pressure profile trace
    default_segments : segment list used for first-run initialisation
    Returns dict of results consumed by the Compare tab and report generator.
    """
    k = lambda name: f"{cid}_{name}"   # namespaced session-state / widget key

    # ── Session state init ────────────────────────────────────────────────────
    if k("segments") not in st.session_state:
        import copy
        _init = default_segments if default_segments is not None else _DEFAULT_SEGMENTS
        st.session_state[k("segments")] = copy.deepcopy(_init)
    if k("gas_species_widget") not in st.session_state:
        st.session_state[k("gas_species_widget")] = ["Air"]
    if k("liquid_species_widget") not in st.session_state:
        st.session_state[k("liquid_species_widget")] = ["Water"]
    if k("flow_mode") not in st.session_state:
        st.session_state[k("flow_mode")] = "gas_liquid"
    if k("vle_fluid_widget") not in st.session_state:
        st.session_state[k("vle_fluid_widget")] = "Water"
    if k("P_bara") not in st.session_state:
        st.session_state[k("P_bara")] = 10.0
    if k("T_C") not in st.session_state:
        st.session_state[k("T_C")] = 25.0
    if k("vle_m_kgs_widget") not in st.session_state:
        st.session_state[k("vle_m_kgs_widget")] = 1.0
    if k("vle_x_widget") not in st.session_state:
        st.session_state[k("vle_x_widget")] = 0.3

    # Migrate segments
    for _seg in st.session_state[k("segments")]:
        _seg.setdefault("kind", "pipe")
        _seg.setdefault("dn", "DN50"); _seg.setdefault("pn", "PN40")
        _seg.setdefault("material", "SS316L"); _seg.setdefault("lined", False)
        _seg.setdefault("liner_material", "FEP"); _seg.setdefault("liner_thickness_mm", 1.0)
        if _seg.get("kind", "pipe") == "pipe":
            if _seg["material"] not in _VALID_MATS:    _seg["material"] = "SS316L"
            if _seg["liner_material"] not in _VALID_LINERS: _seg["liner_material"] = "FEP"

    _lbl_c, _ = st.columns([1, 3])
    _lbl_c.text_input("Case label", key=f"label_{cid}", max_chars=20,
                      help="Name shown in tabs, charts, and reports.")

    col_in, col_out = st.columns([1, 1.2])

    # ── INPUTS ────────────────────────────────────────────────────────────────
    with col_in:
        st.subheader("Inputs")

        # Flow mode selector
        _FM_OPTS = {
            "liquid_only": "Single-phase liquid",
            "gas_only":    "Single-phase gas",
            "gas_liquid":  "Gas + liquid (two-phase)",
            "vle":         "Saturated / VLE",
        }
        _FM_BY_LABEL = {v: fk for fk, v in _FM_OPTS.items()}
        _saved_fm = st.session_state.get(k("flow_mode"), "gas_liquid")
        if _saved_fm not in _FM_OPTS:
            _saved_fm = "gas_liquid"  # migrate old values
        _fm_labels = list(_FM_OPTS.values())
        flow_mode = _FM_BY_LABEL[st.radio(
            "Flow mode",
            _fm_labels,
            horizontal=True,
            key=k("flow_mode_radio"),
            index=_fm_labels.index(_FM_OPTS[_saved_fm]),
            help=(
                "**Single-phase liquid** — incompressible Darcy-Weisbach, CoolProp liquid props.  \n"
                "**Single-phase gas** — isothermal compressible Darcy-Weisbach; gas density "
                "recalculated at each segment inlet via the ideal-gas law.  \n"
                "**Gas + liquid (two-phase)** — six industry correlations (Beggs-Brill, Friedel, "
                "Lockhart-Martinelli, Müller-Steinhagen-Heck, Chisholm, Kim-Mudawar) × two void-fraction "
                "models with pressure marching.  \n"
                "**Saturated / VLE** — single-component pure-fluid saturation; CoolProp derives "
                "phase properties at each segment pressure."
            ),
        )]
        st.session_state[k("flow_mode")] = flow_mode
        _is_vle = (flow_mode == "vle")

        _FM_CAPTIONS = {
            "liquid_only": "Incompressible Darcy-Weisbach. Use for pumped liquid lines — water, "
                           "caustic, solvents, etc. CoolProp provides ρ, μ, and σ.",
            "gas_only":    "Darcy-Weisbach with ideal-gas density. Gas density is updated at each "
                           "segment inlet as pressure drops (isothermal compressible). "
                           "No choking check.",
            "gas_liquid":  "Separate gas and liquid streams flowing together. Choose from six "
                           "industry correlations (Beggs-Brill, Friedel, Lockhart-Martinelli, "
                           "Müller-Steinhagen & Heck, Chisholm, Kim-Mudawar) and two void-fraction "
                           "models. Pressure is marched segment by segment.",
            "vle":         "Single pure component on its own saturation curve — e.g. steam/water, "
                           "propane, ammonia, or a refrigerant. T_sat is derived from pressure. "
                           "Vapour quality x evolves along the pipe via isenthalpic flash; "
                           "you set the inlet x only.",
        }
        st.caption(_FM_CAPTIONS[flow_mode])

        # Inlet Conditions
        with st.container(border=True):
            st.markdown("**Inlet Conditions**")
            if _is_vle:
                P_bara = st.number_input("Inlet Pressure (bara)", min_value=0.1, max_value=300.0,
                                         step=1.0, key=k("P_bara"))
                T_C = None  # derived from saturation in VLE mode
            else:
                p1, p2 = st.columns(2)
                P_bara = p1.number_input("Inlet Pressure (bara)", min_value=1.0, max_value=300.0,
                                         step=1.0, key=k("P_bara"))
                T_C    = p2.number_input("Temperature (°C)", min_value=-60.0, max_value=400.0,
                                         step=5.0, key=k("T_C"))

        # ── Helper: render Gas Phase container ───────────────────────────────
        def _render_gas_inputs():
            _all_species = list(engine.GAS_SPECIES.keys())
            _sel = st.multiselect(
                "Gas species  (select one or more)",
                _all_species, key=k("gas_species_widget"),
                help="Common process gases, hydrocarbons and refrigerants. "
                     "All use the ideal-gas law with CoolProp viscosity.")
            if not _sel:
                st.info("Select at least one gas species.")
            _flows = {}
            if _sel:
                _nc = min(len(_sel), 3)
                _fc = st.columns(_nc)
                for _ci, _sp in enumerate(_sel):
                    _fk = k(f"gflow_{_sp}")
                    if _fk not in st.session_state:
                        st.session_state[_fk] = 100.0
                    _flows[_sp] = _fc[_ci % _nc].number_input(
                        f"{_sp}  (kg/h)", min_value=0.0, step=0.1, key=_fk)
            _cgas = None
            if "Custom" in _sel:
                st.markdown("*Custom gas properties*")
                _cg1, _cg2 = st.columns(2)
                _cg_mw = _cg1.number_input("MW (g/mol)", min_value=1.0, value=28.0,
                                            step=1.0, key=k("cg_mw"))
                _cg_mu = _cg2.number_input("μ (µPa·s)", min_value=1.0, value=18.5,
                                            step=0.5, key=k("cg_mu"))
                _cgas = {"MW_gmol": _cg_mw, "mu_upas": _cg_mu}
            _use_cp = any(engine.GAS_SPECIES.get(sp, {}).get("coolprop_id")
                          for sp in _sel if sp != "Custom")
            return _flows, _cgas, _use_cp

        # ── Helper: render Liquid Phase container ────────────────────────────
        def _render_liquid_inputs(T_C_val, P_bara_val):
            _coolprop_opts = list(engine.LIQUID_COOLPROP_ID.keys())
            _all_liq_opts  = _coolprop_opts + ["KOH solution"]
            _lsp = st.multiselect(
                "Liquid species  (select one or more)",
                _all_liq_opts, key=k("liquid_species_widget"),
                help="CoolProp-backed species can be freely mixed. "
                     "KOH solution uses built-in empirical correlations "
                     "(density ±1 %, viscosity ±15 %) and must be selected alone.")
            _lflows = {}
            if _lsp:
                _lnc = min(len(_lsp), 3)
                _lfc = st.columns(_lnc)
                for _li, _ls in enumerate(_lsp):
                    _lfk = k(f"lflow_{_ls}")
                    if _lfk not in st.session_state:
                        st.session_state[_lfk] = 1000.0
                    _lf = _lfc[_li % _lnc].number_input(
                        f"{_ls}  (kg/h)", min_value=0.0, step=100.0, key=_lfk)
                    if _lf > 0:
                        _lflows[_ls] = _lf

            _liq_type = "Custom"; _q = 0.0; _cliq = None
            _has_koh  = "KOH solution" in _lflows
            _has_cp   = any(s in engine.LIQUID_COOLPROP_ID for s in _lflows)

            if _has_koh and _has_cp:
                st.warning("KOH solution cannot be mixed with other liquid species in this version. "
                           "Remove the other species or use KOH alone.")
                return {}, "Custom", 0.0, None

            if _has_koh:
                # ── KOH built-in path ─────────────────────────────────────────
                _koh_conc = st.slider(
                    "KOH concentration (wt%)", min_value=20, max_value=40,
                    value=int(st.session_state.get(k("koh_conc"), 30)),
                    step=1, key=k("koh_conc"),
                    help="Mass fraction of KOH in water. Typical alkaline electrolyser: 25–32 wt%.")
                _T_koh = T_C_val if T_C_val is not None else 25.0
                _koh_rho, _koh_mu, _koh_sig = engine.koh_properties(_T_koh, _koh_conc)
                _lm1, _lm2, _lm3 = st.columns(3)
                _lm1.metric("ρ", f"{_koh_rho:.1f} kg/m³")
                _lm2.metric("μ", f"{_koh_mu*1e3:.3f} mPa·s")
                _lm3.metric("σ", f"{_koh_sig*1e3:.2f} mN/m")
                _cliq = {"rho_kgm3": _koh_rho,
                         "mu_mpas":  _koh_mu * 1e3,
                         "sigma_mnm": _koh_sig * 1e3,
                         "koh_conc_wt": _koh_conc}
                _q    = _lflows["KOH solution"] / _koh_rho if _koh_rho > 0 else 0.0
                _liq_type = "Custom"

            elif _lflows:
                # ── CoolProp mixture path ─────────────────────────────────────
                try:
                    _T_K  = (T_C_val + 273.15) if T_C_val is not None else 298.15
                    _rl, _mul, _sigl = engine.liquid_mixture_props(
                        _lflows, _T_K, P_bara_val * 1e5)
                    _lm1, _lm2, _lm3 = st.columns(3)
                    _lm1.metric("ρ_mix", f"{_rl:.1f} kg/m³")
                    _lm2.metric("μ_mix", f"{_mul*1e3:.3f} mPa·s")
                    _lm3.metric("σ_mix", f"{_sigl*1e3:.2f} mN/m")
                    _q = sum(_lflows.values()) / _rl
                    if len(_lflows) == 1:
                        _liq_type = next(iter(_lflows)); _cliq = None
                    else:
                        _liq_type = "Custom"
                        _cliq = {"rho_kgm3": _rl, "mu_mpas": _mul * 1e3,
                                 "sigma_mnm": _sigl * 1e3}
                except Exception:
                    pass

            return _lflows, _liq_type, _q, _cliq

        # ── Mode-specific inputs ──────────────────────────────────────────────
        if flow_mode == "vle":
            with st.container(border=True):
                st.markdown("**Saturated Fluid (VLE)**")
                _vle_display_list = list(engine.VLE_FLUID_DISPLAY.keys())
                _vle_names_to_id  = engine.VLE_FLUID_DISPLAY
                _vle_saved_fluid  = st.session_state.get(k("vle_fluid_widget"), "Water")
                _vle_display_saved = next(
                    (dn for dn, fid in _vle_names_to_id.items() if fid == _vle_saved_fluid),
                    _vle_display_list[0])
                _vle_sel_idx = (_vle_display_list.index(_vle_display_saved)
                                if _vle_display_saved in _vle_display_list else 0)
                vle_display_name = st.selectbox(
                    "Fluid", _vle_display_list, index=_vle_sel_idx, key=k("vle_fluid_sel"),
                    help="Single pure component on its own saturation curve. "
                         "T_sat and phase densities are derived from pressure at each segment. "
                         "Quality x evolves along the pipe via isenthalpic flash — "
                         "flashing is computed automatically as pressure drops.")
                vle_fluid_id = _vle_names_to_id[vle_display_name]
                st.session_state[k("vle_fluid_widget")] = vle_fluid_id
                _vv1, _vv2 = st.columns(2)
                vle_m_kgs = _vv1.number_input(
                    "Total mass flow (kg/s)", min_value=0.001, max_value=500.0,
                    step=0.1, format="%.3f", key=k("vle_m_kgs_widget"))
                vle_x = _vv2.slider(
                    "Inlet quality  x  (0 = sat. liquid, 1 = sat. vapour)",
                    min_value=0.0, max_value=1.0, step=0.01, key=k("vle_x_widget"),
                    help="Mass fraction vapour at pipe inlet. Downstream quality is "
                         "calculated automatically via isenthalpic flash.")
                try:
                    _vle_prev = engine.calculate_vle_properties(vle_fluid_id, P_bara, vle_x, vle_m_kgs)
                    _vc1, _vc2, _vc3, _vc4 = st.columns(4)
                    _vc1.metric("T_sat", f"{_vle_prev['T_sat_C']:.1f} °C")
                    _vc2.metric("ρ_liq", f"{_vle_prev['rho_l']:.1f} kg/m³")
                    _vc3.metric("ρ_vap", f"{_vle_prev['rho_g']:.3f} kg/m³")
                    _vc4.metric("σ", f"{_vle_prev['sigma']*1e3:.2f} mN/m")
                except Exception as _vle_err:
                    st.error(f"CoolProp VLE error: {_vle_err}")
                    vle_fluid_id = "Water"; vle_m_kgs = 1.0; vle_x = 0.5
            # Stubs for unused variables
            gas_flows_kgh = {}; liquid_type = f"{vle_fluid_id} (VLE)"; q_lye = 0.0
            liquid_flows_kgh = None; custom_gas = None; custom_liquid = None
            use_coolprop = False; T_C = None

        elif flow_mode == "gas_only":
            vle_fluid_id = None; vle_m_kgs = None; vle_x = None
            with st.container(border=True):
                st.markdown("**Gas Phase**")
                st.caption(
                    "Single-phase gas — Darcy-Weisbach with ideal-gas density. "
                    "Density is updated at each segment inlet (isothermal compressible). "
                    "Choking / Fanno-flow is not modelled.")
                gas_flows_kgh, custom_gas, use_coolprop = _render_gas_inputs()
            # No liquid
            liquid_flows_kgh = {}; q_lye = 0.0; liquid_type = "Custom"; custom_liquid = None

        elif flow_mode == "liquid_only":
            vle_fluid_id = None; vle_m_kgs = None; vle_x = None
            with st.container(border=True):
                st.markdown("**Liquid Phase**")
                st.caption("Single-phase liquid — incompressible Darcy-Weisbach with CoolProp properties.")
                liquid_flows_kgh, liquid_type, q_lye, custom_liquid = _render_liquid_inputs(T_C, P_bara)
            # No gas
            gas_flows_kgh = {}; custom_gas = None; use_coolprop = False

        else:  # gas_liquid (two-phase)
            vle_fluid_id = None; vle_m_kgs = None; vle_x = None
            with st.container(border=True):
                st.markdown("**Gas Phase**")
                gas_flows_kgh, custom_gas, use_coolprop = _render_gas_inputs()
            with st.container(border=True):
                st.markdown("**Liquid Phase**")
                liquid_flows_kgh, liquid_type, q_lye, custom_liquid = _render_liquid_inputs(T_C, P_bara)

        # Pipe Geometry
        with st.container(border=True):
            st.markdown("**Pipe Geometry**")
            current_specs = []
            DN_OPTIONS    = list(engine.PIPE_DATABASE.keys())
            PN_OPTIONS    = ["PN20", "PN25", "PN40"]
            MAT_OPTIONS   = list(engine.MATERIAL_ROUGHNESS.keys())
            LINER_OPTIONS = list(engine.LINER_ROUGHNESS.keys())
            for i, seg in enumerate(st.session_state[k("segments")]):
                _seg_kind = seg.get("kind", "pipe")
                _kind_labels = {"pipe": "Pipe segment", "valve": "Valve", "hx": "Heat exchanger"}
                _seg_hdr_col, _seg_del_col = st.columns([9, 1])
                _seg_hdr_col.markdown(f"**#{i+1} — {_kind_labels.get(_seg_kind, 'Pipe segment')}**")
                if len(st.session_state[k("segments")]) > 1:
                    if _seg_del_col.button("×", key=k(f"del_seg_{i}"), help="Remove this segment"):
                        st.session_state[k("segments")].pop(i)
                        st.rerun()

                # ── VALVE ──────────────────────────────────────────────────────
                if _seg_kind == "valve":
                    _vm1, _vm2, _vm3, _vm4 = st.columns([1.2, 1.0, 1.0, 1.0])
                    _v_mode = _vm1.radio(
                        "Mode", ["Specify ΔP → Kv", "Specify Kv → ΔP"],
                        key=k(f"v_mode_{i}"),
                        index=0 if seg.get("valve_mode", "dp") == "dp" else 1,
                        horizontal=False,
                        help="ΔP→Kv: budget a pressure drop, get required Kv (line sizing).\n"
                             "Kv→ΔP: enter a known Kv, get the pressure drop.")
                    _v_dn = _vm2.selectbox("Nominal DN", DN_OPTIONS, key=k(f"v_dn_{i}"),
                                           index=DN_OPTIONS.index(seg.get("dn", "DN50")))
                    _v_pn = _vm3.selectbox("PN", PN_OPTIONS, key=k(f"v_pn_{i}"),
                                           index=PN_OPTIONS.index(seg.get("pn", "PN40")))
                    _v_char = _vm4.selectbox(
                        "Characteristic", ["equal-percentage", "linear"],
                        key=k(f"v_char_{i}"),
                        index=["equal-percentage", "linear"].index(
                            seg.get("characteristic", "equal-percentage")),
                        help="Equal-percentage (R=50): Kv_eff = Kv_rated × 50^(f−1).  "
                             "Linear: Kv_eff = Kv_rated × f.")

                    _v_mode_key = "dp" if _v_mode.startswith("Specify ΔP") else "kv"

                    if _v_mode_key == "dp":
                        _vc1, _vc2 = st.columns([1.2, 1.0])
                        _v_dp_kpa = _vc1.number_input(
                            "Target ΔP  (kPa)", min_value=0.1, max_value=10000.0,
                            value=float(seg.get("dp_kpa", 50.0)), step=5.0, format="%.1f",
                            key=k(f"v_dp_{i}"),
                            help="Pressure drop budget allocated to this valve.")
                        _v_open = _vc2.slider(
                            "Opening (%)", min_value=1, max_value=100,
                            value=int(seg.get("opening_pct", 100)),
                            key=k(f"v_open_{i}"),
                            help="Valve opening at operating point.")
                        _v_kv = seg.get("Kv_m3h", 10.0)  # kept for storage, computed at runtime
                        st.caption(f"Required Kv calculated from ΔP = {_v_dp_kpa:.1f} kPa "
                                   f"at {_v_open} % opening — shown in results.")
                    else:
                        _vc1, _vc2 = st.columns([1.2, 1.0])
                        _v_kv = _vc1.number_input(
                            "Kv rated  (m³/h·bar^0.5)", min_value=0.001, max_value=50000.0,
                            value=float(seg.get("Kv_m3h", 10.0)), step=1.0, format="%.3f",
                            key=k(f"v_kv_{i}"),
                            help="Valve flow coefficient at full-open. "
                                 "Kv = Cv / 1.156")
                        _v_open = _vc2.slider(
                            "Opening (%)", min_value=1, max_value=100,
                            value=int(seg.get("opening_pct", 100)),
                            key=k(f"v_open_{i}"),
                            help="Valve opening position.")
                        _f_open = max(0.001, _v_open / 100.0)
                        _kv_eff_p = (_v_kv * (50.0 ** (_f_open - 1.0))
                                     if _v_char == "equal-percentage"
                                     else _v_kv * _f_open)
                        _v_dp_kpa = seg.get("dp_kpa", 50.0)  # kept for storage
                        st.caption(
                            f"Kv_eff = **{_kv_eff_p:.3f}**  ·  "
                            f"Cv_rated ≈ {_v_kv * 1.156:.2f}  ·  "
                            f"Cv_eff ≈ {_kv_eff_p * 1.156:.2f}")

                    current_specs.append({
                        "kind": "valve", "dn": _v_dn, "pn": _v_pn,
                        "valve_mode": _v_mode_key,
                        "Kv_m3h": float(_v_kv), "dp_kpa": float(_v_dp_kpa),
                        "opening_pct": float(_v_open),
                        "characteristic": _v_char,
                    })
                    continue

                # ── HEAT EXCHANGER ─────────────────────────────────────────────
                elif _seg_kind == "hx":
                    _hg1, _hg2, _hg3 = st.columns([1.0, 1.0, 1.0])
                    _h_dn = _hg1.selectbox("Nominal DN", DN_OPTIONS, key=k(f"hx_dn_{i}"),
                                           index=DN_OPTIONS.index(seg.get("dn", "DN50")))
                    _h_pn = _hg2.selectbox("PN", PN_OPTIONS, key=k(f"hx_pn_{i}"),
                                           index=PN_OPTIONS.index(seg.get("pn", "PN40")))
                    _h_dir = _hg3.selectbox(
                        "Flow direction", ["Horizontal", "Vertical Upflow", "Vertical Downflow"],
                        key=k(f"hx_dir_{i}"),
                        index=["Horizontal", "Vertical Upflow", "Vertical Downflow"].index(
                            seg.get("type", "Horizontal")))
                    _hg4, _hg5 = st.columns(2)
                    _h_duty = _hg4.number_input(
                        "Heat duty  (kW)", min_value=-100000.0, max_value=100000.0,
                        value=float(seg.get("duty_kw", 0.0)), step=10.0, format="%.1f",
                        key=k(f"hx_duty_{i}"),
                        help="Heat added to the process stream. "
                             "Positive = heating (temperature rises). "
                             "Negative = cooling (temperature falls). "
                             "Set to 0 if you only want the pressure drop.")
                    _h_dp = _hg5.number_input(
                        "Pressure drop  (kPa)", min_value=0.0, max_value=10000.0,
                        value=float(seg.get("dp_kpa", 20.0)), step=1.0, format="%.1f",
                        key=k(f"hx_dp_{i}"),
                        help="Shell- or tube-side ΔP from the equipment datasheet. "
                             "Enter the ΔP for the stream flowing through this pipeline.")
                    if _h_duty > 0:
                        st.caption(f"Heating duty: +{_h_duty:.1f} kW  ·  ΔP = {_h_dp:.1f} kPa")
                    elif _h_duty < 0:
                        st.caption(f"Cooling duty: {_h_duty:.1f} kW  ·  ΔP = {_h_dp:.1f} kPa")
                    else:
                        st.caption(f"No heat exchange  ·  ΔP = {_h_dp:.1f} kPa")
                    current_specs.append({
                        "kind": "hx", "dn": _h_dn, "pn": _h_pn,
                        "type": _h_dir,
                        "duty_kw": _h_duty, "dp_kpa": _h_dp,
                    })
                    continue

                # ── PIPE SEGMENT ───────────────────────────────────────────────
                g1, g2, g3, g4 = st.columns([1.3, 0.8, 0.7, 0.7])
                t = g1.selectbox("Flow direction",
                    ["Horizontal", "Vertical Upflow", "Vertical Downflow"],
                    key=k(f"t_{i}"),
                    index=["Horizontal","Vertical Upflow","Vertical Downflow"].index(seg["type"]))
                dn = g2.selectbox("DN", DN_OPTIONS, key=k(f"dn_{i}"),
                                  index=DN_OPTIONS.index(seg.get("dn","DN50")),
                                  help="Nominal pipe diameter (ANSI B36.10/19). "
                                       "Internal bore depends on DN + PN.")
                pn = g3.selectbox("PN", PN_OPTIONS, key=k(f"pn_{i}"),
                                  index=PN_OPTIONS.index(seg.get("pn","PN40")),
                                  help="Pressure rating class. Determines pipe schedule "
                                       "and wall thickness, setting the internal bore.")
                l  = g4.number_input("Length (m)", min_value=0.1,
                                     value=float(seg["length"]), step=1.0, key=k(f"l_{i}"))

                g5, g8 = st.columns([3.5, 0.65])
                _mat_def = seg.get("material","SS316L")
                mat = g5.selectbox("Material", MAT_OPTIONS, key=k(f"m_{i}"),
                                   index=MAT_OPTIONS.index(_mat_def) if _mat_def in MAT_OPTIONS else 0)
                g8.markdown("<div style='height:1.85rem'></div>", unsafe_allow_html=True)
                lined = g8.checkbox("Lined", value=bool(seg.get("lined",False)), key=k(f"lined_{i}"),
                                    help="Fluoropolymer liner (PTFE/FEP/PFA/PVDF) reduces effective "
                                         "bore and overrides wall roughness.")

                _lmat    = seg.get("liner_material","FEP")
                _lthk_mm = float(seg.get("liner_thickness_mm",1.0))
                if lined:
                    g9, g10 = st.columns([1.6,1.0])
                    _lmat = g9.selectbox("Liner Material", LINER_OPTIONS, key=k(f"lmat_{i}"),
                        index=LINER_OPTIONS.index(_lmat) if _lmat in LINER_OPTIONS else 0)
                    _lthk_mm = g10.number_input("Liner Thickness (mm)", min_value=0.1,
                        max_value=20.0, value=_lthk_mm, step=0.5, key=k(f"lthk_{i}"))

                D_seg = engine.PIPE_DATABASE[dn][pn]
                D_eff = D_seg - 2*(_lthk_mm/1000.0) if lined else D_seg
                rough = engine.LINER_ROUGHNESS[_lmat] if lined else engine.MATERIAL_ROUGHNESS[mat]
                if lined:
                    st.caption(f"Bore {D_seg*1000:.1f} mm → **ID {D_eff*1000:.1f} mm**"
                               f"  ·  ε {rough*1e6:.3g} µm  ·  {mat} + {_lmat} {_lthk_mm:.1f} mm")
                else:
                    st.caption(f"ID {D_seg*1000:.1f} mm  ·  ε {rough*1e6:.2g} µm  ·  {mat}")

                # ── Multi-row fittings ───────────────────────────────────────────
                _fk = k(f"fits_{i}")
                if _fk not in st.session_state:
                    _init_fl = seg.get("fittings_list")
                    if _init_fl is None:
                        _old_f = seg.get("fittings", "None")
                        _old_c = seg.get("fitting_count", 0)
                        _init_fl = ([{"type": _old_f, "qty": int(_old_c)}]
                                    if _old_f in engine.FITTING_Le_over_D and _old_c > 0
                                    else [])
                    st.session_state[_fk] = list(_init_fl)

                _fit_rows = st.session_state[_fk]
                _n_fits   = len(_fit_rows)
                _all_fit_types = list(engine.FITTING_Le_over_D.keys())

                _fl_col, _fa_col = st.columns([4, 1])
                _fl_col.caption("**Minor losses**" if _n_fits else "Minor losses — none")
                if _fa_col.button("+ Add", key=k(f"fadd_{i}"), help="Add a fitting type"):
                    for jj in range(_n_fits):
                        _ft = st.session_state.get(k(f"ftype_{i}_{jj}"))
                        _fq = st.session_state.get(k(f"fqty_{i}_{jj}"), 1)
                        if _ft:
                            st.session_state[_fk][jj]["type"] = _ft
                            st.session_state[_fk][jj]["qty"]  = int(_fq)
                    st.session_state[_fk].append({"type": _all_fit_types[0], "qty": 1})
                    st.rerun()

                for j, _fr in enumerate(list(_fit_rows)):
                    _fc1, _fc2, _fc3 = st.columns([3.5, 0.7, 0.3])
                    _ftype_key = k(f"ftype_{i}_{j}")
                    _fqty_key  = k(f"fqty_{i}_{j}")
                    _cur_type  = _fit_rows[j]["type"]
                    _cur_qty   = _fit_rows[j]["qty"]
                    _fc1.selectbox("Fitting", _all_fit_types,
                        index=_all_fit_types.index(_cur_type) if _cur_type in _all_fit_types else 0,
                        key=_ftype_key, label_visibility="collapsed")
                    _fc2.number_input("Qty", min_value=1, value=_cur_qty,
                        key=_fqty_key, label_visibility="collapsed")
                    if _fc3.button("×", key=k(f"frem_{i}_{j}"), help="Remove this fitting"):
                        for jj in range(_n_fits):
                            _ft = st.session_state.get(k(f"ftype_{i}_{jj}"))
                            _fq = st.session_state.get(k(f"fqty_{i}_{jj}"), 1)
                            if _ft:
                                st.session_state[_fk][jj]["type"] = _ft
                                st.session_state[_fk][jj]["qty"]  = int(_fq)
                        st.session_state[_fk].pop(j)
                        for jj in range(j, _n_fits - 1):
                            st.session_state[k(f"ftype_{i}_{jj}")] = st.session_state[_fk][jj]["type"]
                            st.session_state[k(f"fqty_{i}_{jj}")] = st.session_state[_fk][jj]["qty"]
                        st.session_state.pop(k(f"ftype_{i}_{_n_fits-1}"), None)
                        st.session_state.pop(k(f"fqty_{i}_{_n_fits-1}"), None)
                        st.rerun()

                _fittings_list = []
                for jj in range(_n_fits):
                    _ft = st.session_state.get(k(f"ftype_{i}_{jj}"))
                    _fq = st.session_state.get(k(f"fqty_{i}_{jj}"), 1)
                    if _ft and _ft in engine.FITTING_Le_over_D and int(_fq) > 0:
                        _fittings_list.append({"type": _ft, "qty": int(_fq)})

                current_specs.append({
                    "type": t, "dn": dn, "pn": pn, "material": mat, "length": l,
                    "fittings_list": _fittings_list,
                    "lined": lined, "liner_material": _lmat, "liner_thickness_mm": _lthk_mm,
                })

            st.session_state[k("segments")] = current_specs
            _ab1, _ab2, _ab3, _ab4 = st.columns(4)
            if _ab1.button("+ Pipe segment", key=k("add_seg"), width='stretch'):
                _last = st.session_state[k("segments")][-1]
                _last_pipe = next(
                    (s for s in reversed(st.session_state[k("segments")]) if s.get("kind","pipe") == "pipe"),
                    {"dn": "DN50", "pn": "PN40", "material": "SS316L", "lined": False,
                     "liner_material": "FEP", "liner_thickness_mm": 1.0})
                st.session_state[k("segments")].append({
                    "kind": "pipe", "type": "Horizontal",
                    "dn": _last_pipe.get("dn", "DN50"), "pn": _last_pipe.get("pn", "PN40"),
                    "material": _last_pipe.get("material", "SS316L"),
                    "length": 5.0, "fittings_list": [],
                    "lined": _last_pipe.get("lined", False),
                    "liner_material": _last_pipe.get("liner_material", "FEP"),
                    "liner_thickness_mm": _last_pipe.get("liner_thickness_mm", 1.0),
                })
                st.rerun()
            if _ab2.button("+ Valve", key=k("add_valve"), width='stretch'):
                _last_dn = next(
                    (s.get("dn","DN50") for s in reversed(st.session_state[k("segments")])), "DN50")
                _last_pn = next(
                    (s.get("pn","PN40") for s in reversed(st.session_state[k("segments")])), "PN40")
                st.session_state[k("segments")].append({
                    "kind": "valve", "dn": _last_dn, "pn": _last_pn,
                    "Kv_m3h": 10.0, "opening_pct": 100.0,
                    "characteristic": "equal-percentage",
                })
                st.rerun()
            if _ab3.button("+ Heat exchanger", key=k("add_hx"), width='stretch'):
                _last_dn = next(
                    (s.get("dn","DN50") for s in reversed(st.session_state[k("segments")])), "DN50")
                _last_pn = next(
                    (s.get("pn","PN40") for s in reversed(st.session_state[k("segments")])), "PN40")
                st.session_state[k("segments")].append({
                    "kind": "hx", "dn": _last_dn, "pn": _last_pn,
                    "type": "Horizontal", "duty_kw": 0.0, "dp_kpa": 20.0,
                })
                st.rerun()
            if _ab4.button("− Remove last", key=k("rem_seg"), width='stretch') and \
               len(st.session_state[k("segments")]) > 1:
                st.session_state[k("segments")].pop()
                st.rerun()

        # Calculation Settings — two-phase modes only
        if flow_mode in ("gas_liquid", "vle"):
            with st.container(border=True):
                st.markdown("**Calculation Settings**")
                _cs1, _cs2 = st.columns(2)
                correlation = _cs1.selectbox(
                    "ΔP correlation", engine.TWO_PHASE_CORRELATIONS,
                    key=k("correlation"),
                    help="Two-phase frictional pressure drop correlation. "
                         "Beggs-Brill also accounts for pipe inclination; "
                         "the others add gravity as a separate term.")
                voidage_method = _cs2.selectbox(
                    "Void fraction model", engine.VOIDAGE_METHODS,
                    key=k("voidage_method"),
                    help="Homogeneous: α from the density ratio (fast, conservative). "
                         "Rouhani-1: slip-flow model, better for stratified or annular flow.")
        else:
            # Single-phase: read last-used values from session state (not shown as widgets)
            correlation    = st.session_state.get(k("correlation"),    engine.TWO_PHASE_CORRELATIONS[0])
            voidage_method = st.session_state.get(k("voidage_method"), engine.VOIDAGE_METHODS[0])

    # ── OUTPUTS ───────────────────────────────────────────────────────────────
    with col_out:
        _key_results_ph = st.empty()   # filled after segment loop — ΔP summary

        # Initialise effective liquid state (overwritten in gas+liquid path below)
        _eff_liquid_flows = liquid_flows_kgh
        _eff_gas_flows    = gas_flows_kgh
        _eff_liq_type     = liquid_type
        _eff_q_lye        = q_lye
        _eff_custom_liq   = custom_liquid

        if _is_vle:
            try:
                # Compute inlet enthalpy once — carried through the segment loop
                # so quality evolves correctly as pressure drops (isenthalpic flash).
                _vle_h_inlet = engine.vle_inlet_enthalpy(vle_fluid_id, P_bara, vle_x)
                props = engine.calculate_vle_properties(
                    vle_fluid_id, P_bara, vle_x, vle_m_kgs, h_spec=_vle_h_inlet)
                T_C = props["T_sat_C"]
            except Exception as _vle_calc_err:
                st.error(f"VLE calculation failed: {_vle_calc_err}")
                st.stop()
        else:
            is_valid, warn_list = engine.validate_input_bounds(
                P_bara, T_C, gas_flows_kgh, liquid_type, q_lye,
                liquid_flows_kgh=liquid_flows_kgh)
            for w in warn_list:
                st.warning(w)

            # ── Equilibrium Flash (Peng-Robinson EOS) ────────────────────────
            _flash = engine.flash_pt(gas_flows_kgh, liquid_type, q_lye, T_C, P_bara,
                                     liquid_flows_kgh=liquid_flows_kgh)

            if _flash["feasible"]:
                _vf_pct = _flash["VF_mass"] * 100
                if _vf_pct >= 1.0:
                    st.info(
                        f"Equilibrium flash at {P_bara:.1f} bara / {T_C:.0f} °C — "
                        f"vapour fraction {_vf_pct:.1f} wt%. "
                        "Adjusted compositions used for pressure-drop calculation.",
                        icon="ℹ️",
                    )
                else:
                    st.caption(
                        f"Flash: VF = {_vf_pct:.2f} wt% — effectively single-phase liquid "
                        f"at {P_bara:.1f} bara / {T_C:.0f} °C."
                    )
                with st.expander(
                    f"Feed / phase-split detail  (VF = {_flash['VF_mass']*100:.1f} wt%)",
                    expanded=False):
                    _feed   = _flash["feed_kgh"]
                    _g_ph   = _flash["gas_phase_kgh"]
                    _l_ph   = _flash["liquid_phase_kgh"]
                    _frows  = []
                    for _sp in _flash["species"]:
                        _frows.append({
                            "Species":           _sp,
                            "Feed (kg/h)":       round(_feed.get(_sp, 0), 4),
                            "Gas phase (kg/h)":  round(_g_ph.get(_sp, 0), 4),
                            "Liquid phase (kg/h)": round(_l_ph.get(_sp, 0), 4),
                        })
                    st.dataframe(
                        pd.DataFrame(_frows),
                        column_config={c: st.column_config.NumberColumn(format="%.4f")
                                       for c in ["Feed (kg/h)", "Gas phase (kg/h)",
                                                 "Liquid phase (kg/h)"]},
                        hide_index=True, width='stretch')
                    _fv1, _fv2 = st.columns(2)
                    _fv1.metric("Vapour fraction (mass)", f"{_flash['VF_mass']*100:.2f} %")
                    _fv2.metric("Vapour fraction (mol)",  f"{_flash['VF_mol']*100:.2f} %")
                    st.caption("Peng-Robinson EOS · binary interaction parameters = 0 "
                               "· composition assumed constant along pipe")

                # Derive effective inputs for pressure-drop engine
                _fl_gas = _flash["gas_phase_kgh"]
                _fl_liq = _flash["liquid_phase_kgh"]
                if _fl_gas and _fl_liq:
                    _fl_rho, _fl_mu, _fl_sig = engine.liquid_mixture_props(
                        _fl_liq, T_C + 273.15, P_bara * 1e5)
                    _eff_gas_flows    = _fl_gas
                    _eff_liq_type     = "Custom"
                    _eff_q_lye        = sum(_fl_liq.values()) / _fl_rho
                    _eff_custom_liq   = {"rho_kgm3": _fl_rho,
                                         "mu_mpas":  _fl_mu * 1e3,
                                         "sigma_mnm": _fl_sig * 1e3}
                    _eff_liquid_flows = None  # encoded in Custom above
                elif not _fl_liq:
                    _eff_gas_flows    = _fl_gas or gas_flows_kgh
                    _eff_liq_type     = liquid_type
                    _eff_q_lye        = 0.0
                    _eff_custom_liq   = custom_liquid
                    _eff_liquid_flows = liquid_flows_kgh
                else:
                    _eff_gas_flows    = gas_flows_kgh
                    _eff_liq_type     = liquid_type
                    _eff_q_lye        = q_lye
                    _eff_custom_liq   = custom_liquid
                    _eff_liquid_flows = liquid_flows_kgh
            else:
                if _flash["warnings"]:
                    st.caption(f"Flash equilibrium: {_flash['warnings'][0]}")
                _eff_gas_flows    = gas_flows_kgh
                _eff_liq_type     = liquid_type
                _eff_q_lye        = q_lye
                _eff_custom_liq   = custom_liquid
                _eff_liquid_flows = liquid_flows_kgh

            try:
                props = engine.calculate_two_phase_properties(
                    P_bara, T_C, _eff_gas_flows, _eff_liq_type, _eff_q_lye,
                    custom_gas=custom_gas, custom_liquid=_eff_custom_liq,
                    use_coolprop=use_coolprop,
                    liquid_flows_kgh=_eff_liquid_flows)
            except Exception as _calc_err:
                st.error(f"Property calculation failed: {_calc_err}")
                st.stop()

        # ── Calculate button & segment-loop recompute guard ──────────────────
        _loop_hash_src = {
            "P": P_bara, "T": T_C,
            "eff_gas": {_gk: float(_gv) for _gk, _gv in (_eff_gas_flows or {}).items()},
            "eff_liq": _eff_liq_type, "eff_q": float(_eff_q_lye or 0),
            "eff_cl": json.dumps(_eff_custom_liq, sort_keys=True) if _eff_custom_liq else None,
            "eff_lf": {_gk: float(_gv) for _gk, _gv in (_eff_liquid_flows or {}).items()} if _eff_liquid_flows else None,
            "corr": correlation, "void": voidage_method,
            "is_vle": _is_vle,
            "vle_f": vle_fluid_id if _is_vle else None,
            "vle_x": float(vle_x) if vle_x is not None else None,
            "vle_m": float(vle_m_kgs) if vle_m_kgs is not None else None,
            "cgas": json.dumps(custom_gas, sort_keys=True) if custom_gas else None,
            "ucp": use_coolprop,
            "segs": [(s.get("kind","pipe"), s.get("dn",""), s.get("pn",""),
                      s.get("material",""), float(s.get("length",0)),
                      s.get("type","Horizontal"), bool(s.get("lined",False)),
                      s.get("liner_material","FEP"), float(s.get("liner_thickness_mm",1)),
                      str(sorted((s.get("fittings_list") or []))),
                      s.get("valve_mode","kv"), float(s.get("dp_kpa",50)),
                      float(s.get("Kv_m3h",10)), s.get("characteristic","linear"),
                      float(s.get("opening_pct",100)), float(s.get("hx_dp_kpa",0)))
                     for s in st.session_state[k("segments")]],
        }
        _loop_hash  = hashlib.md5(
            json.dumps(_loop_hash_src, sort_keys=True, default=str).encode()
        ).hexdigest()
        _lc         = st.session_state.get(k("loop_cache"), {})
        _run_calc   = _lc.get("hash") != _loop_hash
        # ─────────────────────────────────────────────────────────────────────

        if _run_calc:
            _calc = compute_pipeline_case(
                P_bara=P_bara,
                T_C=T_C,
                eff_gas_flows=_eff_gas_flows,
                eff_liq_type=_eff_liq_type,
                eff_q_lye=_eff_q_lye,
                eff_custom_liq=_eff_custom_liq,
                eff_liquid_flows=_eff_liquid_flows,
                is_vle=_is_vle,
                vle_fluid_id=vle_fluid_id,
                vle_x=vle_x,
                vle_m_kgs=vle_m_kgs,
                vle_h_inlet=_vle_h_inlet if _is_vle else None,
                segments=st.session_state[k("segments")],
                correlation=correlation,
                voidage_method=voidage_method,
                custom_gas=custom_gas,
                use_coolprop=use_coolprop,
                props=props,
            )
            current_P            = _calc["current_P"]
            current_T_C          = _calc["current_T_C"]
            vle_x                = _calc["vle_x"]
            grid_records         = _calc["grid_records"]
            stream_records       = _calc["stream_records"]
            valve_sizing         = _calc["valve_sizing"]
            cumulative_distance  = _calc["cumulative_distance"]
            cumulative_positions = _calc["cumulative_positions"]
            pressure_profile_x   = _calc["pressure_profile_x"]
            pressure_profile_y   = _calc["pressure_profile_y"]
            regime_bands         = _calc["regime_bands"]
            total_dp_fric_kpa    = _calc["total_dp_fric_kpa"]
            total_dp_grav_kpa    = _calc["total_dp_grav_kpa"]
            total_dp_accel_kpa   = _calc["total_dp_accel_kpa"]
            st.session_state[k("loop_cache")] = {"hash": _loop_hash, **_calc}
        elif _lc:
            current_P            = _lc["current_P"]
            current_T_C          = _lc["current_T_C"]
            vle_x                = _lc["vle_x"]
            grid_records         = _lc["grid_records"]
            stream_records       = _lc["stream_records"]
            valve_sizing         = _lc["valve_sizing"]
            cumulative_distance  = _lc["cumulative_distance"]
            cumulative_positions = _lc["cumulative_positions"]
            pressure_profile_x   = _lc["pressure_profile_x"]
            pressure_profile_y   = _lc["pressure_profile_y"]
            regime_bands         = _lc["regime_bands"]
            total_dp_fric_kpa    = _lc["total_dp_fric_kpa"]
            total_dp_grav_kpa    = _lc["total_dp_grav_kpa"]
            total_dp_accel_kpa   = _lc["total_dp_accel_kpa"]
        # ─────────────────────────────────────────────────────────────────────

        # Erosion — only banner on actionable conditions; OK is shown inline in table
        _max_ratio = max((r["V_m/V_e"] for r in grid_records), default=0.0)
        _worst_seg = next((r["Seg"] for r in grid_records if r["V_m/V_e"]==_max_ratio), "")
        if _max_ratio >= 1.0:
            st.error(f"**Erosion limit exceeded** — Segment {_worst_seg}: "
                     f"V_m/V_e = **{_max_ratio:.2f}** (API RP 14E, C=100). "
                     f"Reduce velocity or increase pipe diameter.")
        elif _max_ratio >= 0.8:
            st.warning(f"**Approaching erosion limit** — Segment {_worst_seg}: "
                       f"V_m/V_e = **{_max_ratio:.2f}** (API RP 14E, limit = 1.0).")

        # ── Valve Sizing Results ──────────────────────────────────────────────
        if valve_sizing:
            with st.container(border=True):
                st.markdown("**Valve Sizing**")
                for _vs in valve_sizing:
                    st.markdown(f"Segment #{_vs['seg']} — {_vs['dn']}/{_vs['pn']}  "
                                f"·  ΔP = {_vs['dp_kpa']:.1f} kPa  "
                                f"·  {_vs['opening']:.0f} % open  "
                                f"·  {_vs['char']}")
                    _vc1, _vc2, _vc3, _vc4, _vc5 = st.columns(5)
                    _vc1.metric("Q",       f"{_vs['Q_m3h']:.3f} m³/h")
                    _vc2.metric("ρ_hom",   f"{_vs['rho_hom']:.1f} kg/m³")
                    _vc3.metric("Kv_eff",  f"{_vs['Kv_eff']:.2f} m³/h·bar^0.5")
                    _vc4.metric("Kv_rated (req.)", f"{_vs['Kv_rated']:.2f} m³/h·bar^0.5")
                    _vc5.metric("Cv_rated (req.)", f"{_vs['Cv_rated']:.2f} US gal/min·psi^0.5")
                    st.caption(f"P_in = {_vs['P_in']:.3f} bara  ·  "
                               f"Kv_rated is the minimum full-open Kv needed to pass "
                               f"{_vs['Q_m3h']:.3f} m³/h at {_vs['dp_kpa']:.1f} kPa "
                               f"with valve at {_vs['opening']:.0f} % opening.")
                    if len(valve_sizing) > 1:
                        st.divider()

        # ── Segment Analysis ──────────────────────────────────────────────────
        _primary_cols = ["Seg", "Type", "Pipe", "ID (mm)", "L (m)", "Fittings",
                         "Regime", "ΔP (kPa)", "ΔP_fric/100m (kPa)",
                         "P_in (bara)", "P_out (bara)", "V_m (m/s)", "V_m/V_e"]
        _detail_cols  = ["V_sg (m/s)", "V_sl (m/s)", "V_e (m/s)",
                         "ΔP_fric (kPa)", "ΔP_grav (kPa)", "ΔP_accel (kPa)",
                         "ρ_g (kg/m³)", "L_eff (m)", "α (void)", "dP/dz (Pa/m)", "Material"]
        _grid_df = pd.DataFrame(grid_records)
        _seg_col_cfg = {
            "ID (mm)":         st.column_config.NumberColumn(format="%.1f"),
            "L (m)":           st.column_config.NumberColumn(format="%.1f"),
            "ΔP (kPa)":        st.column_config.NumberColumn(format="%.3f"),
            "P_in (bara)":     st.column_config.NumberColumn(format="%.4f"),
            "P_out (bara)":    st.column_config.NumberColumn(format="%.4f"),
            "V_m (m/s)":       st.column_config.NumberColumn(format="%.3f"),
            "V_m/V_e":         st.column_config.ProgressColumn(
                                   format="%.3f", min_value=0.0, max_value=1.0),
            "V_sg (m/s)":      st.column_config.NumberColumn(format="%.3f"),
            "V_sl (m/s)":      st.column_config.NumberColumn(format="%.3f"),
            "V_e (m/s)":       st.column_config.NumberColumn(format="%.2f"),
            "ΔP_fric/100m (kPa)": st.column_config.NumberColumn(format="%.2f"),
            "ΔP_fric (kPa)":   st.column_config.NumberColumn(format="%.3f"),
            "ΔP_grav (kPa)":   st.column_config.NumberColumn(format="%.3f"),
            "ΔP_accel (kPa)":  st.column_config.NumberColumn(format="%.3f"),
            "ρ_g (kg/m³)":    st.column_config.NumberColumn(format="%.4f"),
            "L_eff (m)":       st.column_config.NumberColumn(format="%.2f"),
            "α (void)":        st.column_config.NumberColumn(format="%.4f"),
            "dP/dz (Pa/m)":    st.column_config.NumberColumn(format="%.2f"),
        }
        _existing_primary = [c for c in _primary_cols if c in _grid_df.columns]
        st.subheader("Segment Analysis")
        st.dataframe(_grid_df[_existing_primary],
                     column_config=_seg_col_cfg,
                     hide_index=True, width='stretch')
        _existing_detail = [c for c in _detail_cols if c in _grid_df.columns]
        if _existing_detail:
            with st.expander("Detail columns"):
                st.dataframe(_grid_df[["Seg"] + _existing_detail],
                             column_config=_seg_col_cfg,
                             hide_index=True, width='stretch')

        # ── Key Results — fill the top-of-column placeholder ─────────────────
        total_dp_kpa         = ((P_bara*1e5) - current_P) / 1000.0
        outlet_pressure_bara = max(0.1, current_P/1e5)
        pipe_length_m        = sum(s.get("length", 0.0) for s in st.session_state[k("segments")])
        _T_out  = current_T_C
        _T_in   = T_C if T_C is not None else _T_out
        _dT     = _T_out - _T_in
        _has_hx = any(s.get("kind") == "hx" for s in st.session_state[k("segments")])

        with _key_results_ph.container():
            with st.container(border=True):
                # Accent banner: ΔP at a glance
                st.markdown(
                    f'<div style="background:{accent};border-radius:6px;'
                    f'padding:0.45rem 1rem 0.3rem;margin-bottom:0.6rem">'
                    f'<span style="color:white;font-size:1.55rem;font-weight:700">'
                    f'ΔP = {total_dp_kpa:.2f} kPa</span>'
                    f'<span style="color:rgba(255,255,255,0.75);font-size:0.95rem;'
                    f'margin-left:1rem">({total_dp_kpa/100:.4f} bar'
                    f' &nbsp;·&nbsp; {total_dp_kpa*1000:.0f} Pa)</span>'
                    f'</div>',
                    unsafe_allow_html=True)

                # Pressure / temperature metrics
                _kr1, _kr2, _kr3, _kr4 = st.columns(4)
                _kr1.metric("P_in",  f"{P_bara:.3f} bara")
                _kr2.metric("P_out", f"{outlet_pressure_bara:.4f} bara",
                            delta=-(total_dp_kpa / 100), delta_color="inverse")
                _kr3.metric("T_in",  f"{_T_in:.1f} °C")
                _kr4.metric("T_out", f"{_T_out:.1f} °C",
                            delta=(_dT if _has_hx else None))

                # Outlet composition (compact)
                _out_gas_flows = _eff_gas_flows or {}
                _out_liq_kgh   = props.get("m_lye_kgh", 0.0)
                _gas_parts = [f"**{sp}** {m:.3g} kg/h"
                              for sp, m in _out_gas_flows.items() if m > 0]
                _gas_total = sum(_out_gas_flows.values())
                if _gas_parts:
                    st.markdown("**Outlet composition** &nbsp; *(constant along pipe)*",
                                unsafe_allow_html=True)
                    _oc1, _oc2 = st.columns([3, 1])
                    _oc1.caption("Gas phase:  " + "  ·  ".join(_gas_parts)
                                 + f"  ·  Total {_gas_total:.3g} kg/h")
                    _oc2.caption(f"Liquid:  {_out_liq_kgh:.3g} kg/h")

                # ΔP decomposition (collapsed by default)
                with st.expander("ΔP decomposition"):
                    d1, d2, d3 = st.columns(3)
                    d1.metric("Σ Frictional",     f"{total_dp_fric_kpa:.3f} kPa",
                              help="Sum of frictional component across all segments")
                    d2.metric("Σ Gravitational",  f"{total_dp_grav_kpa:.3f} kPa",
                              help="Gravitational head loss (negative = recovery on downflow)")
                    d3.metric("Σ Accelerational", f"{total_dp_accel_kpa:.3f} kPa",
                              help="Residual B&B inclination correction; zero for other correlations.")
                    st.caption(f"Pipe length {pipe_length_m:.1f} m  ·  "
                               f"Eff. length (+ fittings) {cumulative_distance:.1f} m")

        # ΔP decomposition stacked bar chart
        _seg_labels = [r["Seg"] + " " + r["Pipe"] for r in grid_records]
        _dp_fric    = [r["ΔP_fric (kPa)"] for r in grid_records]
        _dp_grav    = [r["ΔP_grav (kPa)"] for r in grid_records]
        _dp_accel   = [r["ΔP_accel (kPa)"] for r in grid_records]
        fig_decomp = go.Figure()
        fig_decomp.add_trace(go.Bar(name="Frictional",    x=_seg_labels, y=_dp_fric,
                                    marker_color="#2563EB", opacity=0.90))
        fig_decomp.add_trace(go.Bar(name="Gravitational", x=_seg_labels, y=_dp_grav,
                                    marker_color="#D97706", opacity=0.90))
        if any(abs(v) > 1e-4 for v in _dp_accel):
            fig_decomp.add_trace(go.Bar(name="Accelerational", x=_seg_labels, y=_dp_accel,
                                        marker_color="#059669", opacity=0.90))
        fig_decomp.update_layout(
            barmode="relative", yaxis_title="ΔP (kPa)", xaxis_title="Segment",
            title=dict(text=f"ΔP Decomposition — {correlation}  ·  void: {voidage_method}",
                       font=dict(size=13), x=0),
            template="plotly_white", height=300,
            margin=dict(l=60, r=20, t=40, b=60),
            paper_bgcolor="white", plot_bgcolor="white",
            xaxis=dict(gridcolor="#F1F5F9", linecolor="#E2E8F0"),
            yaxis=dict(gridcolor="#F1F5F9", linecolor="#E2E8F0", zeroline=True,
                       zerolinecolor="#94A3B8", zerolinewidth=1),
            legend=dict(bgcolor="rgba(255,255,255,0.9)", bordercolor="#E2E8F0",
                        borderwidth=1, orientation="h", y=1.08),
            font=dict(size=12, color="#374151"))
        st.plotly_chart(fig_decomp, width='stretch', key=k("fig_decomp"))

    # ── VISUALISATIONS ────────────────────────────────────────────────────────
    _nodes = [(0.0, 0.0)]
    for _seg in st.session_state[k("segments")]:
        _xl, _yl = _nodes[-1]
        _seg_type = _seg.get("type", "Horizontal")
        _seg_len  = _seg.get("length", 0.0)
        if _seg_type == "Horizontal":
            _nodes.append((_xl+_seg_len, _yl))
        elif _seg_type == "Vertical Upflow":
            _nodes.append((_xl, _yl+_seg_len))
        else:
            _nodes.append((_xl, _yl-_seg_len))

    _DN_LW = {"DN20":2,"DN25":3,"DN40":5,"DN50":7,"DN80":9,"DN100":11,"DN150":15,"DN200":19,"DN250":23}

    fig_sch = go.Figure()
    _seen_reg = set()
    for _i, (_seg, _rec) in enumerate(zip(st.session_state[k("segments")], grid_records)):
        _x0,_y0 = _nodes[_i]; _x1,_y1 = _nodes[_i+1]
        _reg = _rec["Regime"]
        _col = _regime_color(_reg, _REGIME_LINE_KW, "#64748B")
        _lw  = _DN_LW.get(_seg.get("dn","DN50"), 10)
        _show = _reg not in _seen_reg; _seen_reg.add(_reg)
        _seg_kind = _seg.get("kind", "pipe")
        _lh = (f"Liner: {_seg['liner_material']} {_seg['liner_thickness_mm']:.1f} mm"
               f"  →  ID {_rec['ID (mm)']:.1f} mm<br>" if _seg.get("lined") else "")
        _hover = (f"<b>Seg #{_i+1}  {_seg.get('dn','')}/{_seg.get('pn','')}</b><br>"
                  +_lh+f"{_seg.get('type', _seg_kind)},  L={_seg.get('length',0.0):.1f} m<br>"
                  f"Regime: {_reg}<br>ΔP: {_rec['ΔP (kPa)']:.3f} kPa  ·  "
                  f"V_sg: {_rec['V_sg (m/s)']:.3f} m/s<extra></extra>")
        if _seg_kind in ("valve", "hx"):
            # Zero-length inline component — render as a diamond marker
            _sym = "diamond" if _seg_kind == "valve" else "square"
            fig_sch.add_trace(go.Scatter(
                x=[_x0], y=[_y0], mode="markers",
                marker=dict(symbol=_sym, size=14, color=_col,
                            line=dict(color="white", width=2)),
                name=_reg, legendgroup=_reg, showlegend=_show,
                hovertemplate=_hover))
            fig_sch.add_annotation(
                x=_x0, y=_y0,
                text=f"<b>#{_i+1}</b>",
                showarrow=False, font=dict(size=9, color="#1E293B"),
                bgcolor="rgba(255,255,255,0.85)", bordercolor=_col, borderwidth=1.5,
                borderpad=2, xanchor="left", yanchor="bottom", xshift=10)
        else:
            fig_sch.add_trace(go.Scatter(
                x=[_x0,_x1], y=[_y0,_y1], mode="lines",
                line=dict(color=_col, width=_lw), name=_reg, legendgroup=_reg,
                showlegend=_show,
                hovertemplate=_hover))
            fig_sch.add_annotation(
                x=_x0+(_x1-_x0)*0.65, y=_y0+(_y1-_y0)*0.65,
                ax=_x0+(_x1-_x0)*0.5, ay=_y0+(_y1-_y0)*0.5,
                xref="x", yref="y", axref="x", ayref="y",
                showarrow=True, arrowhead=2, arrowsize=1.8, arrowwidth=2.5, arrowcolor=_col)
            _reg_short = _reg.split("/")[0].strip()[:18]
            fig_sch.add_annotation(
                x=(_x0+_x1)/2, y=(_y0+_y1)/2,
                text=f"<b>#{_i+1}</b> {_seg.get('dn','')}<br>"
                     f"<span style='font-size:9px;color:{_col}'>{_reg_short}</span>",
                showarrow=False, font=dict(size=10, color="#1E293B"),
                bgcolor="rgba(255,255,255,0.88)", bordercolor=_col, borderwidth=1.5,
                borderpad=3, xanchor="center", yanchor="middle")

    _all_x=[n[0] for n in _nodes]; _all_y=[n[1] for n in _nodes]
    _xspan=max(_all_x)-min(_all_x); _yspan=max(_all_y)-min(_all_y)
    _xpad=max(_xspan*0.12,2.0); _ypad=max(_yspan*0.12,max(_xspan*0.10,2.0))
    fig_sch.add_trace(go.Scatter(x=[_nodes[0][0]],y=[_nodes[0][1]],
        mode="markers+text", marker=dict(size=14,color="#059669",symbol="circle",
        line=dict(color="white",width=2)), text=["IN"], textposition="bottom center",
        textfont=dict(size=11,color="#059669"), showlegend=False,
        hovertemplate=f"Inlet  ·  {P_bara:.2f} bara<extra></extra>"))
    _otp = "top center" if _nodes[-1][1] >= _nodes[0][1] else "bottom center"
    fig_sch.add_trace(go.Scatter(x=[_nodes[-1][0]],y=[_nodes[-1][1]],
        mode="markers+text", marker=dict(size=14,color="#DC2626",symbol="circle",
        line=dict(color="white",width=2)), text=["OUT"], textposition=_otp,
        textfont=dict(size=11,color="#DC2626"), showlegend=False,
        hovertemplate=f"Outlet  ·  {outlet_pressure_bara:.4f} bara<extra></extra>"))
    fig_sch.update_layout(template="plotly_white", height=440,
        margin=dict(l=60,r=20,t=30,b=50), hovermode="closest",
        paper_bgcolor="white", plot_bgcolor="white",
        xaxis=dict(title="Horizontal Distance (m)", gridcolor="#F1F5F9",
                   zeroline=True, zerolinecolor="#CBD5E1", linecolor="#E2E8F0",
                   range=[min(_all_x)-_xpad, max(_all_x)+_xpad]),
        yaxis=dict(title="Elevation (m)", gridcolor="#F1F5F9",
                   zeroline=True, zerolinecolor="#CBD5E1", linecolor="#E2E8F0",
                   range=[min(_all_y)-_ypad, max(_all_y)+_ypad]),
        legend=dict(title="Flow Regime", bgcolor="rgba(255,255,255,0.9)",
                    bordercolor="#E2E8F0", borderwidth=1, font=dict(size=11)),
        font=dict(size=12,color="#374151"))

    fig_prof = go.Figure()
    _x0p = 0.0
    for _x1p, _rp in zip(cumulative_positions, regime_bands):
        _fill  = _regime_color(_rp, _REGIME_FILL_KW,   "rgba(241,245,249,0.60)")
        _bord  = _regime_color(_rp, _REGIME_BORDER_KW, "#CBD5E1")
        fig_prof.add_shape(type="rect", x0=_x0p, x1=_x1p, y0=0, y1=1, yref="paper",
                           fillcolor=_fill, line=dict(color=_bord,width=1), layer="below")
        fig_prof.add_annotation(x=(_x0p+_x1p)/2, y=0.96, yref="paper",
                                text=_rp[:18], showarrow=False,
                                font=dict(size=9,color="#475569"))
        _x0p = _x1p
    fig_prof.add_trace(go.Scatter(x=pressure_profile_x, y=pressure_profile_y,
        mode="lines+markers", line=dict(color=accent, width=2.5),
        marker=dict(size=7, color=accent),
        hovertemplate="Distance: %{x:.2f} m<br>Pressure: %{y:.4f} bara<extra></extra>"))
    fig_prof.update_layout(xaxis_title="Pipeline Distance (m)", yaxis_title="Pressure (bara)",
        template="plotly_white", height=340, margin=dict(l=60,r=20,t=30,b=50),
        hovermode="x unified", showlegend=False, paper_bgcolor="white", plot_bgcolor="white",
        xaxis=dict(gridcolor="#F1F5F9",zeroline=False,linecolor="#E2E8F0"),
        yaxis=dict(gridcolor="#F1F5F9",zeroline=False,linecolor="#E2E8F0"),
        font=dict(size=12,color="#374151"))

    # ── Flow Regime Map (V_sg vs V_sl, log-log) ─────────────────────────────
    _pipe_recs = [r for r in grid_records if r.get("V_sg (m/s)", 0) > 0 or r.get("V_sl (m/s)", 0) > 0]

    st.divider()
    _tab_names = ["Pipeline Schematic", "Pressure Profile", "Flow Regime Map"]
    if _is_vle:
        _tab_names.append("Phase Distribution")
    _tabs = st.tabs(_tab_names)
    tab_sch, tab_prof_tab, tab_regime_map = _tabs[0], _tabs[1], _tabs[2]
    tab_vle_dist = _tabs[3] if _is_vle else None

    with tab_sch:
        st.plotly_chart(fig_sch, width='stretch', key=k("fig_sch"))
        st.caption("Line width ∝ DN  ·  Colour = flow regime  ·  Regime name on each segment")
    with tab_prof_tab:
        st.plotly_chart(fig_prof, width='stretch', key=k("fig_prof"))
    with tab_regime_map:
        # Determine single-phase status for regime map (need both phases for meaningful map)
        _gas_ok_rm  = bool(_eff_gas_flows) and any(v > 0 for v in (_eff_gas_flows or {}).values())
        _liq_ok_rm  = (bool(_eff_liquid_flows) and any(v > 0 for v in (_eff_liquid_flows or {}).values())) \
                      or (_eff_q_lye or 0) > 0
        _sp_rm = not _is_vle and not (_gas_ok_rm and _liq_ok_rm)
        if _sp_rm:
            _sp_phase_rm = "gas" if _gas_ok_rm else "liquid"
            st.info(
                f"Single-phase {_sp_phase_rm} flow — the flow regime map applies only to "
                "two-phase (gas + liquid) flow. No regime boundaries to display."
            )
        elif _pipe_recs:
            # ── Compute both regime grids ─────────────────────────────────────
            _p      = props
            _D_repr = (_pipe_recs[0]["ID (mm)"] / 1000.0) if _pipe_recs else 0.05
            _common_kw = dict(
                rhol=float(_p["rho_l"]),
                rhog=float(_p["rho_g"]),
                mul=float(_p["mu_l"]),
                mug=float(_p["mu_g"]),
                sigma=float(_p.get("sigma") or 0.072),
                D=float(_D_repr),
                roughness=4.6e-5,
            )
            _td_h, _full_h, _vsl_h, _vsg_h = _compute_regime_grid(**_common_kw, use_horiz=True)
            _td_v, _full_v, _vsl_v, _vsg_v = _compute_regime_grid(**_common_kw, use_horiz=False)

            # ── Split operating points by segment orientation ─────────────────
            _horiz_recs = [r for r in _pipe_recs if r.get("Type", "Horizontal") == "Horizontal"]
            _vert_recs  = [r for r in _pipe_recs if r.get("Type", "Horizontal") != "Horizontal"]

            # ── Build both figures ────────────────────────────────────────────
            _fig_h = _build_regime_fig(
                _td_h, _full_h, _vsl_h, _vsg_h, _horiz_recs,
                "Horizontal — Taitel-Dukler + Mandhane-Gregory-Aziz",
            )
            _fig_v = _build_regime_fig(
                _td_v, _full_v, _vsl_v, _vsg_v, _vert_recs,
                "Vertical — Wallis / void-fraction",
            )

            # ── Side-by-side display ──────────────────────────────────────────
            _col_h, _col_v = st.columns(2)
            with _col_h:
                st.plotly_chart(_fig_h, width='stretch', key=k("fig_regime_h"))
            with _col_v:
                st.plotly_chart(_fig_v, width='stretch', key=k("fig_regime_v"))

            st.caption(
                f"Background zones computed for inlet conditions: "
                f"ρ_l = {_p['rho_l']:.1f} kg/m³, ρ_g = {_p['rho_g']:.4f} kg/m³, "
                f"D = {_D_repr*1000:.1f} mm.  "
                "Left: Taitel-Dukler (1976) + Mandhane-Gregory-Aziz (1974).  "
                "Right: Wallis annular criterion + void-fraction thresholds.  "
                "● horizontal segment  ◆ vertical segment"
            )
        else:
            st.info("No pipe segments with velocity data to plot.")

    # ── VLE PHASE DISTRIBUTION TAB ────────────────────────────────────────────
    if _is_vle and tab_vle_dist is not None and stream_records:
        with tab_vle_dist:
            _vd_dist   = pressure_profile_x                              # [0, L1, L1+L2, …]
            _vd_P      = [sr["P (bara)"]          for sr in stream_records]
            _vd_T      = [sr["T (°C)"]            for sr in stream_records]
            _vd_x      = [sr["x (−)"]             for sr in stream_records]
            _vkey_vap  = f"{vle_fluid_id} vapour  kg/h"
            _vkey_liq  = f"{vle_fluid_id} liquid  kg/h"
            _vd_vap    = [sr.get(_vkey_vap, 0.0)  for sr in stream_records]
            _vd_liq    = [sr.get(_vkey_liq, 0.0)  for sr in stream_records]
            _vd_labels = [sr["Stream"]             for sr in stream_records]

            # ── Chart 1: P and T_sat vs distance ─────────────────────────────
            _fig_pt = go.Figure()
            _fig_pt.add_trace(go.Scatter(
                x=_vd_dist, y=_vd_P, name="P (bara)",
                mode="lines+markers", line=dict(color="#2563EB", width=2.5),
                marker=dict(size=7), yaxis="y1",
                hovertemplate="Distance: %{x:.2f} m<br>P: %{y:.3f} bara<extra></extra>",
            ))
            _fig_pt.add_trace(go.Scatter(
                x=_vd_dist, y=_vd_T, name="T_sat (°C)",
                mode="lines+markers", line=dict(color="#D97706", width=2.5, dash="dash"),
                marker=dict(size=7, symbol="diamond"), yaxis="y2",
                hovertemplate="Distance: %{x:.2f} m<br>T_sat: %{y:.2f} °C<extra></extra>",
            ))
            _fig_pt.update_layout(
                template="plotly_white", height=240,
                margin=dict(l=60, r=60, t=30, b=40),
                xaxis=dict(title="Cumulative distance (m)", gridcolor="#F1F5F9"),
                yaxis=dict(title="Pressure (bara)", color="#2563EB",
                           gridcolor="#F1F5F9", zeroline=False),
                yaxis2=dict(title="T_sat (°C)", color="#D97706",
                            overlaying="y", side="right", zeroline=False),
                legend=dict(orientation="h", x=0.01, y=1.08, bgcolor="rgba(0,0,0,0)"),
                hovermode="x unified", font=dict(size=11, color="#374151"),
            )
            st.plotly_chart(_fig_pt, width='stretch', key=k("fig_vle_pt"))

            # ── Chart 2: vapour/liquid split as stacked area ──────────────────
            _fig_split = go.Figure()
            _fig_split.add_trace(go.Scatter(
                x=_vd_dist, y=_vd_liq, name="Liquid (kg/h)",
                mode="lines", line=dict(color="#3B82F6", width=0),
                fill="tozeroy", fillcolor="rgba(59,130,246,0.25)",
                hovertemplate="Distance: %{x:.2f} m<br>Liquid: %{y:.1f} kg/h<extra></extra>",
                stackgroup="split",
            ))
            _fig_split.add_trace(go.Scatter(
                x=_vd_dist, y=_vd_vap, name="Vapour (kg/h)",
                mode="lines", line=dict(color="#F97316", width=0),
                fill="tonexty", fillcolor="rgba(249,115,22,0.25)",
                hovertemplate="Distance: %{x:.2f} m<br>Vapour: %{y:.1f} kg/h<extra></extra>",
                stackgroup="split",
            ))
            # Quality x as an overlay line on the right axis
            _fig_split.add_trace(go.Scatter(
                x=_vd_dist, y=_vd_x, name="Quality x (−)",
                mode="lines+markers", line=dict(color="#7C3AED", width=2, dash="dot"),
                marker=dict(size=6), yaxis="y2",
                hovertemplate="Distance: %{x:.2f} m<br>x: %{y:.4f}<extra></extra>",
            ))
            _fig_split.update_layout(
                template="plotly_white", height=260,
                margin=dict(l=60, r=60, t=30, b=50),
                xaxis=dict(title="Cumulative distance (m)", gridcolor="#F1F5F9"),
                yaxis=dict(title="Flow (kg/h)", gridcolor="#F1F5F9", zeroline=False),
                yaxis2=dict(title="Quality x (−)", color="#7C3AED",
                            overlaying="y", side="right",
                            range=[-0.02, 1.02], zeroline=False),
                legend=dict(orientation="h", x=0.01, y=1.10, bgcolor="rgba(0,0,0,0)"),
                hovermode="x unified", font=dict(size=11, color="#374151"),
            )
            st.plotly_chart(_fig_split, width='stretch', key=k("fig_vle_split"))

            # ── Table ─────────────────────────────────────────────────────────
            _vd_table = []
            for _i, _sr in enumerate(stream_records):
                _vd_table.append({
                    "Stream":         _sr["Stream"],
                    "P (bara)":       _sr["P (bara)"],
                    "T_sat (°C)":     _sr["T (°C)"],
                    "x (−)":          _sr["x (−)"],
                    "Vapour (kg/h)":  _sr.get(_vkey_vap, 0.0),
                    "Liquid (kg/h)":  _sr.get(_vkey_liq, 0.0),
                    "α (−)":          _sr.get("α (−)", "—"),
                })
            st.dataframe(
                pd.DataFrame(_vd_table),
                hide_index=True, width='stretch',
                column_config={
                    "P (bara)":      st.column_config.NumberColumn(format="%.3f"),
                    "T_sat (°C)":    st.column_config.NumberColumn(format="%.2f"),
                    "x (−)":         st.column_config.NumberColumn(format="%.5f"),
                    "Vapour (kg/h)": st.column_config.NumberColumn(format="%.2f"),
                    "Liquid (kg/h)": st.column_config.NumberColumn(format="%.2f"),
                    "α (−)":         st.column_config.NumberColumn(format="%.4f"),
                },
            )

    # ── HEAT & MASS BALANCE (full width, below charts) ───────────────────────
    if stream_records:
        _hmb_rows_def = [("P  (bara)", "P (bara)"), ("T  (°C)", "T (°C)")]
        if _is_vle:
            _hmb_rows_def += [
                (f"  {vle_fluid_id} vapour  kg/h", f"{vle_fluid_id} vapour  kg/h"),
                (f"  {vle_fluid_id} liquid  kg/h", f"{vle_fluid_id} liquid  kg/h"),
            ]
        else:
            for _gsp in (_eff_gas_flows or {}):
                _hmb_rows_def.append((f"  {_gsp}  kg/h", f"gas:{_gsp}"))
            if _eff_gas_flows:
                _hmb_rows_def.append(("  Σ gas  kg/h", "Ṁ_gas (kg/h)"))
            for _lsp in (_eff_liquid_flows or {}):
                _hmb_rows_def.append((f"  {_lsp}  kg/h", f"liq:{_lsp}"))
            if _eff_liquid_flows:
                _hmb_rows_def.append(("  Σ liquid  kg/h", "Ṁ_liq (kg/h)"))
        _hmb_rows_def += [
            ("x  (−)",       "x (−)"),
            ("α  (−)",       "α (−)"),
            ("ρ_hom  kg/m³", "ρ_hom (kg/m³)"),
        ]
        _hmb_table = []
        for _prop_label, _key in _hmb_rows_def:
            _row = {"Property": _prop_label}
            for _sr in stream_records:
                _row[_sr["Stream"]] = _sr.get(_key, None)
            _hmb_table.append(_row)
        _hmb_df      = pd.DataFrame(_hmb_table)
        _stream_cols = [sr["Stream"] for sr in stream_records]
        _num_col_cfg = {col: st.column_config.NumberColumn(format="%.4g")
                        for col in _stream_cols}
        st.subheader("Heat & Mass Balance")
        st.dataframe(_hmb_df, column_config=_num_col_cfg,
                     hide_index=True, width='stretch')


    # ── METHOD SENSITIVITY ────────────────────────────────────────────────────
    # Detect single-phase: only one phase present → all correlations identical.
    _gas_present_ms  = bool(_eff_gas_flows) and any(v > 0 for v in (_eff_gas_flows or {}).values())
    _liq_present_ms  = (bool(_eff_liquid_flows) and any(v > 0 for v in (_eff_liquid_flows or {}).values())) \
                       or (_eff_q_lye or 0) > 0
    _is_single_phase = not _is_vle and not (_gas_present_ms and _liq_present_ms)

    # Hash the inputs so we only recompute when something actually changes.
    _ms_hash_src = {
        "P": P_bara, "T": T_C, "liq": _eff_liq_type, "q": _eff_q_lye,
        "gas": {kk: float(vv) for kk, vv in (_eff_gas_flows or {}).items()},
        "liq_flows": {kk: float(vv) for kk, vv in (_eff_liquid_flows or {}).items()},
        "vle_fluid": vle_fluid_id, "vle_x": float(vle_x) if vle_x is not None else None,
        "vle_m": float(vle_m_kgs) if vle_m_kgs is not None else None,
        "segs": [(s.get("type", s.get("kind", "")), s.get("dn", ""), s.get("length", 0.0))
                 for s in st.session_state[k("segments")]],
    }
    _ms_hash = hashlib.md5(json.dumps(_ms_hash_src, sort_keys=True).encode()).hexdigest()
    if st.session_state.get(k("ms_hash")) != _ms_hash:
        with st.spinner("Running method sensitivity (12 combinations)…"):
            st.session_state[k("sens_data_case")] = engine.run_sensitivity(
                P_bara, T_C if T_C is not None else 20.0,
                _eff_gas_flows or {}, _eff_liq_type, _eff_q_lye,
                st.session_state[k("segments")],
                custom_gas=custom_gas, custom_liquid=_eff_custom_liq,
                liquid_flows_kgh=_eff_liquid_flows if not _is_vle else None,
                vle_fluid=vle_fluid_id if _is_vle else None,
                vle_x_mass=vle_x if _is_vle else None,
                vle_m_total_kgs=vle_m_kgs if _is_vle else None,
            )
            st.session_state[k("ms_hash")] = _ms_hash
    _sens_data = st.session_state.get(k("sens_data_case"), [])

    with st.expander("Method Sensitivity — ΔP across all correlations × void-fraction models",
                     expanded=False):
        if _is_single_phase:
            _sp_phase = "gas" if _gas_present_ms else "liquid"
            st.info(
                f"Single-phase {_sp_phase} flow — all two-phase correlations reduce to "
                "Darcy-Weisbach and return identical results. Method sensitivity is not applicable."
            )
        elif not _sens_data:
            st.info("No sensitivity results available yet.")
        else:
            # Build label list and determine currently-selected method label
            _sel_corr_s = _CORR_SHORT.get(correlation, correlation)
            _sel_void_s = _VOID_SHORT.get(voidage_method, voidage_method)
            _sel_lbl_s  = f"{_sel_corr_s} / {_sel_void_s}"

            _s_labels, _s_dp, _s_ok = [], [], []
            for _sr in _sens_data:
                _lc = _CORR_SHORT.get(_sr["correlation"], _sr["correlation"])
                _lv = _VOID_SHORT.get(_sr["voidage"], _sr["voidage"])
                _s_labels.append(f"{_lc} / {_lv}")
                _s_dp.append(_sr["total_dp_kpa"] if _sr["ok"] else None)
                _s_ok.append(_sr["ok"])

            # Bar colors: highlight the selected combo
            _s_colors = ["#2563EB" if _l == _sel_lbl_s else "#93C5FD" for _l in _s_labels]

            _fig_ms = go.Figure()
            _fig_ms.add_trace(go.Bar(
                x=[v for v in _s_dp],
                y=_s_labels,
                orientation="h",
                marker=dict(color=_s_colors, line=dict(color="#1E40AF", width=0.5)),
                text=[f"{v:.2f}" if v is not None else "—" for v in _s_dp],
                textposition="outside",
                hovertemplate="%{y}<br>Total ΔP: %{x:.3f} kPa<extra></extra>",
            ))
            # Dashed vline at selected method
            _sel_dp_s = next(
                (v for v, l in zip(_s_dp, _s_labels) if l == _sel_lbl_s and v is not None), None)
            if _sel_dp_s is not None:
                _fig_ms.add_vline(
                    x=_sel_dp_s,
                    line=dict(color="#1D4ED8", width=1.5, dash="dash"),
                    annotation_text=f"Selected: {_sel_dp_s:.2f} kPa",
                    annotation_position="top right",
                    annotation_font=dict(size=9, color="#1D4ED8"),
                )
            _valid_dp_s = [v for v in _s_dp if v is not None]
            _x_lo = min(_valid_dp_s) * 0.85 if _valid_dp_s else 0
            _x_hi = max(_valid_dp_s) * 1.18 if _valid_dp_s else 1
            _fig_ms.update_layout(
                template="plotly_white", height=380,
                margin=dict(l=10, r=60, t=36, b=50),
                xaxis=dict(title="Total ΔP  (kPa)", range=[_x_lo, _x_hi],
                           gridcolor="#F1F5F9"),
                yaxis=dict(autorange="reversed"),
                font=dict(size=11, color="#374151"),
                title=dict(text=(f"Method sensitivity  —  selected: <b>{_sel_lbl_s}</b>"
                                 f"  ({_sel_dp_s:.2f} kPa)" if _sel_dp_s else "Method sensitivity"),
                           font=dict(size=12), x=0),
                showlegend=False,
            )
            st.plotly_chart(_fig_ms, width='stretch', key=k("fig_method_sens"))

            # Data table
            _ms_rows = []
            for _sr, _lbl in zip(_sens_data, _s_labels):
                _ms_rows.append({
                    "Method": _lbl,
                    "ΔP (kPa)": round(_sr["total_dp_kpa"], 3) if _sr["ok"] and _sr["total_dp_kpa"] is not None else None,
                    "Status": "OK" if _sr["ok"] else f"Error: {_sr.get('error','?')}",
                    "Selected": "★" if _lbl == _sel_lbl_s else "",
                })
            st.dataframe(
                pd.DataFrame(_ms_rows),
                hide_index=True,
                width='stretch',
                column_config={
                    "ΔP (kPa)": st.column_config.NumberColumn(format="%.3f"),
                },
            )

    # ── EXPORTS ───────────────────────────────────────────────────────────────
    st.divider()
    ex_tab_w, ex_tab_x = st.tabs(["Export Word (.docx)", "Export Excel (.xlsx)"])
    with ex_tab_w:
        _rpt_hash = hashlib.md5(json.dumps({
            "P": P_bara, "T": T_C,
            "gas_flows": {_kk: float(_vv) for _kk,_vv in gas_flows_kgh.items()},
            "liq_flows": {_kk: float(_vv) for _kk,_vv in (liquid_flows_kgh or {}).items()},
            "segs": [(s.get("type", s.get("kind","")),s.get("dn",""),s.get("pn",""),s.get("length",0.0),
                      tuple(sorted((f["type"],f["qty"]) for f in s.get("fittings_list",[]))),
                      s.get("lined",False),
                      s.get("liner_material","FEP"),s.get("liner_thickness_mm",1.0))
                     for s in st.session_state[k("segments")]],
        }, sort_keys=True).encode()).hexdigest()
        if st.session_state.get(k("rpt_hash")) != _rpt_hash:
            st.session_state[k("rpt_hash")]    = None
            st.session_state[k("rpt_bytes")]   = None
            st.session_state[k("rpt_hash_ch")] = None
            st.session_state[k("rpt_bytes_ch")]= None
        def _rpt_kwargs():
            return dict(
                P_bara=P_bara, T_C=T_C,
                gas_flows_kgh=gas_flows_kgh,
                liquid_type=liquid_type, q_lye=q_lye,
                props=props, grid_records=grid_records,
                segments=st.session_state[k("segments")],
                total_dp_kpa=total_dp_kpa,
                outlet_pressure_bara=outlet_pressure_bara,
                pipe_length_m=pipe_length_m,
                cumulative_distance=cumulative_distance,
                case_label=st.session_state.get(f"label_{cid}", f"Case {cid.upper()}"),
                flow_mode=flow_mode,
                custom_liquid=_eff_custom_liq,
                stream_records=stream_records if stream_records else None,
                sensitivity_results=_sens_data if not _is_single_phase else None,
            )
        _wg1, _wg2 = st.columns(2)
        with _wg1:
            if st.button("Generate (tables only)", type="primary",
                         width='stretch', key=k("gen_rpt")):
                with st.spinner("Building document…"):
                    _buf = report_generator.generate_report(
                        **_rpt_kwargs(), fig_sch=None, fig_prof=None)
                    st.session_state[k("rpt_bytes")] = _buf.getvalue()
                    st.session_state[k("rpt_hash")]  = _rpt_hash
        with _wg2:
            if st.button("Generate with charts", width='stretch',
                         key=k("gen_rpt_ch")):
                with st.spinner("Building document with charts…"):
                    _buf = report_generator.generate_report(
                        **_rpt_kwargs(), fig_sch=fig_sch, fig_prof=fig_prof)
                    st.session_state[k("rpt_bytes_ch")] = _buf.getvalue()
                    st.session_state[k("rpt_hash_ch")]  = _rpt_hash
        _wd1, _wd2 = st.columns(2)
        with _wd1:
            if st.session_state.get(k("rpt_bytes")) and \
               st.session_state.get(k("rpt_hash")) == _rpt_hash:
                st.download_button("Download tables only  (.docx)",
                    data=st.session_state[k("rpt_bytes")],
                    file_name=f"hydraulic_report_case_{cid}.docx",
                    mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                    width='stretch', key=k("dl_rpt"))
            elif st.session_state.get(k("rpt_bytes")):
                st.info("Inputs changed — regenerate.")
        with _wd2:
            if st.session_state.get(k("rpt_bytes_ch")) and \
               st.session_state.get(k("rpt_hash_ch")) == _rpt_hash:
                st.download_button("Download with charts  (.docx)",
                    data=st.session_state[k("rpt_bytes_ch")],
                    file_name=f"hydraulic_report_case_{cid}_charts.docx",
                    mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                    width='stretch', key=k("dl_rpt_ch"))
            elif st.session_state.get(k("rpt_bytes_ch")):
                st.info("Inputs changed — regenerate.")

    with ex_tab_x:
        def _build_xlsx_case():
            from io import BytesIO
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            BLUE="2563EB"; LGRAY="F1F5F9"; STRIPE="F8FAFC"
            _hf=Font(bold=True,color="FFFFFF",size=10,name="Calibri")
            _hfill=PatternFill("solid",fgColor=BLUE)
            _sf=Font(bold=True,size=9,name="Calibri")
            _sfill=PatternFill("solid",fgColor=LGRAY)
            _df=Font(size=9,name="Calibri")
            _af=PatternFill("solid",fgColor=STRIPE)
            _th=Side(style="thin",color="CBD5E1")
            _bdr=Border(left=_th,right=_th,top=_th,bottom=_th)
            def _hdr(ws,row,text,nc):
                c=ws.cell(row=row,column=1,value=text); c.font=_hf; c.fill=_hfill; c.border=_bdr
                for col in range(2,nc+2):
                    cc=ws.cell(row=row,column=col); cc.fill=_hfill; cc.border=_bdr
                return row+1
            def _sec(ws,row,text,nc):
                c=ws.cell(row=row,column=1,value=text); c.font=_sf; c.fill=_sfill; c.border=_bdr
                for col in range(2,nc+2):
                    cc=ws.cell(row=row,column=col); cc.fill=_sfill; cc.border=_bdr
                return row+1
            def _dat(ws,row,label,value,alt=False):
                fill=_af if alt else None
                lc=ws.cell(row=row,column=1,value=label); lc.font=_df; lc.border=_bdr
                if fill: lc.fill=fill
                vc=ws.cell(row=row,column=2,value=value); vc.font=_df; vc.border=_bdr
                if fill: vc.fill=fill
                return row+1
            wb=openpyxl.Workbook()
            ws=wb.active; ws.title="System"; ws.freeze_panes="A2"
            r=_hdr(ws,1,"System Summary",1)
            alt=False
            r=_sec(ws,r,"Boundary Conditions",1)
            for lbl,val in [("Inlet Pressure (bara)",round(P_bara,4)),
                             ("Temperature (°C)",round(T_C,2)),
                             ("Outlet Pressure (bara)",round(outlet_pressure_bara,4)),
                             ("Total ΔP (kPa)",round(total_dp_kpa,4)),
                             ("Total ΔP (bar)",round(total_dp_kpa/100,6)),
                             ("Pipe Length (m)",round(pipe_length_m,3)),
                             ("Effective Length (m)",round(cumulative_distance,3))]:
                r=_dat(ws,r,lbl,val,alt); alt=not alt
            r+=1; alt=False
            r=_sec(ws,r,"Gas Phase  —  inlet conditions",1)
            for _sp,_fl in gas_flows_kgh.items():
                r=_dat(ws,r,f"{_sp} mass flow (kg/h)",round(float(_fl),4),alt); alt=not alt
            for lbl,val in [("Total gas (kg/h)",round(props["m_gas_total_kgh"],4)),
                             ("ρ_g (kg/m³)",round(props["rho_g"],4)),
                             ("MW_mix (g/mol)",round(props["MW_mix_gmol"],3)),
                             ("x_gas (%)",round(props["x_gas"]*100,5)),
                             ("α void (%)",round(props["alpha"]*100,4))]:
                r=_dat(ws,r,lbl,val,alt); alt=not alt
            if props.get("P_sat_H2O_pa",0)>0:
                for lbl,val in [("H₂O vapour (kg/h)",round(props["m_vapor_h2o_kgh"],5)),
                                 ("P_sat H₂O (bara)",round(props["P_sat_H2O_pa"]/1e5,5))]:
                    r=_dat(ws,r,lbl,val,alt); alt=not alt
            r+=1; alt=False
            r=_sec(ws,r,f"Liquid Phase  —  {liquid_type}",1)
            for lbl,val in [("Volume flow (m³/h)",round(q_lye,4)),
                             ("Mass flow (kg/h)",round(props["m_lye_kgh"],3)),
                             ("ρ_l (kg/m³)",round(props["rho_l"],3)),
                             ("μ_l (mPa·s)",round(props["mu_l"]*1e3,4)),
                             ("σ (mN/m)",round(props["sigma"]*1e3,4))]:
                r=_dat(ws,r,lbl,val,alt); alt=not alt
            ws.column_dimensions["A"].width=42; ws.column_dimensions["B"].width=20
            ws2=wb.create_sheet("Segments"); ws2.freeze_panes="B2"
            n=len(grid_records)
            c=ws2.cell(row=1,column=1,value="Parameter"); c.font=_hf; c.fill=_hfill; c.border=_bdr
            for j,rec in enumerate(grid_records):
                lbl=f"Seg {rec['Seg']}  {rec['Pipe']}"
                cc=ws2.cell(row=1,column=j+2,value=lbl); cc.font=_hf; cc.fill=_hfill; cc.border=_bdr
            _ROWS=[("Orientation","Type"),("Pipe class","Pipe"),("Inner diameter (mm)","ID (mm)"),
                   ("Material","Material"),("Physical length (m)","L (m)"),
                   ("Effective length (m)","L_eff (m)"),
                   ("P_in  Inlet pressure (bara)","P_in (bara)"),
                   ("P_out  Outlet pressure (bara)","P_out (bara)"),
                   ("ΔP  Pressure drop (kPa)","ΔP (kPa)"),
                   ("dP/dz  Pressure gradient (Pa/m)","dP/dz (Pa/m)"),
                   ("Gas density ρ_g (kg/m³)","ρ_g (kg/m³)"),("Flow regime","Regime"),
                   ("V_sg  Superficial gas (m/s)","V_sg (m/s)"),
                   ("V_sl  Superficial liquid (m/s)","V_sl (m/s)"),
                   ("V_m  Mixture velocity (m/s)","V_m (m/s)"),
                   ("V_e  Erosion limit API RP 14E (m/s)","V_e (m/s)"),
                   ("V_m/V_e  Erosion ratio (–)","V_m/V_e")]
            for i,(label,key) in enumerate(_ROWS):
                row=i+2; alt=(i%2==1)
                lc=ws2.cell(row=row,column=1,value=label); lc.font=_df; lc.border=_bdr
                if alt: lc.fill=_af
                for j,rec in enumerate(grid_records):
                    vc=ws2.cell(row=row,column=j+2,value=rec.get(key,"")); vc.font=_df; vc.border=_bdr
                    if alt: vc.fill=_af
            ws2.column_dimensions["A"].width=46
            for j in range(n):
                ws2.column_dimensions[get_column_letter(j+2)].width=18
            buf=BytesIO(); wb.save(buf); buf.seek(0); return buf

        _xc1, _xc2 = st.columns([1, 2])
        with _xc1:
            if st.button("Generate Excel", width='stretch', key=k("gen_xl")):
                try:
                    st.session_state[k("xl_bytes")] = _build_xlsx_case().getvalue()
                except Exception as _xe:
                    st.error(f"Excel export failed: {_xe}")
        with _xc2:
            if st.session_state.get(k("xl_bytes")):
                st.download_button("Download  (.xlsx)",
                    data=st.session_state[k("xl_bytes")],
                    file_name=f"hydraulic_data_case_{cid}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    width='stretch', key=k("dl_xl"))

    return {
        "P_bara":               P_bara,
        "T_C":                  T_C,
        "gas_flows_kgh":        gas_flows_kgh,
        "liquid_type":          liquid_type,
        "q_lye":                q_lye,
        "liquid_flows_kgh":     liquid_flows_kgh,
        "props":                props,
        "grid_records":         grid_records,
        "total_dp_kpa":         total_dp_kpa,
        "total_dp_fric_kpa":    total_dp_fric_kpa,
        "total_dp_grav_kpa":    total_dp_grav_kpa,
        "total_dp_accel_kpa":   total_dp_accel_kpa,
        "outlet_pressure_bara": outlet_pressure_bara,
        "pipe_length_m":        pipe_length_m,
        "cumulative_distance":  cumulative_distance,
        "pressure_profile_x":   pressure_profile_x,
        "pressure_profile_y":   pressure_profile_y,
        "segments":             st.session_state[k("segments")],
        "correlation":          correlation,
        "voidage_method":       voidage_method,
        "custom_gas":           custom_gas,
        "custom_liquid":        custom_liquid,
        "fig_sch":              fig_sch,
        "fig_prof":             fig_prof,
        # VLE mode fields (None when not in VLE mode)
        "flow_mode":            "vle" if _is_vle else "gas_liquid",
        "vle_fluid":            vle_fluid_id,
        "vle_x_mass":           vle_x,
        "vle_m_total_kgs":      vle_m_kgs,
    }


# ============================================================================
# TOP-LEVEL TABS
# ============================================================================
# ============================================================================
# GOAL-SEEK HELPERS  — used by the Compare tab
# ============================================================================

def _calc_dp_at_p(res, P_bara_override):
    """
    Re-run the pipeline in res at a different inlet pressure.
    Returns (total_dp_kpa, outlet_bara).  Handles pipe, valve, and HX segments.
    """
    current_P   = P_bara_override * 1e5
    current_T_C = res.get("T_C", 25.0)
    total_dp    = 0.0
    corr  = res.get("correlation",     engine.TWO_PHASE_CORRELATIONS[0])
    void  = res.get("voidage_method",  engine.VOIDAGE_METHODS[0])
    cgas  = res.get("custom_gas")
    cliq  = res.get("custom_liquid")
    _is_vle_res = (res.get("flow_mode") == "vle" and res.get("vle_fluid"))

    def _get_props():
        if _is_vle_res:
            return engine.calculate_vle_properties(
                res["vle_fluid"], max(1e4, current_P) / 1e5,
                res["vle_x_mass"], res["vle_m_total_kgs"])
        return engine.calculate_two_phase_properties(
            max(1e4, current_P) / 1e5, current_T_C,
            res["gas_flows_kgh"], res["liquid_type"], res["q_lye"],
            custom_gas=cgas, custom_liquid=cliq,
            liquid_flows_kgh=res.get("liquid_flows_kgh"))

    for seg in res["segments"]:
        _kind = seg.get("kind", "pipe")

        if _kind == "valve":
            props    = _get_props()
            v_res    = engine.calculate_valve_dp(
                props, seg.get("Kv_m3h", 1.0),
                seg.get("opening_pct", 100.0),
                seg.get("characteristic", "equal-percentage"))
            total_dp  += v_res["dP_Pa"]
            current_P -= v_res["dP_Pa"]
            current_P  = max(1e4, current_P)
            continue

        if _kind == "hx":
            dp_pa      = seg.get("dp_kpa", 0.0) * 1000.0
            duty_kw    = seg.get("duty_kw", 0.0)
            if duty_kw != 0.0:
                props = _get_props()
                Cp = engine.estimate_mixture_cp(
                    props, current_T_C + 273.15, max(1e4, current_P))
                if props["m_total_kgs"] > 0 and Cp > 0:
                    current_T_C += (duty_kw * 1000.0) / (props["m_total_kgs"] * Cp)
            total_dp  += dp_pa
            current_P -= dp_pa
            current_P  = max(1e4, current_P)
            continue

        # Pipe segment
        D_seg  = engine.PIPE_DATABASE[seg["dn"]][seg["pn"]]
        lined  = seg.get("lined", False)
        lthk_m = seg.get("liner_thickness_mm", 1.0) / 1000.0
        lmat   = seg.get("liner_material", "FEP")
        D_eff  = D_seg - 2 * lthk_m if lined else D_seg
        rough  = (engine.LINER_ROUGHNESS[lmat] if lined
                  else engine.MATERIAL_ROUGHNESS[seg.get("material", "SS316L")])
        props  = _get_props()
        angle  = {"Horizontal": 0.0, "Vertical Upflow": np.pi / 2.0,
                  "Vertical Downflow": -np.pi / 2.0}[seg["type"]]
        le_fit = sum_le_fit(seg, D_eff)
        seg_res = engine.calculate_segment_pressure_drop(
            props, D_eff, rough, seg["length"] + le_fit, angle,
            correlation=corr, voidage_method=void)
        total_dp  += seg_res["dP_Pa"]
        current_P -= seg_res["dP_Pa"]
        current_P  = max(1e4, current_P)

    return total_dp / 1000.0, current_P / 1e5


def _calc_regimes_at_p(res, P_bara_override):
    """Run segments at the given inlet pressure; return list of dicts with
    seg label, pipe label, and regime string — one entry per segment."""
    current_P = P_bara_override * 1e5
    corr = res.get("correlation",    engine.TWO_PHASE_CORRELATIONS[0])
    void = res.get("voidage_method", engine.VOIDAGE_METHODS[0])
    cgas = res.get("custom_gas")
    cliq = res.get("custom_liquid")
    _is_vle_res = (res.get("flow_mode") == "vle" and res.get("vle_fluid"))
    current_T_C = res.get("T_C", 25.0)
    out = []
    for i, seg in enumerate(res["segments"]):
        _kind = seg.get("kind", "pipe")
        label = f"#{i + 1}"
        pipe_label = f"{seg.get('dn', '—')}/{seg.get('pn', '—')}"

        if _kind == "valve":
            if _is_vle_res:
                props = engine.calculate_vle_properties(
                    res["vle_fluid"], current_P / 1e5,
                    res["vle_x_mass"], res["vle_m_total_kgs"])
            else:
                props = engine.calculate_two_phase_properties(
                    current_P / 1e5, current_T_C,
                    res["gas_flows_kgh"], res["liquid_type"], res["q_lye"],
                    custom_gas=cgas, custom_liquid=cliq,
                    liquid_flows_kgh=res.get("liquid_flows_kgh"))
            vres = engine.calculate_valve_dp(
                props, seg.get("Kv_m3h", 1.0), seg.get("opening_pct", 100.0),
                seg.get("characteristic", "equal-percentage"))
            out.append({"seg": label, "pipe": pipe_label, "regime": "Valve"})
            current_P -= vres["dP_Pa"]
            current_P  = max(1e4, current_P)
            continue

        if _kind == "hx":
            duty_kw = seg.get("duty_kw", 0.0)
            if duty_kw != 0.0:
                if _is_vle_res:
                    props = engine.calculate_vle_properties(
                        res["vle_fluid"], current_P / 1e5,
                        res["vle_x_mass"], res["vle_m_total_kgs"])
                else:
                    props = engine.calculate_two_phase_properties(
                        current_P / 1e5, current_T_C,
                        res["gas_flows_kgh"], res["liquid_type"], res["q_lye"],
                        custom_gas=cgas, custom_liquid=cliq,
                        liquid_flows_kgh=res.get("liquid_flows_kgh"))
                Cp = engine.estimate_mixture_cp(props, current_T_C + 273.15, max(1e4, current_P))
                if props["m_total_kgs"] > 0 and Cp > 0:
                    current_T_C += (duty_kw * 1000.0) / (props["m_total_kgs"] * Cp)
            out.append({"seg": label, "pipe": pipe_label, "regime": "Heat Exchanger"})
            current_P -= seg.get("dp_kpa", 0.0) * 1000.0
            current_P  = max(1e4, current_P)
            continue

        D_seg  = engine.PIPE_DATABASE[seg["dn"]][seg["pn"]]
        lined  = seg.get("lined", False)
        lthk_m = seg.get("liner_thickness_mm", 1.0) / 1000.0
        lmat   = seg.get("liner_material", "FEP")
        D_eff  = D_seg - 2 * lthk_m if lined else D_seg
        rough  = (engine.LINER_ROUGHNESS[lmat] if lined
                  else engine.MATERIAL_ROUGHNESS[seg.get("material", "SS316L")])
        if _is_vle_res:
            props = engine.calculate_vle_properties(
                res["vle_fluid"], current_P / 1e5,
                res["vle_x_mass"], res["vle_m_total_kgs"])
        else:
            props  = engine.calculate_two_phase_properties(
                current_P / 1e5, current_T_C,
                res["gas_flows_kgh"], res["liquid_type"], res["q_lye"],
                custom_gas=cgas, custom_liquid=cliq,
                liquid_flows_kgh=res.get("liquid_flows_kgh"))
        angle  = {"Horizontal": 0.0, "Vertical Upflow": np.pi / 2.0,
                  "Vertical Downflow": -np.pi / 2.0}[seg["type"]]
        le_fit = sum_le_fit(seg, D_eff)
        seg_res = engine.calculate_segment_pressure_drop(
            props, D_eff, rough, seg["length"] + le_fit, angle,
            correlation=corr, voidage_method=void)
        out.append({
            "seg":    label,
            "pipe":   pipe_label,
            "regime": seg_res["regime"],
        })
        current_P -= seg_res["dP_Pa"]
        current_P  = max(1e4, current_P)
    return out


def _march_header_simple(tap_dists, gas_per_tap, liq_per_tap,
                          hdr, P_start_Pa, T_C, liquid_type, corr, void,
                          custom_liquid=None):
    """Pressure-march one header arm from farthest tap to T-junction.

    tap_dists    : list of distances from T (m) for each tap — sorted internally.
    gas_per_tap  : {species: kg/h} for ONE A-line.
    liq_per_tap  : m³/h liquid for ONE A-line.
    hdr          : dict with dn, pn, material, lined, liner_material,
                   liner_thickness_mm, fittings_list.
    custom_liquid: optional {rho_kgm3, mu_mpas, sigma_mnm} for Custom liquid type.
    Returns (total_dp_Pa, P_T_Pa, dp_fric_Pa, dp_grav_Pa, records).
    """
    if not tap_dists or not gas_per_tap or liq_per_tap <= 0:
        return 0.0, P_start_Pa, 0.0, 0.0, []

    dists = sorted(tap_dists, reverse=True)   # farthest first
    n     = len(dists)
    boundaries = dists + [0.0]                # segment endpoints, T at 0

    D_nom  = engine.PIPE_DATABASE[hdr["dn"]][hdr["pn"]]
    lined  = hdr.get("lined", False)
    lthk_m = hdr.get("liner_thickness_mm", 1.0) / 1000.0
    lmat   = hdr.get("liner_material", "FEP")
    D_eff  = D_nom - 2 * lthk_m if lined else D_nom
    rough  = (engine.LINER_ROUGHNESS[lmat] if lined
              else engine.MATERIAL_ROUGHNESS[hdr.get("material", "SS316L")])
    le_fit = sum_le_fit(hdr, D_eff)

    current_P   = P_start_Pa
    total_dp    = dp_fric = dp_grav = 0.0
    running_gas = {}
    running_liq = 0.0
    records     = []

    for i in range(n):
        for sp, kgh in gas_per_tap.items():
            running_gas[sp] = running_gas.get(sp, 0.0) + kgh
        running_liq += liq_per_tap

        seg_len = boundaries[i] - boundaries[i + 1]
        if seg_len <= 1e-6:
            continue

        eff_len = seg_len + le_fit / n    # distribute fittings equally
        props_seg = engine.calculate_two_phase_properties(
            current_P / 1e5, T_C, running_gas, liquid_type, running_liq,
            custom_liquid=custom_liquid)
        seg_res = engine.calculate_segment_pressure_drop(
            props_seg, D_eff, rough, eff_len, 0.0,
            correlation=corr, voidage_method=void)

        dP_Pa = seg_res["dP_Pa"]
        end_P = current_P - dP_Pa
        V_m   = seg_res["Vsg"] + seg_res["Vsl"]
        V_e, _ = engine.calculate_erosion_velocity(
            props_seg["rho_g"], props_seg["rho_l"], props_seg["x_gas"])

        records.append({
            "Seg":          f"#{i+1}",
            "Taps in seg":  i + 1,
            "From T (m)":   round(boundaries[i], 2),
            "To T (m)":     round(boundaries[i + 1], 2),
            "L (m)":        round(seg_len, 3),
            "Pipe":         f"{hdr['dn']}/{hdr['pn']}",
            "ID (mm)":      round(D_eff * 1000, 1),
            "Regime":       seg_res["regime"],
            "ΔP (kPa)":     round(dP_Pa / 1000, 3),
            "P_in (bara)":  round(current_P / 1e5, 4),
            "P_out (bara)": round(end_P / 1e5, 4),
            "V_m (m/s)":    round(V_m, 3),
            "V_m/V_e":      round(V_m / V_e if V_e > 0 else 0.0, 3),
            "Q_gas_kgh":    round(sum(running_gas.values()), 3),
            "Q_liq_m3h":    round(running_liq, 3),
            "ΔP_fric (kPa)": round(seg_res["dP_fric_Pa"] / 1000, 3),
            "ΔP_grav (kPa)": round(seg_res["dP_grav_Pa"] / 1000, 3),
        })

        total_dp  += dP_Pa
        dp_fric   += seg_res["dP_fric_Pa"]
        dp_grav   += seg_res["dP_grav_Pa"]
        current_P  = max(1e4, end_P)

    return total_dp, current_P, dp_fric, dp_grav, records


def _march_single_seg(seg, P_in_Pa, T_C, gas_flows, liquid_type, q_lye, corr, void,
                       custom_liquid=None):
    """ΔP for one horizontal pipe segment carrying combined flow.
    Returns (dp_Pa, P_out_Pa, dp_fric_Pa, dp_grav_Pa, record_dict).
    """
    D_nom = engine.PIPE_DATABASE[seg["dn"]][seg["pn"]]
    mat   = seg.get("material", "SS316L")
    rough = engine.MATERIAL_ROUGHNESS.get(mat, engine.MATERIAL_ROUGHNESS["SS316L"])
    le_fit = sum_le_fit(seg, D_nom)

    props   = engine.calculate_two_phase_properties(
        P_in_Pa / 1e5, T_C, gas_flows, liquid_type, q_lye,
        custom_liquid=custom_liquid)
    seg_res = engine.calculate_segment_pressure_drop(
        props, D_nom, rough, seg["length"] + le_fit, 0.0,
        correlation=corr, voidage_method=void)

    dp_Pa = seg_res["dP_Pa"]
    P_out = max(1e4, P_in_Pa - dp_Pa)
    V_m   = seg_res["Vsg"] + seg_res["Vsl"]
    V_e, _ = engine.calculate_erosion_velocity(
        props["rho_g"], props["rho_l"], props["x_gas"])

    record = {
        "Seg": "T-seg",
        "Pipe": f"{seg['dn']}/{seg['pn']}",
        "ID (mm)": round(D_nom * 1000, 1),
        "L (m)": seg["length"],
        "Regime": seg_res["regime"],
        "ΔP (kPa)": round(dp_Pa / 1000, 3),
        "P_in (bara)": round(P_in_Pa / 1e5, 4),
        "P_out (bara)": round(P_out / 1e5, 4),
        "V_m (m/s)": round(V_m, 3),
        "V_m/V_e": round(V_m / V_e if V_e > 0 else 0.0, 3),
        "Q_gas_kgh": round(sum(gas_flows.values()), 3),
        "Q_liq_m3h": round(q_lye, 3),
        "ΔP_fric (kPa)": round(seg_res["dP_fric_Pa"] / 1000, 3),
        "ΔP_grav (kPa)": round(seg_res["dP_grav_Pa"] / 1000, 3),
    }
    return dp_Pa, P_out, seg_res["dP_fric_Pa"], seg_res["dP_grav_Pa"], record


def _calc_header_dp_at_p(res_hdr, P_in_bara):
    """Re-run header (both arms) + T-segment at an overridden tap inlet pressure.
    Returns (total_dp_kpa, P_separator_bara).
    """
    P_start = P_in_bara * 1e5
    hdr   = res_hdr["header_pipe"]
    gpt   = res_hdr["gas_per_tap"]
    lpt   = res_hdr["liq_per_tap"]
    T_C   = res_hdr["T_C"]
    liq   = res_hdr["liquid_type"]
    cliq  = res_hdr.get("custom_liquid")
    corr  = res_hdr.get("correlation",    engine.TWO_PHASE_CORRELATIONS[0])
    void  = res_hdr.get("voidage_method", engine.VOIDAGE_METHODS[0])

    dp_l, P_T_l, *_ = _march_header_simple(
        res_hdr["left_taps"],  gpt, lpt, hdr, P_start, T_C, liq, corr, void,
        custom_liquid=cliq)
    dp_r, P_T_r, *_ = _march_header_simple(
        res_hdr["right_taps"], gpt, lpt, hdr, P_start, T_C, liq, corr, void,
        custom_liquid=cliq)

    dp_worst = max(dp_l, dp_r)
    P_T      = min(P_T_l, P_T_r)      # worst-arm pressure arriving at T

    t_seg = res_hdr.get("t_seg")
    if t_seg and t_seg.get("length", 0) > 0:
        dp_t, P_sep, *_ = _march_single_seg(
            t_seg, P_T, T_C, res_hdr["gas_flows_kgh"], liq,
            res_hdr["q_lye"], corr, void, custom_liquid=cliq)
        return (dp_worst + dp_t) / 1000.0, P_sep / 1e5

    return dp_worst / 1000.0, P_T / 1e5


def _apply_dn_override(res, dn_alt):
    """Return a deep copy of a branch result dict with dn replaced in all segments.
    Header pipe and T-segment are not modified."""
    import copy
    r = copy.deepcopy(res)
    for seg in r.get("segments", []):
        seg["dn"] = dn_alt
    return r


def _goal_seek_header(res_hdr, P_target_sep, tol=0.0005, max_iter=25):
    """Find the header inlet pressure (= line outlet) to achieve P_target_sep
    at the separator (end of T-segment).  Returns a result dict.
    """
    P_in = P_target_sep + res_hdr["total_dp_kpa"] / 100.0
    dp = P_sep = 0.0
    for i in range(max_iter):
        dp, P_sep = _calc_header_dp_at_p(res_hdr, P_in)
        error = P_sep - P_target_sep
        if abs(error) < tol:
            return {"P_hdr_in": P_in, "P_sep": P_sep, "dp_hdr": dp,
                    "iterations": i + 1, "converged": True}
        P_in -= error
    return {"P_hdr_in": P_in, "P_sep": P_sep, "dp_hdr": dp,
            "iterations": max_iter, "converged": abs(P_sep - P_target_sep) < tol * 20}


def _forward_stack(res_line, res_hdr, P_source_bara):
    """Forward march from a fixed inlet pressure to the separator.
    No iteration — deterministic in one pass.
    Returns a result dict compatible with _goal_seek_stack output.
    """
    dp_line, P_line_out = _calc_dp_at_p(res_line, P_source_bara)
    dp_hdr,  P_sep      = _calc_header_dp_at_p(res_hdr, P_line_out)
    return {
        "P_line_in":  P_source_bara,
        "P_line_out": P_line_out,
        "P_sep":      P_sep,
        "dp_line":    dp_line,
        "dp_hdr":     dp_hdr,
        "converged":  True,
        "iterations": 1,
    }


def _goal_seek_stack(res_line, res_hdr, P_target_sep, tol=0.0005, max_iter=30):
    """Find the line inlet pressure (= stack outlet) to achieve P_target_sep
    at the separator.  Flow: line → header arms → T-segment → separator.
    Returns a result dict.
    """
    P_line_in = (P_target_sep
                 + res_line["total_dp_kpa"] / 100.0
                 + res_hdr["total_dp_kpa"]  / 100.0)

    dp_line = dp_hdr = 0.0
    P_line_out = P_sep = 0.0

    for i in range(max_iter):
        dp_line, P_line_out = _calc_dp_at_p(res_line, P_line_in)
        dp_hdr,  P_sep      = _calc_header_dp_at_p(res_hdr, P_line_out)
        error = P_sep - P_target_sep
        if abs(error) < tol:
            return {
                "P_line_in":  P_line_in,  "P_line_out": P_line_out,
                "P_sep":      P_sep,
                "dp_line":    dp_line,    "dp_hdr":     dp_hdr,
                "iterations": i + 1,      "converged":  True,
            }
        P_line_in -= error
        P_line_in = max(0.1, P_line_in)  # prevent runaway below hard vacuum

    return {
        "P_line_in":  P_line_in,  "P_line_out": P_line_out,
        "P_sep":      P_sep,
        "dp_line":    dp_line,    "dp_hdr":     dp_hdr,
        "iterations": max_iter,
        "converged":  abs(P_sep - P_target_sep) < tol * 20,
    }


# ============================================================================
# HEADER PIPING SCHEMATIC
# ============================================================================
def _make_header_schematic(
        left_positions, right_positions, t_seg_spec,
        P_inlet_bara, worst_arm,
        dp_l_kpa, dp_r_kpa,
        P_T_l_bara, P_T_r_bara, P_sep_bara):
    """Return a Plotly figure showing the physical header piping layout."""

    PIPE_Y  = 1.8   # header pipe y-level
    TAP_TOP = 3.1   # top of tap risers (flow enters from branch above)
    TSEG_Y  = 0.5   # T-segment / separator level

    max_l  = max(left_positions,  default=0.5)
    max_r  = max(right_positions, default=0.5)
    t_len  = float(t_seg_spec.get("length", 1.0))
    pad    = max((max_l + max_r) * 0.06, 0.5)

    LEFT_C  = "#2563EB"
    RIGHT_C = "#D97706"
    TSEG_C  = "#475569"
    JOINT_C = "#1E293B"

    shapes = []

    # ── Header pipe: left arm ─────────────────────────────────────────────────
    x_left_end  = -max_l if left_positions  else -0.7
    x_right_end =  max_r if right_positions else  0.7
    shapes.append(dict(type="line",
        x0=x_left_end, y0=PIPE_Y, x1=0, y1=PIPE_Y,
        line=dict(color=LEFT_C if left_positions else "#CBD5E1", width=9)))

    # ── Header pipe: right arm ────────────────────────────────────────────────
    shapes.append(dict(type="line",
        x0=0, y0=PIPE_Y, x1=x_right_end, y1=PIPE_Y,
        line=dict(color=RIGHT_C if right_positions else "#CBD5E1", width=9)))

    # ── T-segment: vertical drop then horizontal to separator ─────────────────
    shapes.append(dict(type="line",
        x0=0, y0=PIPE_Y, x1=0, y1=TSEG_Y,
        line=dict(color=TSEG_C, width=14)))
    if t_len > 0:
        shapes.append(dict(type="line",
            x0=0, y0=TSEG_Y, x1=t_len, y1=TSEG_Y,
            line=dict(color=TSEG_C, width=14)))

    # ── Separator box ─────────────────────────────────────────────────────────
    sep_x = t_len if t_len > 0 else 0
    sep_w, sep_h = 1.3, 0.75
    shapes.append(dict(type="rect",
        x0=sep_x, y0=TSEG_Y - sep_h / 2, x1=sep_x + sep_w, y1=TSEG_Y + sep_h / 2,
        line=dict(color=JOINT_C, width=2), fillcolor="#DBEAFE"))

    # ── Tap risers ────────────────────────────────────────────────────────────
    left_sorted  = sorted(left_positions,  reverse=True)   # farthest first
    right_sorted = sorted(right_positions)                  # nearest first
    for pos in left_sorted:
        shapes.append(dict(type="line",
            x0=-pos, y0=PIPE_Y, x1=-pos, y1=TAP_TOP,
            line=dict(color=LEFT_C, width=3)))
    for pos in right_sorted:
        shapes.append(dict(type="line",
            x0=pos, y0=PIPE_Y, x1=pos, y1=TAP_TOP,
            line=dict(color=RIGHT_C, width=3)))

    fig = go.Figure()

    # ── Tap inlet markers (triangles pointing down = flow direction) ──────────
    tap_xs   = [-p for p in left_sorted] + list(right_sorted)
    tap_cs   = [LEFT_C] * len(left_sorted) + [RIGHT_C] * len(right_sorted)
    tap_lbls = (
        [f"L{len(left_sorted) - i}<br>{p:.1f} m" for i, p in enumerate(left_sorted)] +
        [f"R{i + 1}<br>{p:.1f} m"                for i, p in enumerate(right_sorted)]
    )
    if tap_xs:
        fig.add_trace(go.Scatter(
            x=tap_xs, y=[TAP_TOP] * len(tap_xs),
            mode="markers+text",
            marker=dict(symbol="triangle-down", size=14, color=tap_cs,
                        line=dict(color="white", width=1)),
            text=tap_lbls,
            textposition="top center",
            textfont=dict(size=9),
            hoverinfo="skip", showlegend=False,
        ))

    # ── T-junction marker ─────────────────────────────────────────────────────
    fig.add_trace(go.Scatter(
        x=[0], y=[PIPE_Y],
        mode="markers+text",
        marker=dict(symbol="circle", size=20, color=JOINT_C,
                    line=dict(color="white", width=2)),
        text=["<b>T</b>"],
        textposition="middle center",
        textfont=dict(size=11, color="white"),
        hoverinfo="skip", showlegend=False,
    ))

    # ── Annotations ───────────────────────────────────────────────────────────
    anns = []

    # Tap inlet pressure banner (top-left)
    anns.append(dict(
        x=x_left_end, y=TAP_TOP + 0.55,
        text=f"<b>Tap inlet pressure: {P_inlet_bara:.3f} bara</b>",
        showarrow=False, xanchor="left",
        font=dict(size=10, color="#1E293B"),
        xref="x", yref="y",
    ))

    # Left arm: ΔP label + flow arrow
    if left_positions:
        w_mark = "  ⚠ governing" if worst_arm == "Left" else ""
        anns.append(dict(
            x=-max_l * 0.5, y=PIPE_Y - 0.35,
            text=f"ΔP = {dp_l_kpa:.2f} kPa{w_mark}",
            showarrow=False, xanchor="center",
            font=dict(size=9, color=LEFT_C),
            xref="x", yref="y",
        ))
        anns.append(dict(
            x=-max_l * 0.18, y=PIPE_Y + 0.28,
            ax=-max_l * 0.72, ay=PIPE_Y + 0.28,
            xref="x", yref="y", axref="x", ayref="y",
            showarrow=True, arrowhead=3, arrowsize=1.3,
            arrowwidth=2, arrowcolor=LEFT_C,
        ))
        anns.append(dict(
            x=x_left_end + 0.1, y=PIPE_Y + 0.28,
            text="<i>← Left arm</i>",
            showarrow=False, xanchor="left",
            font=dict(size=9, color=LEFT_C),
            xref="x", yref="y",
        ))

    # Right arm: ΔP label + flow arrow
    if right_positions:
        w_mark = "  ⚠ governing" if worst_arm == "Right" else ""
        anns.append(dict(
            x=max_r * 0.5, y=PIPE_Y - 0.35,
            text=f"ΔP = {dp_r_kpa:.2f} kPa{w_mark}",
            showarrow=False, xanchor="center",
            font=dict(size=9, color=RIGHT_C),
            xref="x", yref="y",
        ))
        anns.append(dict(
            x=max_r * 0.18, y=PIPE_Y + 0.28,
            ax=max_r * 0.72, ay=PIPE_Y + 0.28,
            xref="x", yref="y", axref="x", ayref="y",
            showarrow=True, arrowhead=3, arrowsize=1.3,
            arrowwidth=2, arrowcolor=RIGHT_C,
        ))
        anns.append(dict(
            x=x_right_end - 0.1, y=PIPE_Y + 0.28,
            text="<i>Right arm →</i>",
            showarrow=False, xanchor="right",
            font=dict(size=9, color=RIGHT_C),
            xref="x", yref="y",
        ))

    # T-segment: pressure at T and flow arrow down
    anns.append(dict(
        x=0.25, y=(PIPE_Y + TSEG_Y) / 2,
        text=f"P_T ≈ {min(P_T_l_bara, P_T_r_bara):.3f} bara",
        showarrow=False, xanchor="left",
        font=dict(size=9, color=TSEG_C),
        xref="x", yref="y",
    ))
    anns.append(dict(
        x=0, y=TSEG_Y + 0.18,
        ax=0, ay=PIPE_Y - 0.18,
        xref="x", yref="y", axref="x", ayref="y",
        showarrow=True, arrowhead=3, arrowsize=1.2,
        arrowwidth=2, arrowcolor=TSEG_C,
    ))
    if t_len > 0:
        anns.append(dict(
            x=t_len * 0.5, y=TSEG_Y + 0.22,
            text=f"T-seg  {t_len:.1f} m",
            showarrow=False, xanchor="center",
            font=dict(size=9, color=TSEG_C),
            xref="x", yref="y",
        ))

    # Separator label
    anns.append(dict(
        x=sep_x + sep_w / 2, y=TSEG_Y,
        text=f"<b>SEP</b><br>{P_sep_bara:.3f} bara",
        showarrow=False, xanchor="center",
        font=dict(size=9, color="#1E40AF"),
        xref="x", yref="y",
    ))

    x_lo = x_left_end  - pad - 4.0   # extra room for inlet pressure label
    x_hi = sep_x + sep_w + pad + 0.5

    fig.update_layout(
        shapes=shapes,
        annotations=anns,
        showlegend=False,
        height=340,
        margin=dict(l=10, r=10, t=35, b=10),
        xaxis=dict(visible=False, range=[x_lo, x_hi]),
        yaxis=dict(visible=False, range=[-0.3, TAP_TOP + 1.4]),
        paper_bgcolor="white",
        plot_bgcolor="white",
        title=dict(text="Header piping layout", font=dict(size=13, color="#1E293B"), x=0.5),
    )
    return fig


# ============================================================================
# HEADER CASE RUNNER  (Case C — uniform header with n A-line taps on each side)
# ============================================================================
def run_header_case(cid: str = "c", accent: str = "#059669",
                    results_a: dict = None) -> dict:
    """
    Uniform header with n A-line taps on each side of the T-junction.
    Flow enters from n_left + n_right copies of Case A and exits at T.
    Tap positions (distance from T) are individually configurable.
    """
    k = lambda name: f"{cid}_{name}"

    # ── Flow source: read from results_a, fall back to session state defaults
    _ra          = results_a or {}
    gas_per_tap  = dict(_ra.get("gas_flows_kgh") or {"H₂": 5.0})
    liq_per_tap  = float(_ra.get("q_lye") or 0.5)
    liquid_type  = str(_ra.get("liquid_type") or "Custom")
    custom_liquid_hdr = _ra.get("custom_liquid")   # None for single-species CoolProp
    T_C_a        = float(_ra.get("T_C") or 60.0)

    col_in, col_out = st.columns([1, 1.2])

    with col_in:
        st.subheader("Inputs")

        with st.container(border=True):
            st.markdown("**Fluid — from Case A (read-only)**")
            _gas_str = "  ·  ".join(f"{sp}: {v:.2f} kg/h" for sp, v in gas_per_tap.items())
            st.caption(f"Gas per tap:  {_gas_str}")
            st.caption(f"Liquid per tap:  {liq_per_tap:.3f} m³/h  ·  {liquid_type or 'none'}")

        with st.container(border=True):
            st.markdown("**Process Conditions**")
            p1, p2 = st.columns(2)
            P_target_sep = p1.number_input(
                "Target separator pressure (bara)",
                min_value=1.0, max_value=200.0,
                value=float(st.session_state.get(k("P_target_sep"), 16.5)),
                step=0.5, format="%.2f", key=k("P_target_sep"),
                help="Pressure required at the T-junction / separator connection. "
                     "The required tap inlet pressure is found automatically.")
            T_C = p2.number_input("Temperature (°C)",
                                  min_value=5.0, max_value=95.0,
                                  value=float(st.session_state.get(k("T_C"), T_C_a)),
                                  step=5.0, key=k("T_C"))

        with st.container(border=True):
            st.markdown("**Header Pipe** — uniform along full length")
            DN_OPT  = list(engine.PIPE_DATABASE.keys())
            PN_OPT  = ["PN20", "PN25", "PN40"]
            MAT_OPT = list(engine.MATERIAL_ROUGHNESS.keys())
            h1, h2, h3 = st.columns(3)
            hdr_dn  = h1.selectbox("DN",       DN_OPT,  key=k("hdr_dn"),
                                   index=DN_OPT.index(st.session_state.get(k("hdr_dn"), "DN100"))
                                         if st.session_state.get(k("hdr_dn"), "DN100") in DN_OPT else 0)
            hdr_pn  = h2.selectbox("PN",       PN_OPT,  key=k("hdr_pn"),
                                   index=PN_OPT.index(st.session_state.get(k("hdr_pn"), "PN40"))
                                         if st.session_state.get(k("hdr_pn"), "PN40") in PN_OPT else 2)
            hdr_mat = h3.selectbox("Material", MAT_OPT, key=k("hdr_mat"),
                                   index=MAT_OPT.index(st.session_state.get(k("hdr_mat"), "SS316L"))
                                         if st.session_state.get(k("hdr_mat"), "SS316L") in MAT_OPT else 0)
            h4, h5, h6 = st.columns(3)
            hdr_lined = h4.checkbox("Lined", value=bool(st.session_state.get(k("hdr_lined"), False)),
                                    key=k("hdr_lined"))
            LINER_OPT = list(engine.LINER_ROUGHNESS.keys())
            if hdr_lined:
                hdr_lmat  = h5.selectbox("Liner", LINER_OPT, key=k("hdr_lmat"))
                hdr_lthk  = h6.number_input("Thickness (mm)", min_value=0.1, max_value=20.0,
                                             value=float(st.session_state.get(k("hdr_lthk"), 1.0)),
                                             step=0.5, key=k("hdr_lthk"))
            else:
                hdr_lmat = "FEP";  hdr_lthk = 1.0
            _D_nom = engine.PIPE_DATABASE[hdr_dn][hdr_pn]
            _D_eff = _D_nom - 2 * hdr_lthk / 1000.0 if hdr_lined else _D_nom
            st.caption(f"ID {_D_eff*1000:.1f} mm")
            st.caption("**Header fittings**" if st.session_state.get(k("hdr_fits"), []) else "Header fittings — none")
            _hdr_fk = k("hdr_fits")
            if _hdr_fk not in st.session_state:
                st.session_state[_hdr_fk] = []
            _hdr_fit_rows  = st.session_state[_hdr_fk]
            _hdr_n_fits    = len(_hdr_fit_rows)
            _all_fit_types = list(engine.FITTING_Le_over_D.keys())
            if st.button("+ Add fitting", key=k("hdr_fadd"), help="Add a fitting to the header"):
                for jj in range(_hdr_n_fits):
                    _ft = st.session_state.get(k(f"hdr_ftype_{jj}"))
                    _fq = st.session_state.get(k(f"hdr_fqty_{jj}"), 1)
                    if _ft:
                        st.session_state[_hdr_fk][jj]["type"] = _ft
                        st.session_state[_hdr_fk][jj]["qty"]  = int(_fq)
                st.session_state[_hdr_fk].append({"type": _all_fit_types[0], "qty": 1})
                st.rerun()
            for j, _hfr in enumerate(list(_hdr_fit_rows)):
                _hfc1, _hfc2, _hfc3 = st.columns([3.5, 0.7, 0.3])
                _hfc1.selectbox("Fitting", _all_fit_types,
                    index=_all_fit_types.index(_hfr["type"]) if _hfr["type"] in _all_fit_types else 0,
                    key=k(f"hdr_ftype_{j}"), label_visibility="collapsed")
                _hfc2.number_input("Qty", min_value=1, value=_hfr["qty"],
                    key=k(f"hdr_fqty_{j}"), label_visibility="collapsed")
                if _hfc3.button("×", key=k(f"hdr_frem_{j}"), help="Remove"):
                    for jj in range(_hdr_n_fits):
                        _ft = st.session_state.get(k(f"hdr_ftype_{jj}"))
                        _fq = st.session_state.get(k(f"hdr_fqty_{jj}"), 1)
                        if _ft:
                            st.session_state[_hdr_fk][jj]["type"] = _ft
                            st.session_state[_hdr_fk][jj]["qty"]  = int(_fq)
                    st.session_state[_hdr_fk].pop(j)
                    for jj in range(j, _hdr_n_fits - 1):
                        st.session_state[k(f"hdr_ftype_{jj}")] = st.session_state[_hdr_fk][jj]["type"]
                        st.session_state[k(f"hdr_fqty_{jj}")] = st.session_state[_hdr_fk][jj]["qty"]
                    st.session_state.pop(k(f"hdr_ftype_{_hdr_n_fits-1}"), None)
                    st.session_state.pop(k(f"hdr_fqty_{_hdr_n_fits-1}"), None)
                    st.rerun()

        _hdr_fits_list = []
        _hdr_fk = k("hdr_fits")
        for jj in range(len(st.session_state.get(_hdr_fk, []))):
            _ft = st.session_state.get(k(f"hdr_ftype_{jj}"))
            _fq = st.session_state.get(k(f"hdr_fqty_{jj}"), 1)
            if _ft and _ft in engine.FITTING_Le_over_D and int(_fq) > 0:
                _hdr_fits_list.append({"type": _ft, "qty": int(_fq)})
        hdr_spec = {
            "dn": hdr_dn, "pn": hdr_pn, "material": hdr_mat,
            "lined": hdr_lined, "liner_material": hdr_lmat, "liner_thickness_mm": hdr_lthk,
            "fittings_list": _hdr_fits_list,
        }

        with st.container(border=True):
            st.markdown("**Calculation Settings**")
            _cs1, _cs2 = st.columns(2)
            correlation    = _cs1.selectbox("ΔP correlation",
                                            engine.TWO_PHASE_CORRELATIONS, key=k("correlation"))
            voidage_method = _cs2.selectbox("Void fraction",
                                            engine.VOIDAGE_METHODS, key=k("voidage_method"))

        # ── Tap position editor ────────────────────────────────────────────────
        def _tap_editor(side, default_n=3, default_spacing=2.5):
            """Render n + position inputs for one arm. Returns list of distances (m)."""
            n = st.number_input(f"Taps — {side} arm", min_value=0, max_value=8,
                                value=int(st.session_state.get(k(f"n_{side}"), default_n)),
                                step=1, key=k(f"n_{side}"))
            if n == 0:
                return []
            # Show positions in rows of 4
            positions = []
            cols_per_row = 4
            _rows = [range(i, min(i + cols_per_row, n))
                     for i in range(0, n, cols_per_row)]
            for _row_idxs in _rows:
                _cols = st.columns(len(_row_idxs))
                for _ci, _ti in enumerate(_row_idxs):
                    _pk = k(f"pos_{side}_{_ti}")
                    if _pk not in st.session_state:
                        st.session_state[_pk] = float((_ti + 1) * default_spacing)
                    positions.append(_cols[_ci].number_input(
                        f"{side[0].upper()}{_ti+1} (m)", min_value=0.1, step=0.5,
                        key=_pk))
            return positions

        st.markdown("---")
        st.markdown("#### Tap Positions  *(distance from T-junction)*")
        _diag_left  = st.session_state.get(k("n_left"), 3)
        _diag_right = st.session_state.get(k("n_right"), 3)
        st.caption(
            "  ".join([f"←─ L{i}" for i in range(int(_diag_left), 0, -1)])
            + "  ──T──  "
            + "  ".join([f"R{i} ─→" for i in range(1, int(_diag_right) + 1)])
        )
        lc, rc = st.columns(2)
        with lc:
            st.markdown("**← Left arm**")
            left_positions  = _tap_editor("left")
        with rc:
            st.markdown("**Right arm →**")
            right_positions = _tap_editor("right")

        with st.container(border=True):
            st.markdown("**T-Segment** *(at junction — may be larger than header)*")
            st.caption("Carries combined flow from all taps. Horizontal. End = separator connection.")
            ts1, ts2, ts3, ts4 = st.columns(4)
            t_dn  = ts1.selectbox("DN", DN_OPT, key=k("t_dn"),
                                   index=DN_OPT.index(st.session_state.get(k("t_dn"), "DN150"))
                                         if st.session_state.get(k("t_dn"), "DN150") in DN_OPT else 0)
            t_pn  = ts2.selectbox("PN", PN_OPT, key=k("t_pn"),
                                   index=PN_OPT.index(st.session_state.get(k("t_pn"), "PN40"))
                                         if st.session_state.get(k("t_pn"), "PN40") in PN_OPT else 2)
            t_mat = ts3.selectbox("Material", MAT_OPT, key=k("t_mat"),
                                   index=MAT_OPT.index(st.session_state.get(k("t_mat"), "SS316L"))
                                         if st.session_state.get(k("t_mat"), "SS316L") in MAT_OPT else 0)
            t_len = ts4.number_input("Length (m)", min_value=0.0,
                                      value=float(st.session_state.get(k("t_len"), 1.0)),
                                      step=0.5, key=k("t_len"))
            _t_D = engine.PIPE_DATABASE[t_dn][t_pn]
            st.caption(f"ID {_t_D*1000:.1f} mm")

        t_seg_spec = {"dn": t_dn, "pn": t_pn, "material": t_mat,
                      "length": t_len, "fittings_list": []}

    # ── CALCULATION ───────────────────────────────────────────────────────────
    with col_out:
        st.subheader("Results")

        n_left  = len(left_positions)
        n_right = len(right_positions)
        n_total = n_left + n_right
        total_gas = {sp: kgh * n_total for sp, kgh in gas_per_tap.items()}
        total_liq = liq_per_tap * n_total

        # ── Goal-seek: find tap inlet pressure for target separator pressure ──
        _hdr_for_gs = {
            "header_pipe":    hdr_spec,
            "gas_per_tap":    gas_per_tap,
            "liq_per_tap":    liq_per_tap,
            "T_C":            T_C,
            "liquid_type":    liquid_type,
            "custom_liquid":  custom_liquid_hdr,
            "left_taps":      left_positions,
            "right_taps":     right_positions,
            "t_seg":          t_seg_spec,
            "correlation":    correlation,
            "voidage_method": voidage_method,
            "gas_flows_kgh":  total_gas,
            "q_lye":          total_liq,
            "total_dp_kpa":   0.5,   # coarse starting offset; converges regardless
        }
        if left_positions or right_positions:
            _gs = _goal_seek_header(_hdr_for_gs, P_target_sep)
            P_inlet_bara = _gs["P_hdr_in"]
            _gs_resid    = abs(_gs["P_sep"] - P_target_sep) * 1000
            if _gs["converged"]:
                st.success(
                    f"Goal-seek converged in {_gs['iterations']} iter  ·  "
                    f"residual {_gs_resid:.2f} mbar"
                )
            else:
                st.warning(
                    f"Goal-seek did not fully converge ({_gs['iterations']} iter, "
                    f"residual {_gs_resid:.1f} mbar)"
                )
        else:
            P_inlet_bara = P_target_sep
            st.info("Add taps to compute the required inlet pressure.")

        # ── Forward march at the found inlet pressure ─────────────────────────
        P_start = P_inlet_bara * 1e5
        dp_l_Pa, P_T_l, fric_l, grav_l, rec_l = _march_header_simple(
            left_positions,  gas_per_tap, liq_per_tap,
            hdr_spec, P_start, T_C, liquid_type, correlation, voidage_method,
            custom_liquid=custom_liquid_hdr)
        dp_r_Pa, P_T_r, fric_r, grav_r, rec_r = _march_header_simple(
            right_positions, gas_per_tap, liq_per_tap,
            hdr_spec, P_start, T_C, liquid_type, correlation, voidage_method,
            custom_liquid=custom_liquid_hdr)

        dp_l_kpa   = dp_l_Pa / 1000.0
        dp_r_kpa   = dp_r_Pa / 1000.0
        P_T_l_bara = P_T_l / 1e5
        P_T_r_bara = P_T_r / 1e5
        worst_arm  = "Left" if dp_l_kpa >= dp_r_kpa else "Right"
        dp_worst   = max(dp_l_kpa, dp_r_kpa)
        P_T_worst  = min(P_T_l_bara, P_T_r_bara)

        # T-segment calculation
        dp_t_kpa   = 0.0
        P_sep_bara = P_T_worst
        rec_t      = None
        if t_seg_spec["length"] > 0 and total_gas and total_liq > 0:
            try:
                dp_t_Pa, P_sep, fric_t, grav_t, rec_t = _march_single_seg(
                    t_seg_spec, P_T_worst * 1e5, T_C, total_gas, liquid_type,
                    total_liq, correlation, voidage_method,
                    custom_liquid=custom_liquid_hdr)
                dp_t_kpa   = dp_t_Pa / 1000.0
                P_sep_bara = P_sep / 1e5
            except Exception as _e:
                st.warning(f"T-segment calculation failed: {_e}")

        with st.container(border=True):
            st.subheader("T-Junction")
            _c1, _c2, _c3, _c4 = st.columns(4)
            _c1.metric("Required tap inlet pressure",
                       f"{P_inlet_bara:.4f} bara",
                       delta=f"Target sep: {P_target_sep:.2f} bara",
                       delta_color="off",
                       help="Pressure required at each header tap = branch line outlet pressure")
            _c2.metric("Left arm ΔP",
                       f"{dp_l_kpa:.3f} kPa",
                       delta=f"{P_inlet_bara:.4f} → {P_T_l_bara:.4f} bara",
                       delta_color="off")
            _c3.metric("Right arm ΔP",
                       f"{dp_r_kpa:.3f} kPa",
                       delta=f"{P_inlet_bara:.4f} → {P_T_r_bara:.4f} bara",
                       delta_color="off")
            _c4.metric("Separator connection pressure",
                       f"{P_sep_bara:.4f} bara",
                       delta=f"T-seg ΔP: {dp_t_kpa:.3f} kPa",
                       delta_color="off")
            st.caption(
                f"Total flow at T:  "
                + "  ·  ".join(f"{sp} {v:.2f} kg/h" for sp, v in total_gas.items())
                + f"  ·  Liquid {total_liq:.3f} m³/h"
                + f"  ·  ({n_left}L + {n_right}R = {n_total} taps)"
            )

        # ── Piping schematic ──────────────────────────────────────────────────
        fig_sch_hdr = _make_header_schematic(
            left_positions, right_positions, t_seg_spec,
            P_inlet_bara, worst_arm,
            dp_l_kpa, dp_r_kpa,
            P_T_l_bara, P_T_r_bara, P_sep_bara)
        st.plotly_chart(fig_sch_hdr, width='stretch',
                        key=f"{cid}_hdr_schematic")

        _col_hdr = ["Seg", "Taps in seg", "From T (m)", "To T (m)", "L (m)",
                    "Pipe", "ID (mm)", "Regime",
                    "Q_gas_kgh", "Q_liq_m3h",
                    "ΔP_fric (kPa)", "ΔP_grav (kPa)", "ΔP (kPa)",
                    "P_in (bara)", "P_out (bara)", "V_m (m/s)", "V_m/V_e"]
        _col_cfg_hdr = {
            "ID (mm)":       st.column_config.NumberColumn(format="%.1f"),
            "From T (m)":    st.column_config.NumberColumn(format="%.2f"),
            "To T (m)":      st.column_config.NumberColumn(format="%.2f"),
            "Q_gas_kgh":     st.column_config.NumberColumn(label="Q gas (kg/h)", format="%.3f"),
            "Q_liq_m3h":     st.column_config.NumberColumn(label="Q liq (m³/h)", format="%.3f"),
            "P_in (bara)":   st.column_config.NumberColumn(format="%.4f"),
            "P_out (bara)":  st.column_config.NumberColumn(format="%.4f"),
            "ΔP (kPa)":      st.column_config.NumberColumn(format="%.3f"),
            "ΔP_fric (kPa)": st.column_config.NumberColumn(format="%.3f"),
            "ΔP_grav (kPa)": st.column_config.NumberColumn(format="%.3f"),
            "V_m (m/s)":     st.column_config.NumberColumn(format="%.3f"),
            "V_m/V_e":       st.column_config.NumberColumn(format="%.3f"),
        }
        tl, tr = st.columns(2)
        with tl:
            st.markdown(f"**Left arm**  {'⚠️ governing' if worst_arm == 'Left' else ''}")
            if rec_l:
                _df_l = pd.DataFrame(rec_l)
                st.dataframe(_df_l[[c for c in _col_hdr if c in _df_l.columns]],
                             column_config=_col_cfg_hdr,
                             hide_index=True, width='stretch')
        with tr:
            st.markdown(f"**Right arm**  {'⚠️ governing' if worst_arm == 'Right' else ''}")
            if rec_r:
                _df_r = pd.DataFrame(rec_r)
                st.dataframe(_df_r[[c for c in _col_hdr if c in _df_r.columns]],
                             column_config=_col_cfg_hdr,
                             hide_index=True, width='stretch')

        if rec_t:
            st.markdown("**T-segment** (junction → separator)")
            _df_t = pd.DataFrame([rec_t])
            _t_cols = [c for c in _col_hdr if c in _df_t.columns]
            st.dataframe(_df_t[_t_cols], column_config=_col_cfg_hdr,
                         hide_index=True, width='stretch')

        fig_hdr = None
        if rec_l or rec_r:
            fig_hdr = go.Figure()
            def _arm_trace(records, label, color, positions):
                if not records:
                    return
                dists_sorted = sorted(positions, reverse=True)
                xs = [dists_sorted[0]] if dists_sorted else [0.0]
                ys = [P_inlet_bara]
                for r in records:
                    xs.append(r["To T (m)"])
                    ys.append(r["P_out (bara)"])
                fig_hdr.add_trace(go.Scatter(
                    x=xs, y=ys, mode="lines+markers", name=label,
                    line=dict(color=color, width=2), marker=dict(size=6)))
            _arm_trace(rec_l, "Left arm",  "#2563EB", left_positions)
            _arm_trace(rec_r, "Right arm", "#D97706", right_positions)
            fig_hdr.update_layout(
                xaxis_title="Distance from T-junction (m)",
                xaxis=dict(autorange="reversed"),
                yaxis_title="Pressure (bara)",
                height=300, margin=dict(l=40, r=20, t=30, b=40),
                legend=dict(orientation="h", y=1.1),
                template="plotly_white",
                paper_bgcolor="white", plot_bgcolor="white",
                xaxis_gridcolor="#F1F5F9", yaxis_gridcolor="#F1F5F9",
            )
            st.plotly_chart(fig_hdr, width='stretch',
                            key=f"{cid}_hdr_pressure_profile")

    # ── RETURN DICT ────────────────────────────────────────────────────────────
    try:
        _props_out = engine.calculate_two_phase_properties(
            P_T_worst, T_C, total_gas, liquid_type, total_liq) if total_liq > 0 and total_gas else {}
    except Exception:
        _props_out = {}

    _all_recs  = ([dict(r, Seg=f"L{r['Seg']}") for r in rec_l] +
                  [dict(r, Seg=f"R{r['Seg']}") for r in rec_r])
    _pipe_len  = (max(left_positions,  default=[0.0]) +
                  max(right_positions, default=[0.0]))
    _total_dp_fric = (fric_l + fric_r) / 1000.0
    _total_dp_grav = (grav_l + grav_r) / 1000.0

    return {
        "P_bara":               P_inlet_bara,
        "P_target_sep":         P_target_sep,
        "T_C":                  T_C,
        "total_dp_kpa":         dp_worst + dp_t_kpa,
        "outlet_pressure_bara": P_sep_bara,
        "outlet_pressure_mbar": P_sep_bara * 1000.0,
        "P_T_bara":             P_T_worst,
        "P_separator_bara":     P_sep_bara,
        "dp_header_kpa":        dp_worst,
        "dp_t_seg_kpa":         dp_t_kpa,
        "total_dp_fric_kpa":    _total_dp_fric,
        "total_dp_grav_kpa":    _total_dp_grav,
        "pipe_length_m":        _pipe_len,
        "cumulative_distance":  _pipe_len,
        "liquid_type":          liquid_type,
        "gas_flows_kgh":        total_gas,
        "q_lye":                total_liq,
        "props":                _props_out,
        "segments":             [],
        "grid_records":         _all_recs,
        "correlation":          correlation,
        "voidage_method":       voidage_method,
        "fig_sch":              fig_sch_hdr,
        "fig_prof":             fig_hdr,
        # Fields for goal-seek re-run
        "left_taps":            left_positions,
        "right_taps":           right_positions,
        "gas_per_tap":          gas_per_tap,
        "liq_per_tap":          liq_per_tap,
        "header_pipe":          hdr_spec,
        "t_seg":                t_seg_spec,
        # Per-arm detail
        "dp_left_kpa":          dp_l_kpa,
        "dp_right_kpa":         dp_r_kpa,
        "P_T_left_bara":        P_T_l_bara,
        "P_T_right_bara":       P_T_r_bara,
        "worst_arm":            worst_arm,
        "n_left":               n_left,
        "n_right":              n_right,
    }


if "label_a" not in st.session_state:
    st.session_state["label_a"] = "A"
if "label_b" not in st.session_state:
    st.session_state["label_b"] = "B"

# Forward any pending goal-seek apply values BEFORE widgets render
if "stack_apply_a_pending" in st.session_state:
    st.session_state["a_P_bara"] = max(0.1, st.session_state.pop("stack_apply_a_pending"))
if "stack_apply_b_pending" in st.session_state:
    st.session_state["b_P_bara"] = max(0.1, st.session_state.pop("stack_apply_b_pending"))

_la = st.session_state.get("label_a") or "A"
_lb = st.session_state.get("label_b") or "B"
_lc = f"Header {_la}"
_ld = f"Header {_lb}"

_group = st.segmented_control(
    "Workspace", ["Pipeline Cases", "Engineering Tools"],
    default="Pipeline Cases", label_visibility="collapsed",
    key="main_group",
)

if _group != "Engineering Tools":
    tab_a, tab_b, tab_c, tab_d, tab_cmp, tab_gs = st.tabs(
        [_la, _lb, _lc, _ld, "Compare", "Goal Seek"])

    _CORR_SHORT = {
        "Beggs-Brill": "BB", "Friedel": "Friedel",
        "Lockhart_Martinelli": "L-M", "Muller_Steinhagen_Heck": "MSH",
        "Chisholm": "Chisholm", "Kim_Mudawar": "Kim-M",
    }
    _VOID_SHORT = {
        "Homogeneous": "Homo",
        "Rouhani-1 (slip)": "Rouhani-1",
    }

    with tab_a:
        results_a = run_case("a", accent="#2563EB")

    with tab_b:
        results_b = run_case("b", accent="#D97706")

    with tab_c:
        st.info(
            f"**{_lc}**  "
            f"A uniform pipe with n {_la} taps on each side of a central T-junction. "
            f"Each tap feeds one copy of {_la}'s branch flow. "
            "Worst-arm ΔP = farthest tap → T → separator.",
            icon="ℹ️",
        )
        results_c = run_header_case("c", accent="#059669", results_a=results_a)

    with tab_d:
        st.info(
            f"**{_ld}**  "
            f"A uniform pipe with n {_lb} taps on each side of a central T-junction. "
            f"Each tap feeds one copy of {_lb}'s branch flow. "
            "Worst-arm ΔP = farthest tap → T → separator.",
            icon="ℹ️",
        )
        results_d = run_header_case("d", accent="#7C3AED", results_a=results_b)

    # ============================================================================
    # COMPARE TAB
    # ============================================================================

    def _sens_hash(ra, rb):
        """Stable hash of all inputs that determine sensitivity results."""
        _data = {
            "a": {
                "P": ra["P_bara"], "T": ra["T_C"],
                "gas": {_k: float(_v) for _k, _v in ra["gas_flows_kgh"].items()},
                "liq": ra["liquid_type"], "lye": ra["q_lye"],
                "segs": [(s.get("type", s.get("kind","")), s.get("dn",""), s.get("pn",""), float(s.get("length",0.0)),
                          tuple(sorted((f["type"],f["qty"]) for f in s.get("fittings_list",[]))),
                          bool(s.get("lined", False)), s.get("liner_material", "FEP"),
                          float(s.get("liner_thickness_mm", 1.0)))
                         for s in ra["segments"]],
            },
            "b": {
                "P": rb["P_bara"], "T": rb["T_C"],
                "gas": {_k: float(_v) for _k, _v in rb["gas_flows_kgh"].items()},
                "liq": rb["liquid_type"], "lye": rb["q_lye"],
                "segs": [(s.get("type", s.get("kind","")), s.get("dn",""), s.get("pn",""), float(s.get("length",0.0)),
                          tuple(sorted((f["type"],f["qty"]) for f in s.get("fittings_list",[]))),
                          bool(s.get("lined", False)), s.get("liner_material", "FEP"),
                          float(s.get("liner_thickness_mm", 1.0)))
                         for s in rb["segments"]],
            },
        }
        return hashlib.md5(json.dumps(_data, sort_keys=True).encode()).hexdigest()


    with tab_cmp:
        _la = st.session_state.get("label_a") or "A"
        _lb = st.session_state.get("label_b") or "B"
        st.subheader(f"{_la}  vs.  {_lb}")

        ra, rb = results_a, results_b

        # ── Side-by-side headline metrics ─────────────────────────────────────────
        with st.container(border=True):
            _cl, _cm, _cr = st.columns([2, 2, 2])

            _cl.markdown("**Metric**")
            _cm.markdown(f"**{_la}**")
            _cr.markdown(f"**{_lb}**")
            st.divider()

            def _cmp_row(label, va, vb, fmt="{}", better="lower", unit=""):
                """Render one comparison row with delta badge."""
                _cl, _cm, _cr = st.columns([2, 2, 2])
                try:
                    _delta = vb - va
                    _pct   = (_delta / abs(va) * 100) if abs(va) > 1e-9 else 0.0
                    _sign  = "+" if _delta > 0 else ""
                    _delta_str = f"{_sign}{fmt.format(_delta)} {unit}  ({_sign}{_pct:.1f}%)"
                    if better == "lower":
                        _color = "normal" if _delta > 1e-9 else ("inverse" if _delta < -1e-9 else "off")
                    elif better == "higher":
                        _color = "inverse" if _delta > 1e-9 else ("normal" if _delta < -1e-9 else "off")
                    else:  # neutral — show delta magnitude without colour judgement
                        _color = "off"
                except Exception:
                    _delta_str = "—"
                    _color = "off"
                _cl.markdown(label)
                _cm.metric(_la, f"{fmt.format(va)} {unit}", label_visibility="collapsed")
                _cr.metric(_lb, f"{fmt.format(vb)} {unit}", delta=_delta_str,
                           delta_color=_color, label_visibility="collapsed")

            _cmp_row("Inlet pressure",        ra["P_bara"],               rb["P_bara"],               fmt="{:.2f}", unit="bara", better="neutral")
            _cmp_row("Outlet pressure",       ra["outlet_pressure_bara"], rb["outlet_pressure_bara"], fmt="{:.4f}", unit="bara", better="higher")
            _cmp_row("Total ΔP",              ra["total_dp_kpa"],         rb["total_dp_kpa"],         fmt="{:.3f}", unit="kPa",  better="lower")
            _cmp_row("  ↳ Frictional",        ra["total_dp_fric_kpa"],    rb["total_dp_fric_kpa"],    fmt="{:.3f}", unit="kPa",  better="lower")
            _cmp_row("  ↳ Gravitational",     ra["total_dp_grav_kpa"],    rb["total_dp_grav_kpa"],    fmt="{:.3f}", unit="kPa",  better="neutral")
            _cmp_row("Pipe length",           ra["pipe_length_m"],        rb["pipe_length_m"],        fmt="{:.1f}", unit="m",    better="lower")
            _cmp_row("Effective length",      ra["cumulative_distance"],  rb["cumulative_distance"],  fmt="{:.1f}", unit="m",    better="lower")
            _max_a = max((r["V_m/V_e"] for r in ra["grid_records"]), default=0.0)
            _max_b = max((r["V_m/V_e"] for r in rb["grid_records"]), default=0.0)
            _cmp_row("Worst erosion ratio V_m/V_e",  _max_a, _max_b,
                     fmt="{:.3f}", unit="— (limit 1.0)", better="lower")

        # ── Overlaid pressure profiles ─────────────────────────────────────────────
        st.markdown("#### Pressure Profiles")
        fig_cmp = go.Figure()
        fig_cmp.add_trace(go.Scatter(
            x=ra["pressure_profile_x"], y=ra["pressure_profile_y"],
            mode="lines+markers", name=_la,
            line=dict(color="#2563EB", width=2.5), marker=dict(size=7, color="#2563EB"),
            hovertemplate=f"{_la}  |  Distance: %{{x:.2f}} m<br>Pressure: %{{y:.4f}} bara<extra></extra>"))
        fig_cmp.add_trace(go.Scatter(
            x=rb["pressure_profile_x"], y=rb["pressure_profile_y"],
            mode="lines+markers", name=_lb,
            line=dict(color="#D97706", width=2.5, dash="dash"), marker=dict(size=7, color="#D97706"),
            hovertemplate=f"{_lb}  |  Distance: %{{x:.2f}} m<br>Pressure: %{{y:.4f}} bara<extra></extra>"))
        fig_cmp.update_layout(
            xaxis_title="Pipeline Distance (m)", yaxis_title="Pressure (bara)",
            template="plotly_white", height=360, margin=dict(l=60,r=20,t=30,b=50),
            hovermode="x unified", paper_bgcolor="white", plot_bgcolor="white",
            xaxis=dict(gridcolor="#F1F5F9", linecolor="#E2E8F0"),
            yaxis=dict(gridcolor="#F1F5F9", linecolor="#E2E8F0"),
            legend=dict(bgcolor="rgba(255,255,255,0.9)", bordercolor="#E2E8F0", borderwidth=1),
            font=dict(size=12, color="#374151"))
        st.plotly_chart(fig_cmp, width='stretch', key="fig_cmp")

        # ── Per-segment ΔP comparison ──────────────────────────────────────────────
        st.markdown("#### Pressure Drop by Segment")
        _segs_a = [f"A-{r['Seg']} {r['Pipe']}" for r in ra["grid_records"]]
        _segs_b = [f"B-{r['Seg']} {r['Pipe']}" for r in rb["grid_records"]]
        _dp_a   = [r["ΔP (kPa)"] for r in ra["grid_records"]]
        _dp_b   = [r["ΔP (kPa)"] for r in rb["grid_records"]]
        fig_bar = go.Figure()
        fig_bar.add_trace(go.Bar(name=_la, x=_segs_a, y=_dp_a,
                                 marker_color="#2563EB", opacity=0.85))
        fig_bar.add_trace(go.Bar(name=_lb, x=_segs_b, y=_dp_b,
                                 marker_color="#D97706", opacity=0.85))
        fig_bar.update_layout(
            barmode="group", yaxis_title="ΔP (kPa)", xaxis_title="Segment",
            template="plotly_white", height=320, margin=dict(l=60,r=20,t=20,b=60),
            paper_bgcolor="white", plot_bgcolor="white",
            xaxis=dict(gridcolor="#F1F5F9", linecolor="#E2E8F0"),
            yaxis=dict(gridcolor="#F1F5F9", linecolor="#E2E8F0"),
            legend=dict(bgcolor="rgba(255,255,255,0.9)", bordercolor="#E2E8F0", borderwidth=1),
            font=dict(size=12, color="#374151"))
        st.plotly_chart(fig_bar, width='stretch', key="fig_bar")

        # ── Method Sensitivity Analysis ────────────────────────────────────────────
        st.markdown("#### Method Sensitivity Analysis")

        _sh = _sens_hash(ra, rb)
        if st.session_state.get("sens_hash") != _sh:
            st.session_state.pop("sens_a", None)
            st.session_state.pop("sens_b", None)
            st.session_state.pop("sens_hash", None)

        _sc1, _sc2 = st.columns([1, 3])
        with _sc1:
            _run_sens = st.button("Run Sensitivity Analysis", type="primary",
                                  width='stretch', key="run_sens")
        with _sc2:
            if "sens_a" in st.session_state and "sens_b" in st.session_state:
                st.success("Results shown below — click again to refresh after input changes.")
            else:
                st.caption(
                    "Runs all 12 combinations (6 correlations × 2 void-fraction models). "
                    f"Quantifies the full ΔP range for {_la} and {_lb} due to method uncertainty.")

        if _run_sens:
            with st.spinner("Running 24 calculations (12 per case)…"):
                st.session_state["sens_a"] = engine.run_sensitivity(
                    ra["P_bara"], ra.get("T_C"), ra["gas_flows_kgh"], ra["liquid_type"], ra["q_lye"],
                    ra["segments"],
                    custom_gas=ra.get("custom_gas"), custom_liquid=ra.get("custom_liquid"),
                    vle_fluid=ra.get("vle_fluid"), vle_x_mass=ra.get("vle_x_mass"),
                    vle_m_total_kgs=ra.get("vle_m_total_kgs"))
                st.session_state["sens_b"] = engine.run_sensitivity(
                    rb["P_bara"], rb.get("T_C"), rb["gas_flows_kgh"], rb["liquid_type"], rb["q_lye"],
                    rb["segments"],
                    custom_gas=rb.get("custom_gas"), custom_liquid=rb.get("custom_liquid"),
                    vle_fluid=rb.get("vle_fluid"), vle_x_mass=rb.get("vle_x_mass"),
                    vle_m_total_kgs=rb.get("vle_m_total_kgs"))
                st.session_state["sens_hash"] = _sh
            st.rerun()

        fig_sens = None
        if "sens_a" in st.session_state and "sens_b" in st.session_state:
            _sa = st.session_state["sens_a"]
            _sb = st.session_state["sens_b"]

            _ylabels, _dp_a_vals, _dp_b_vals, _ok_a, _ok_b = [], [], [], [], []
            for _r_a, _r_b in zip(_sa, _sb):
                _c = _CORR_SHORT.get(_r_a["correlation"], _r_a["correlation"])
                _v = _VOID_SHORT.get(_r_a["voidage"], _r_a["voidage"])
                _ylabels.append(f"{_c} / {_v}")
                _dp_a_vals.append(_r_a["total_dp_kpa"] if _r_a["ok"] else None)
                _dp_b_vals.append(_r_b["total_dp_kpa"] if _r_b["ok"] else None)
                _ok_a.append(_r_a["ok"])
                _ok_b.append(_r_b["ok"])

            _va = [v for v in _dp_a_vals if v is not None]
            _vb = [v for v in _dp_b_vals if v is not None]

            fig_sens = go.Figure()

            # Shaded range bands
            if len(_va) >= 2:
                fig_sens.add_vrect(x0=min(_va), x1=max(_va),
                                   fillcolor="rgba(37,99,235,0.07)", line_width=0,
                                   annotation_text=f"{_la} range", annotation_position="top left",
                                   annotation_font=dict(size=9, color="#2563EB"))
            if len(_vb) >= 2:
                fig_sens.add_vrect(x0=min(_vb), x1=max(_vb),
                                   fillcolor="rgba(217,119,6,0.07)", line_width=0,
                                   annotation_text=f"{_lb} range", annotation_position="bottom right",
                                   annotation_font=dict(size=9, color="#D97706"))

            # Connecting lines — one trace with None separators
            _conn_x, _conn_y = [], []
            for _ci in range(len(_ylabels)):
                if _ok_a[_ci] and _ok_b[_ci]:
                    _conn_x.extend([_dp_a_vals[_ci], _dp_b_vals[_ci], None])
                    _conn_y.extend([_ylabels[_ci],   _ylabels[_ci],   None])
            if _conn_x:
                fig_sens.add_trace(go.Scatter(
                    x=_conn_x, y=_conn_y, mode="lines",
                    line=dict(color="#CBD5E1", width=1.2),
                    showlegend=False, hoverinfo="skip"))

            # Case B dots (amber diamonds, drawn first so A circles sit on top)
            _xb_p = [_dp_b_vals[_ci] for _ci in range(len(_ylabels)) if _ok_b[_ci]]
            _yb_p = [_ylabels[_ci]   for _ci in range(len(_ylabels)) if _ok_b[_ci]]
            fig_sens.add_trace(go.Scatter(
                x=_xb_p, y=_yb_p, mode="markers", name=_lb,
                marker=dict(color="#D97706", size=11, symbol="diamond",
                            line=dict(color="#92400E", width=1.5)),
                hovertemplate=f"{_lb}  |  %{{y}}<br>Total ΔP: %{{x:.3f}} kPa<extra></extra>"))

            # dots (blue circles)
            _xa_p = [_dp_a_vals[_ci] for _ci in range(len(_ylabels)) if _ok_a[_ci]]
            _ya_p = [_ylabels[_ci]   for _ci in range(len(_ylabels)) if _ok_a[_ci]]
            fig_sens.add_trace(go.Scatter(
                x=_xa_p, y=_ya_p, mode="markers", name=_la,
                marker=dict(color="#2563EB", size=11, symbol="circle",
                            line=dict(color="#1E40AF", width=1.5)),
                hovertemplate=f"{_la}  |  %{{y}}<br>Total ΔP: %{{x:.3f}} kPa<extra></extra>"))

            # Dashed reference lines for the currently-selected method in each case
            _sel_lbl_a = (f"{_CORR_SHORT.get(ra['correlation'], ra['correlation'])} / "
                          f"{_VOID_SHORT.get(ra['voidage_method'], ra['voidage_method'])}")
            _sel_lbl_b = (f"{_CORR_SHORT.get(rb['correlation'], rb['correlation'])} / "
                          f"{_VOID_SHORT.get(rb['voidage_method'], rb['voidage_method'])}")
            _sel_a_dp = next(
                (r["total_dp_kpa"] for r in _sa if r["ok"]
                 and f"{_CORR_SHORT.get(r['correlation'], r['correlation'])} / "
                     f"{_VOID_SHORT.get(r['voidage'], r['voidage'])}" == _sel_lbl_a), None)
            _sel_b_dp = next(
                (r["total_dp_kpa"] for r in _sb if r["ok"]
                 and f"{_CORR_SHORT.get(r['correlation'], r['correlation'])} / "
                     f"{_VOID_SHORT.get(r['voidage'], r['voidage'])}" == _sel_lbl_b), None)
            if _sel_a_dp is not None:
                fig_sens.add_vline(x=_sel_a_dp,
                                   line=dict(color="#2563EB", width=1.5, dash="dash"),
                                   annotation_text=f"{_la} selected: {_sel_a_dp:.2f} kPa",
                                   annotation_position="top right",
                                   annotation_font=dict(size=9, color="#2563EB"))
            if _sel_b_dp is not None:
                fig_sens.add_vline(x=_sel_b_dp,
                                   line=dict(color="#D97706", width=1.5, dash="dot"),
                                   annotation_text=f"{_lb} selected: {_sel_b_dp:.2f} kPa",
                                   annotation_position="bottom right",
                                   annotation_font=dict(size=9, color="#D97706"))

            fig_sens.update_layout(
                title=dict(text=f"Total ΔP — all 12 method combinations  (● {_la}  ◆ {_lb})",
                           font=dict(size=12), x=0),
                xaxis_title="Total ΔP (kPa)", yaxis_title=None,
                template="plotly_white", height=460,
                margin=dict(l=155, r=40, t=55, b=50),
                hovermode="closest", paper_bgcolor="white", plot_bgcolor="white",
                xaxis=dict(gridcolor="#F1F5F9", linecolor="#E2E8F0"),
                yaxis=dict(gridcolor="#F1F5F9", linecolor="#E2E8F0",
                           categoryorder="array",
                           categoryarray=list(reversed(_ylabels))),
                legend=dict(bgcolor="rgba(255,255,255,0.9)", bordercolor="#E2E8F0",
                            borderwidth=1, orientation="h", y=1.10, x=0),
                font=dict(size=11, color="#374151"))
            st.plotly_chart(fig_sens, width='stretch', key="fig_sens")

            _n_failed = sum(1 for ok in _ok_a + _ok_b if not ok)
            if _n_failed > 0:
                st.caption(f"{_n_failed} combination(s) failed to converge and are excluded from the chart.")

            # Summary table
            if _va and _vb:
                _a_min, _a_max = min(_va), max(_va)
                _b_min, _b_max = min(_vb), max(_vb)
                _overlap = _a_min <= _b_max and _b_min <= _a_max
                st.dataframe(pd.DataFrame([
                    {"": _la,
                     "Min (kPa)":      f"{_a_min:.3f}",
                     "Selected (kPa)": f"{ra['total_dp_kpa']:.3f}",
                     "Max (kPa)":      f"{_a_max:.3f}",
                     "Spread (kPa)":   f"{_a_max - _a_min:.3f}"},
                    {"": _lb,
                     "Min (kPa)":      f"{_b_min:.3f}",
                     "Selected (kPa)": f"{rb['total_dp_kpa']:.3f}",
                     "Max (kPa)":      f"{_b_max:.3f}",
                     "Spread (kPa)":   f"{_b_max - _b_min:.3f}"},
                ]), hide_index=True, width='stretch')
                if _overlap:
                    st.warning(
                        f"Ranges **overlap** — the relative ordering of {_la} vs {_lb} depends on "
                        "which correlation is chosen.")
                else:
                    st.success(
                        "Ranges **do not overlap** — one case is unambiguously lower-ΔP "
                        "across all methods.")

            # ── Flow Regime Consistency ────────────────────────────────────────────
            st.markdown("**Flow Regime Consistency across Methods**")
            st.caption(
                "Regime is independent of ΔP correlation — Vsg, Vsl, and angle are fixed. "
                "Only the void fraction model (α) can shift vertical-segment thresholds "
                "(bubble/slug/churn). Each column header shows the first correlation that "
                "produced that regime; all other correlations with the same void model agree.")

            def _regime_table(sens_results, segments_list, label):
                """Build regime DataFrame for one case across all converged combinations."""
                ok_results = [r for r in sens_results if r["ok"] and r["segment_regimes"]]
                if not ok_results or not segments_list:
                    return None

                n_segs = len(segments_list)
                # For each segment, collect the set of unique regimes across all combinations
                rows = []
                for i, seg in enumerate(segments_list):
                    seg_id   = f"#{i+1}"
                    orient   = seg.get("type", seg.get("kind", "—")).replace(
                                   "Vertical Upflow", "V Up").replace(
                                   "Vertical Downflow", "V Down").replace("Horizontal", "Horiz")
                    dn_str   = f"{seg['dn']}/{seg['pn']}"
                    # Group regimes by void fraction model (correlation doesn't change regime)
                    by_void = {}
                    for r in ok_results:
                        v = _VOID_SHORT.get(r["voidage"], r["voidage"])
                        regime_str = r["segment_regimes"][i] if i < len(r["segment_regimes"]) else "—"
                        by_void.setdefault(v, set()).add(regime_str)
                    # Build row: one column per void model, showing unique regimes found
                    row = {"Seg": seg_id, "Pipe": dn_str, "Orient": orient}
                    for v_short, regime_set in by_void.items():
                        row[v_short] = " / ".join(sorted(regime_set)) if regime_set else "—"
                    # Consensus across ALL combinations
                    all_regimes = set()
                    for r in ok_results:
                        if i < len(r["segment_regimes"]):
                            all_regimes.add(r["segment_regimes"][i])
                    row["Unanimous"] = "✓" if len(all_regimes) == 1 else f"✗  ({len(all_regimes)} distinct)"
                    rows.append(row)
                return pd.DataFrame(rows)

            _rt_a = _regime_table(_sa, ra["segments"], _la)
            _rt_b = _regime_table(_sb, rb["segments"], _lb)

            _rca, _rcb = st.columns(2)
            with _rca:
                st.markdown(f"**{_la}**")
                if _rt_a is not None:
                    st.dataframe(_rt_a, hide_index=True, width='stretch')
                else:
                    st.caption("No regime data available.")
            with _rcb:
                st.markdown(f"**{_lb}**")
                if _rt_b is not None:
                    st.dataframe(_rt_b, hide_index=True, width='stretch')
                else:
                    st.caption("No regime data available.")

        # ── Reports ───────────────────────────────────────────────────────────────
        st.divider()
        st.markdown("#### Reports")
        _sens_avail = ("sens_a" in st.session_state and "sens_b" in st.session_state
                       and fig_sens is not None)

        def _build_sens_data():
            if _sens_avail:
                return {"sa": st.session_state["sens_a"],
                        "sb": st.session_state["sens_b"],
                        "fig": fig_sens}
            return None

        def _build_stack_dp_data():
            _sh = st.session_state.get("stack_gsr_h2")
            _so = st.session_state.get("stack_gsr_o2")
            if _sh and _so:
                return {
                    "label_a": _la,
                    "label_b": _lb,
                    "gsr_h2":  _sh,
                    "gsr_o2":  _so,
                    "P_sep_h2": st.session_state.get("stack_sep_h2"),
                    "P_sep_o2": st.session_state.get("stack_sep_o2"),
                }
            return None

        def _build_dn_study_data():
            _dn_p = st.session_state.get("dn_study_dn_primary")
            _dn_a = st.session_state.get("dn_study_dn_alt")
            _gh_a = st.session_state.get("dn_study_gsr_h2_alt")
            _go_a = st.session_state.get("dn_study_gsr_o2_alt")
            _gh_p = st.session_state.get("stack_gsr_h2")
            _go_p = st.session_state.get("stack_gsr_o2")
            if not all([_dn_p, _dn_a, _gh_a, _go_a, _gh_p, _go_p]):
                return None
            _dp_p_mb = (_gh_p["P_line_in"] - _go_p["P_line_in"]) * 1000.0
            _dp_a_mb = (_gh_a["P_line_in"] - _go_a["P_line_in"]) * 1000.0
            _seg0    = ra["segments"][0] if ra.get("segments") else {}
            _pn0     = _seg0.get("pn", "PN20")
            _lined0  = _seg0.get("lined", False)
            _lthk0_m = _seg0.get("liner_thickness_mm", 1.0) / 1000.0
            _D_p_b   = engine.PIPE_DATABASE.get(_dn_p, {}).get(
                            _pn0, list(engine.PIPE_DATABASE.get(_dn_p, {None: 0}).values())[0])
            _D_a_b   = engine.PIPE_DATABASE.get(_dn_a, {}).get(
                            _pn0, list(engine.PIPE_DATABASE.get(_dn_a, {None: 0}).values())[0])
            _D_p_e   = _D_p_b - 2*_lthk0_m if _lined0 else _D_p_b
            _D_a_e   = _D_a_b - 2*_lthk0_m if _lined0 else _D_a_b
            _scale   = (_D_p_e / _D_a_e)**2 if _D_a_e > 0 else 1.0
            _rec_a0  = ra["grid_records"][0]  if ra.get("grid_records")  else {}
            _rec_b0  = rb["grid_records"][0]  if rb.get("grid_records")  else {}
            return {
                "dn_primary": _dn_p, "dn_alt": _dn_a,
                "label_a": _la, "label_b": _lb,
                "gsr_h2_primary": _gh_p, "gsr_o2_primary": _go_p,
                "gsr_h2_alt":     _gh_a, "gsr_o2_alt":     _go_a,
                "dp_gen_primary_mbar": _dp_p_mb,
                "dp_gen_alt_mbar":     _dp_a_mb,
                "vel_data": {
                    "vm_a_primary": float(_rec_a0.get("V_m (m/s)", 0)),
                    "vm_b_primary": float(_rec_b0.get("V_m (m/s)", 0)),
                    "vm_a_alt":     float(_rec_a0.get("V_m (m/s)", 0)) * _scale,
                    "vm_b_alt":     float(_rec_b0.get("V_m (m/s)", 0)) * _scale,
                    "ve_a":         float(_rec_a0.get("V_e (m/s)", 0)),
                    "ve_b":         float(_rec_b0.get("V_e (m/s)", 0)),
                    "D_p_mm":       _D_p_e * 1000,
                    "D_a_mm":       _D_a_e * 1000,
                    "vel_scale":    _scale,
                },
                "p_sep_h2": st.session_state.get("stack_sep_h2", 0),
                "p_sep_o2": st.session_state.get("stack_sep_o2", 0),
            }

        # ── Comparison report ─────────────────────────────────────────────────────
        _cmp_btn_col, _cmp_dl_col = st.columns(2)
        with _cmp_btn_col:
            if st.button("Generate Comparison (tables only)", type="primary",
                         width='stretch', key="gen_cmp_rpt"):
                with st.spinner("Building comparison report…"):
                    try:
                        _sd = _build_sens_data()
                        _cbuf = report_generator.generate_comparison_report(
                            results_a=ra, results_b=rb,
                            label_a=_la, label_b=_lb,
                            fig_cmp=None, fig_bar=None,
                            sensitivity_data={**_sd, "fig": None} if _sd else None,
                            stack_dp=_build_stack_dp_data())
                        st.session_state["cmp_rpt_bytes"] = _cbuf.getvalue()
                        st.success("Ready.")
                    except Exception as _e:
                        st.error(f"Failed: {_e}")
            if st.button("Generate Comparison with charts", width='stretch',
                         key="gen_cmp_rpt_ch"):
                with st.spinner("Building comparison report with charts…"):
                    try:
                        _cbuf = report_generator.generate_comparison_report(
                            results_a=ra, results_b=rb,
                            label_a=_la, label_b=_lb,
                            fig_cmp=fig_cmp, fig_bar=fig_bar,
                            sensitivity_data=_build_sens_data(),
                            stack_dp=_build_stack_dp_data())
                        st.session_state["cmp_rpt_bytes_ch"] = _cbuf.getvalue()
                        st.success("Ready.")
                    except Exception as _e:
                        st.error(f"Failed: {_e}")
        with _cmp_dl_col:
            if st.session_state.get("cmp_rpt_bytes"):
                st.download_button(
                    "Download Comparison tables  (.docx)",
                    data=st.session_state["cmp_rpt_bytes"],
                    file_name="report_comparison.docx",
                    mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                    width='stretch', key="dl_cmp_rpt")
            if st.session_state.get("cmp_rpt_bytes_ch"):
                st.download_button(
                    "Download Comparison with charts  (.docx)",
                    data=st.session_state["cmp_rpt_bytes_ch"],
                    file_name="report_comparison_charts.docx",
                    mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                    width='stretch', key="dl_cmp_rpt_ch")

        # ── Combined report ───────────────────────────────────────────────────────
        _cmb_btn_col, _cmb_dl_col = st.columns(2)
        with _cmb_btn_col:
            if st.button("Generate Combined (tables only)", type="primary",
                         width='stretch', key="gen_combined_rpt"):
                with st.spinner("Building combined report…"):
                    try:
                        _sd = _build_sens_data()
                        _combined_buf = report_generator.generate_combined_report(
                            cases=[ra, rb, results_c, results_d],
                            case_labels=[_la, _lb, _lc, _ld],
                            fig_cmp=None, fig_bar=None,
                            sensitivity_data={**_sd, "fig": None} if _sd else None,
                            stack_dp=_build_stack_dp_data(),
                            dn_study_data=_build_dn_study_data())
                        st.session_state["combined_rpt_bytes"] = _combined_buf.getvalue()
                        st.success("Ready.")
                    except Exception as _e:
                        st.error(f"Failed: {_e}")
            if st.button("Generate Combined with charts", width='stretch',
                         key="gen_combined_rpt_ch"):
                with st.spinner("Building combined report with charts…"):
                    try:
                        _combined_buf = report_generator.generate_combined_report(
                            cases=[ra, rb, results_c, results_d],
                            case_labels=[_la, _lb, _lc, _ld],
                            fig_cmp=fig_cmp, fig_bar=fig_bar,
                            sensitivity_data=_build_sens_data(),
                            stack_dp=_build_stack_dp_data(),
                            dn_study_data=_build_dn_study_data())
                        st.session_state["combined_rpt_bytes_ch"] = _combined_buf.getvalue()
                        st.success("Ready.")
                    except Exception as _e:
                        st.error(f"Failed: {_e}")
        with _cmb_dl_col:
            if st.session_state.get("combined_rpt_bytes"):
                st.download_button(
                    "Download Combined tables  (.docx)",
                    data=st.session_state["combined_rpt_bytes"],
                    file_name="report_combined.docx",
                    mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                    width='stretch', key="dl_combined_rpt")
            if st.session_state.get("combined_rpt_bytes_ch"):
                st.download_button(
                    "Download Combined with charts  (.docx)",
                    data=st.session_state["combined_rpt_bytes_ch"],
                    file_name="report_combined_charts.docx",
                    mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                    width='stretch', key="dl_combined_rpt_ch")

        st.caption("ℹ Individual case reports (A, B) are available in their respective tabs.")
        if not _sens_avail:
            st.caption("ℹ Run the Sensitivity Analysis above first to include it in the reports.")
        if not st.session_state.get("stack_gsr_h2"):
            st.caption("ℹ Run the Goal Seek calculation first to include it in the reports.")
        if not st.session_state.get("dn_study_dn_primary"):
            st.caption("ℹ Run the DN Study first to include it in the Combined report.")

        st.divider()

        # ── Full segment tables side by side ──────────────────────────────────────
        st.markdown("#### Segment Detail")
        _col_cfg = {
            "ID (mm)":          st.column_config.NumberColumn(format="%.1f"),
            "P_in (bara)":      st.column_config.NumberColumn(format="%.4f"),
            "P_out (bara)":     st.column_config.NumberColumn(format="%.4f"),
            "α (void)":         st.column_config.NumberColumn(format="%.4f"),
            "ΔP_fric (kPa)":    st.column_config.NumberColumn(format="%.3f"),
            "ΔP_grav (kPa)":    st.column_config.NumberColumn(format="%.3f"),
            "ΔP (kPa)":         st.column_config.NumberColumn(format="%.3f"),
            "V_m (m/s)":        st.column_config.NumberColumn(format="%.3f"),
            "V_m/V_e":          st.column_config.NumberColumn(format="%.3f"),
        }
        _cmp_cols = ["Seg","Pipe","ID (mm)","Type","L (m)","Regime",
                     "α (void)","V_m (m/s)","V_m/V_e",
                     "ΔP_fric (kPa)","ΔP_grav (kPa)","ΔP (kPa)",
                     "P_in (bara)","P_out (bara)"]
        _ta, _tb = st.columns(2)
        with _ta:
            st.markdown(f"**{_la}**")
            st.dataframe(pd.DataFrame(ra["grid_records"])[_cmp_cols],
                         column_config=_col_cfg, hide_index=True, width='stretch')
        with _tb:
            st.markdown(f"**{_lb}**")
            st.dataframe(pd.DataFrame(rb["grid_records"])[_cmp_cols],
                         column_config=_col_cfg, hide_index=True, width='stretch')


    # ============================================================================
    # GOAL SEEK TAB
    # ============================================================================
    with tab_gs:
        _la = st.session_state.get("label_a") or "A"
        _lb = st.session_state.get("label_b") or "B"
        _lc = f"{_la} Header"
        _ld = f"{_lb} Header"

        st.subheader("Goal Seek")

        _gs_direction = st.radio(
            "Mode",
            [
                "Fix separator pressure — back-calculate inlet",
                "Fix source pressure — forward-calculate separator",
            ],
            horizontal=True,
            key="gs_direction",
            label_visibility="collapsed",
        )
        _sink_mode = _gs_direction.startswith("Fix separator")

        # ── Sink-fixed mode (back-calculation) ───────────────────────────────────
        if _sink_mode:
            st.caption(
                f"Goal-seek both the {_la} system ({_la} → {_lc}) and {_lb} system ({_lb} → {_ld}) "
                "to find the required line inlet pressures for given separator pressures. "
                f"**Generator ΔP** = P_inlet_{_la} − P_inlet_{_lb}."
            )

            with st.container(border=True):
                st.markdown("##### Separator Target Pressures")
                _sk_col1, _sk_col2 = st.columns(2)
                with _sk_col1:
                    _p_sep_h2 = st.number_input(
                        f"{_la} separator pressure (bara)",
                        min_value=0.1, max_value=200.0,
                        value=float(round(results_c.get("P_separator_bara",
                                                        results_a["outlet_pressure_bara"]), 3)),
                        step=0.1, format="%.3f",
                        key="stack_p_sep_h2",
                        help=f"Target pressure at the {_la} gas-liquid separator.",
                    )
                with _sk_col2:
                    _p_sep_o2 = st.number_input(
                        f"{_lb} separator pressure (bara)",
                        min_value=0.1, max_value=200.0,
                        value=float(round(results_d.get("P_separator_bara",
                                                        results_b["outlet_pressure_bara"]), 3)),
                        step=0.1, format="%.3f",
                        key="stack_p_sep_o2",
                        help=f"Target pressure at the {_lb} gas-liquid separator.",
                    )
                _sk_run = st.button("Calculate", type="primary",
                                    width='stretch', key="stack_run")

            if _sk_run:
                with st.spinner(f"Solving {_la} and {_lb} systems…"):
                    _gsr_h2 = _goal_seek_stack(results_a, results_c, _p_sep_h2)
                    _gsr_o2 = _goal_seek_stack(results_b, results_d, _p_sep_o2)
                st.session_state["stack_gsr_h2"] = _gsr_h2
                st.session_state["stack_gsr_o2"] = _gsr_o2
                st.session_state["stack_sep_h2"] = _p_sep_h2
                st.session_state["stack_sep_o2"] = _p_sep_o2
                st.session_state["gs_mode_used"] = "sink"

        # ── Source-fixed mode (forward march) ────────────────────────────────────
        else:
            st.caption(
                f"Fix the {_la} and {_lb} source (inlet) pressures and march forward through "
                f"the branch pipeline and header to find the resulting separator pressures. "
                f"**Generator ΔP** = P_source_{_la} − P_source_{_lb} (set by you)."
            )

            with st.container(border=True):
                st.markdown("##### Source (Inlet) Pressures")
                _fwd_col1, _fwd_col2 = st.columns(2)
                with _fwd_col1:
                    _p_src_h2 = st.number_input(
                        f"{_la} source pressure (bara)",
                        min_value=0.1, max_value=200.0,
                        value=float(round(results_a.get("inlet_pressure_bara",
                                                        results_a["outlet_pressure_bara"]), 3)),
                        step=0.1, format="%.3f",
                        key="gs_p_src_h2",
                        help=f"Fixed inlet pressure for the {_la} branch.",
                    )
                with _fwd_col2:
                    _p_src_o2 = st.number_input(
                        f"{_lb} source pressure (bara)",
                        min_value=0.1, max_value=200.0,
                        value=float(round(results_b.get("inlet_pressure_bara",
                                                        results_b["outlet_pressure_bara"]), 3)),
                        step=0.1, format="%.3f",
                        key="gs_p_src_o2",
                        help=f"Fixed inlet pressure for the {_lb} branch.",
                    )
                _fwd_run = st.button("Calculate", type="primary",
                                     width='stretch', key="gs_fwd_run")

            if _fwd_run:
                with st.spinner(f"Marching {_la} and {_lb} systems forward…"):
                    _gsr_h2 = _forward_stack(results_a, results_c, _p_src_h2)
                    _gsr_o2 = _forward_stack(results_b, results_d, _p_src_o2)
                st.session_state["stack_gsr_h2"] = _gsr_h2
                st.session_state["stack_gsr_o2"] = _gsr_o2
                st.session_state["stack_sep_h2"] = _gsr_h2["P_sep"]
                st.session_state["stack_sep_o2"] = _gsr_o2["P_sep"]
                st.session_state["gs_mode_used"] = "source"

        # ── Shared results display ────────────────────────────────────────────────
        _gsr_h2 = st.session_state.get("stack_gsr_h2")
        _gsr_o2 = st.session_state.get("stack_gsr_o2")

        if _gsr_h2 and _gsr_o2:
            _shown_sep_h2 = st.session_state.get("stack_sep_h2", 0.0)
            _shown_sep_o2 = st.session_state.get("stack_sep_o2", 0.0)

            _conv_h2 = _gsr_h2["converged"]
            _conv_o2 = _gsr_o2["converged"]
            if _conv_h2 and _conv_o2:
                _iter_note = (
                    f"{_la} converged in {_gsr_h2['iterations']} iter.  "
                    f"{_lb} converged in {_gsr_o2['iterations']} iter."
                    if _sink_mode else
                    f"Forward march complete — {_la} and {_lb} solved in one pass."
                )
                st.success(_iter_note)
            else:
                if not _conv_h2:
                    st.warning(f"{_la} did not fully converge ({_gsr_h2['iterations']} iter, "
                               f"residual {abs(_gsr_h2['P_sep'] - _shown_sep_h2)*1000:.1f} mbar)")
                if not _conv_o2:
                    st.warning(f"{_lb} did not fully converge ({_gsr_o2['iterations']} iter, "
                               f"residual {abs(_gsr_o2['P_sep'] - _shown_sep_o2)*1000:.1f} mbar)")

            st.divider()

            _p_in_a = _gsr_h2["P_line_in"]
            _p_in_b = _gsr_o2["P_line_in"]
            _dp_stack = _p_in_a - _p_in_b

            with st.container(border=True):
                st.markdown(f"##### Generator ΔP  (P_inlet_{_la} − P_inlet_{_lb})")
                _sk1, _sk2, _sk3 = st.columns(3)
                _sk1.metric(
                    f"{_la} inlet pressure",
                    f"{_p_in_a:.4f} bara",
                    delta=f"{_la} ΔP = {_gsr_h2['dp_line']:.3f} kPa",
                    delta_color="off",
                )
                _sk2.metric(
                    f"{_lb} inlet pressure",
                    f"{_p_in_b:.4f} bara",
                    delta=f"{_lb} ΔP = {_gsr_o2['dp_line']:.3f} kPa",
                    delta_color="off",
                )
                _dp_stack_kpa  = _dp_stack * 100.0
                _dp_stack_mbar = _dp_stack_kpa * 10.0
                _sk3.metric(
                    f"Generator ΔP  ({_la} − {_lb})",
                    f"{_dp_stack:.4f} bara",
                    delta=f"{_dp_stack_kpa:.2f} kPa  ·  {_dp_stack_mbar:.1f} mbar",
                    delta_color="off",
                )

            st.markdown("##### System Pressure Breakdown")
            with st.container(border=True):
                _bk1, _bk2 = st.columns(2)
                with _bk1:
                    st.markdown(f"**{_la} system  ({_la} → {_lc} → Separator)**")
                    st.metric(f"{_la} inlet",
                              f"{_gsr_h2['P_line_in']:.4f} bara")
                    st.metric(f"{_la} outlet / {_lc} inlet",
                              f"{_gsr_h2['P_line_out']:.4f} bara",
                              delta=f"{_la} ΔP = {_gsr_h2['dp_line']:.3f} kPa",
                              delta_color="off")
                    st.metric(f"{_la} Separator",
                              f"{_gsr_h2['P_sep']:.4f} bara",
                              delta=f"{_lc}+T ΔP = {_gsr_h2['dp_hdr']:.3f} kPa",
                              delta_color="off")
                with _bk2:
                    st.markdown(f"**{_lb} system  ({_lb} → {_ld} → Separator)**")
                    st.metric(f"{_lb} inlet",
                              f"{_gsr_o2['P_line_in']:.4f} bara")
                    st.metric(f"{_lb} outlet / {_ld} inlet",
                              f"{_gsr_o2['P_line_out']:.4f} bara",
                              delta=f"{_lb} ΔP = {_gsr_o2['dp_line']:.3f} kPa",
                              delta_color="off")
                    st.metric(f"{_lb} Separator",
                              f"{_gsr_o2['P_sep']:.4f} bara",
                              delta=f"{_ld}+T ΔP = {_gsr_o2['dp_hdr']:.3f} kPa",
                              delta_color="off")

            st.divider()
            st.markdown("##### Apply Results to Cases")
            st.caption(
                "Write the required inlet pressures back into Case A and Case B "
                "so the branch calculations reflect the goal-seek solution."
            )
            _ap1, _ap2 = st.columns(2)
            if _ap1.button(
                f"Apply {_p_in_a:.4f} bara → {_la}",
                width='stretch', key="stack_apply_a",
            ):
                st.session_state["stack_apply_a_pending"] = float(_p_in_a)
                st.rerun()
            if _ap2.button(
                f"Apply {_p_in_b:.4f} bara → {_lb}",
                width='stretch', key="stack_apply_b",
            ):
                st.session_state["stack_apply_b_pending"] = float(_p_in_b)
                st.rerun()

        # ============================================================================
        # DN STUDY (merged into Goal Seek tab)
        # ============================================================================
        st.divider()
        st.subheader("DN Study — Branch Line Size Comparison")
        st.caption(
            "Re-runs the full system (branches A and B + goal-seek) with a different branch DN. "
            "Header sizes are unchanged. All other inputs (flows, pressure, correlation) are identical."
        )

        _prereq_ok = (
            results_a is not None and results_b is not None
            and results_c is not None and results_d is not None
        )
        _gsr_ok = st.session_state.get("gs_mode_used") is not None

        if not _prereq_ok:
            st.warning("Run all four cases (A, B, C, D) before using DN Study.")
        elif not _gsr_ok:
            st.warning("Run the Goal Seek calculation above before using DN Study.")
        else:
            _dn_primary = results_a["segments"][0]["dn"] if results_a.get("segments") else "DN50"
            _dn_all_opts = list(engine.PIPE_DATABASE.keys())
            _dn_alt_opts = [dn for dn in _dn_all_opts if dn != _dn_primary]
            _dn_default_idx = _dn_alt_opts.index("DN40") if "DN40" in _dn_alt_opts else 0

            _inp_col, _res_col = st.columns([1, 2])

            with _inp_col:
                with st.container(border=True):
                    st.markdown("**Study Settings**")
                    dn_alt = st.selectbox(
                        "Alternative branch DN", _dn_alt_opts,
                        index=_dn_default_idx, key="dn_study_alt_sel"
                    )
                    _p_sep_h2_dn = st.session_state.get("stack_sep_h2", 16.5)
                    _p_sep_o2_dn = st.session_state.get("stack_sep_o2", 16.5)
                    st.caption(
                        f"Separator targets (from Goal Seek above):  "
                        f"{_la} {_p_sep_h2_dn:.2f} bara  ·  {_lb} {_p_sep_o2_dn:.2f} bara"
                    )
                    _run_dn_study = st.button(
                        "Run DN Study", type="primary",
                        width='stretch', key="dn_study_run"
                    )

            if _run_dn_study:
                with st.spinner(f"Computing {_dn_primary} vs {dn_alt}…"):
                    _ra_alt = _apply_dn_override(results_a, dn_alt)
                    _rb_alt = _apply_dn_override(results_b, dn_alt)
                    _gsr_h2_alt = _goal_seek_stack(_ra_alt, results_c, _p_sep_h2_dn)
                    _gsr_o2_alt = _goal_seek_stack(_rb_alt, results_d, _p_sep_o2_dn)
                    _reg_a_alt  = _calc_regimes_at_p(_ra_alt, _gsr_h2_alt["P_line_in"])
                    _reg_b_alt  = _calc_regimes_at_p(_rb_alt, _gsr_o2_alt["P_line_in"])
                st.session_state["dn_study_dn_primary"] = _dn_primary
                st.session_state["dn_study_dn_alt"]     = dn_alt
                st.session_state["dn_study_gsr_h2_alt"] = _gsr_h2_alt
                st.session_state["dn_study_gsr_o2_alt"] = _gsr_o2_alt
                st.session_state["dn_study_regimes_a"]  = _reg_a_alt
                st.session_state["dn_study_regimes_b"]  = _reg_b_alt
                st.rerun()

            _gsr_h2_p = st.session_state.get("stack_gsr_h2")
            _gsr_o2_p = st.session_state.get("stack_gsr_o2")
            _gsr_h2_a = st.session_state.get("dn_study_gsr_h2_alt")
            _gsr_o2_a = st.session_state.get("dn_study_gsr_o2_alt")
            _dn_p_lbl = st.session_state.get("dn_study_dn_primary", _dn_primary)
            _dn_a_lbl = st.session_state.get("dn_study_dn_alt", "")

            if _gsr_h2_a and _gsr_o2_a and _dn_a_lbl:
                with _res_col:
                    # ── Generator ΔP ──────────────────────────────────────────────
                    _dp_gen_p_mbar = (_gsr_h2_p["P_line_in"] - _gsr_o2_p["P_line_in"]) * 1000.0
                    _dp_gen_a_mbar = (_gsr_h2_a["P_line_in"] - _gsr_o2_a["P_line_in"]) * 1000.0
                    _dp_gen_delta  = _dp_gen_a_mbar - _dp_gen_p_mbar

                    with st.container(border=True):
                        st.markdown("**Generator ΔP**")
                        _gc1, _gc2, _gc3 = st.columns(3)
                        _gc1.metric(f"{_dn_p_lbl} (primary)", f"{_dp_gen_p_mbar:.1f} mbar")
                        _gc2.metric(
                            f"{_dn_a_lbl} (alternative)",
                            f"{_dp_gen_a_mbar:.1f} mbar",
                            delta=f"{_dp_gen_delta:+.1f} mbar",
                            delta_color="off",
                        )
                        _winner = _dn_p_lbl if abs(_dp_gen_p_mbar) <= abs(_dp_gen_a_mbar) else _dn_a_lbl
                        _gc3.metric("Lower |ΔP|", _winner)

                    # ── ΔP comparison table ───────────────────────────────────────
                    _dp_rows = []
                    for _case_lbl, _dp_p_kpa, _dp_a_kpa in [
                        (f"{_la} branch", _gsr_h2_p["dp_line"], _gsr_h2_a["dp_line"]),
                        (f"{_lb} branch", _gsr_o2_p["dp_line"], _gsr_o2_a["dp_line"]),
                        (f"{_la} header", _gsr_h2_p["dp_hdr"],  _gsr_h2_a["dp_hdr"]),
                        (f"{_lb} header", _gsr_o2_p["dp_hdr"],  _gsr_o2_a["dp_hdr"]),
                    ]:
                        _pct = ((_dp_a_kpa - _dp_p_kpa) / _dp_p_kpa * 100
                                if _dp_p_kpa and abs(_dp_p_kpa) > 1e-9 else 0.0)
                        _dp_rows.append({
                            "Case":                      _case_lbl,
                            f"{_dn_p_lbl} ΔP (kPa)":    round(_dp_p_kpa, 3),
                            f"{_dn_a_lbl} ΔP (kPa)":    round(_dp_a_kpa, 3),
                            "Change (%)":                f"{_pct:+.1f}",
                        })
                    st.dataframe(
                        pd.DataFrame(_dp_rows),
                        hide_index=True, width='stretch',
                        column_config={
                            f"{_dn_p_lbl} ΔP (kPa)": st.column_config.NumberColumn(format="%.3f"),
                            f"{_dn_a_lbl} ΔP (kPa)": st.column_config.NumberColumn(format="%.3f"),
                        }
                    )

                    # ── Velocity estimate (first segment, ID-ratio scaling) ────────
                    _seg0     = results_a["segments"][0]
                    _pn0      = _seg0["pn"]
                    _lined0   = _seg0.get("lined", False)
                    _lthk0_m  = _seg0.get("liner_thickness_mm", 1.0) / 1000.0
                    _D_p_bore = engine.PIPE_DATABASE[_dn_p_lbl].get(
                        _pn0, list(engine.PIPE_DATABASE[_dn_p_lbl].values())[0])
                    _D_a_bore = engine.PIPE_DATABASE[_dn_a_lbl].get(
                        _pn0, list(engine.PIPE_DATABASE[_dn_a_lbl].values())[0])
                    _D_p_eff  = _D_p_bore - 2 * _lthk0_m if _lined0 else _D_p_bore
                    _D_a_eff  = _D_a_bore - 2 * _lthk0_m if _lined0 else _D_a_bore
                    _vel_scale = (_D_p_eff / _D_a_eff) ** 2 if _D_a_eff > 0 else 1.0

                    _rec_a0 = results_a["grid_records"][0] if results_a.get("grid_records") else {}
                    _rec_b0 = results_b["grid_records"][0] if results_b.get("grid_records") else {}
                    _vm_a_p = float(_rec_a0.get("V_m (m/s)", 0))
                    _vm_b_p = float(_rec_b0.get("V_m (m/s)", 0))
                    _ve_a   = float(_rec_a0.get("V_e (m/s)", 0))
                    _ve_b   = float(_rec_b0.get("V_e (m/s)", 0))
                    _vm_a_a = _vm_a_p * _vel_scale
                    _vm_b_a = _vm_b_p * _vel_scale
                    _ratio_a = _vm_a_a / _ve_a if _ve_a > 0 else 0.0
                    _ratio_b = _vm_b_a / _ve_b if _ve_b > 0 else 0.0

                    with st.container(border=True):
                        st.markdown("**Inlet velocity — first segment (estimated)**")
                        _vc1, _vc2 = st.columns(2)
                        with _vc1:
                            st.markdown(f"*{_la} branch*")
                            st.metric(f"{_dn_p_lbl}", f"{_vm_a_p:.2f} m/s",
                                      help=f"V_m/V_e = {_vm_a_p/_ve_a:.2f}" if _ve_a > 0 else None)
                            _dc_a = "inverse" if _ratio_a > 1.0 else ("normal" if _ratio_a > 0.8 else "off")
                            st.metric(f"{_dn_a_lbl}", f"{_vm_a_a:.2f} m/s",
                                      delta=f"V_m/V_e = {_ratio_a:.2f}", delta_color=_dc_a)
                        with _vc2:
                            st.markdown(f"*{_lb} branch*")
                            st.metric(f"{_dn_p_lbl}", f"{_vm_b_p:.2f} m/s",
                                      help=f"V_m/V_e = {_vm_b_p/_ve_b:.2f}" if _ve_b > 0 else None)
                            _dc_b = "inverse" if _ratio_b > 1.0 else ("normal" if _ratio_b > 0.8 else "off")
                            st.metric(f"{_dn_a_lbl}", f"{_vm_b_a:.2f} m/s",
                                      delta=f"V_m/V_e = {_ratio_b:.2f}", delta_color=_dc_b)
                        st.caption(
                            f"Velocity scales as (ID ratio)²: "
                            f"{_dn_p_lbl} {_D_p_eff*1000:.1f} mm → {_dn_a_lbl} {_D_a_eff*1000:.1f} mm  "
                            f"·  factor {_vel_scale:.2f}×.  "
                            f"Erosion limit V_e from primary case (API RP 14E, C = 100)."
                        )

                    # ── Flow Regime ───────────────────────────────────────────────
                    _reg_a_stored = st.session_state.get("dn_study_regimes_a", [])
                    _reg_b_stored = st.session_state.get("dn_study_regimes_b", [])
                    if _reg_a_stored or _reg_b_stored:
                        with st.container(border=True):
                            st.markdown("**Flow Regime by Segment**")
                            _rr1, _rr2 = st.columns(2)
                            with _rr1:
                                st.caption(f"*{_la} branch*")
                                _prim_a_recs = results_a.get("grid_records", [])
                                _rrA = []
                                for _ri, _ar in enumerate(_reg_a_stored):
                                    _pr = _prim_a_recs[_ri]["Regime"] if _ri < len(_prim_a_recs) else "—"
                                    _rrA.append({
                                        "Seg":     _ar["seg"],
                                        _dn_p_lbl: _pr,
                                        _dn_a_lbl: _ar["regime"],
                                        "Changed": "⚠" if _pr != _ar["regime"] else "✓",
                                    })
                                st.dataframe(pd.DataFrame(_rrA), hide_index=True,
                                             width='stretch')
                            with _rr2:
                                st.caption(f"*{_lb} branch*")
                                _prim_b_recs = results_b.get("grid_records", [])
                                _rrB = []
                                for _ri, _br in enumerate(_reg_b_stored):
                                    _pr = _prim_b_recs[_ri]["Regime"] if _ri < len(_prim_b_recs) else "—"
                                    _rrB.append({
                                        "Seg":     _br["seg"],
                                        _dn_p_lbl: _pr,
                                        _dn_a_lbl: _br["regime"],
                                        "Changed": "⚠" if _pr != _br["regime"] else "✓",
                                    })
                                st.dataframe(pd.DataFrame(_rrB), hide_index=True,
                                             width='stretch')

                    # ── Recommendation ────────────────────────────────────────────
                    _vel_ok        = _ratio_a <= 1.0 and _ratio_b <= 1.0
                    _dp_alt_better = abs(_dp_gen_a_mbar) < abs(_dp_gen_p_mbar)
                    if _dp_alt_better and _vel_ok:
                        st.success(
                            f"**{_dn_a_lbl}** gives lower Generator |ΔP| "
                            f"({_dp_gen_a_mbar:.1f} vs {_dp_gen_p_mbar:.1f} mbar) "
                            f"with acceptable velocities."
                        )
                    elif _dp_alt_better and not _vel_ok:
                        st.warning(
                            f"**{_dn_a_lbl}** gives lower Generator |ΔP| but estimated "
                            f"V_m/V_e exceeds 1.0 — verify erosion before selecting."
                        )
                    else:
                        st.info(
                            f"**{_dn_p_lbl}** (primary) gives lower Generator |ΔP|. "
                            f"{_dn_a_lbl} appears undersized for this duty."
                        )

                    # ── Report ────────────────────────────────────────────────────
                    st.divider()
                    _dn_rpt_col1, _dn_rpt_col2 = st.columns(2)
                    with _dn_rpt_col1:
                        if st.button("Generate DN Study Report", width='stretch',
                                     key="dn_study_gen_rpt"):
                            try:
                                _dn_buf = report_generator.generate_dn_study_report(
                                    dn_primary=_dn_p_lbl,
                                    dn_alt=_dn_a_lbl,
                                    label_a=_la, label_b=_lb,
                                    gsr_h2_primary=_gsr_h2_p,
                                    gsr_o2_primary=_gsr_o2_p,
                                    gsr_h2_alt=_gsr_h2_a,
                                    gsr_o2_alt=_gsr_o2_a,
                                    dp_gen_primary_mbar=_dp_gen_p_mbar,
                                    dp_gen_alt_mbar=_dp_gen_a_mbar,
                                    vel_data={
                                        "vm_a_primary": _vm_a_p, "vm_b_primary": _vm_b_p,
                                        "vm_a_alt":     _vm_a_a, "vm_b_alt":     _vm_b_a,
                                        "ve_a":         _ve_a,   "ve_b":         _ve_b,
                                        "D_p_mm":       _D_p_eff * 1000,
                                        "D_a_mm":       _D_a_eff * 1000,
                                        "vel_scale":    _vel_scale,
                                    },
                                    p_sep_h2=_p_sep_h2_dn,
                                    p_sep_o2=_p_sep_o2_dn,
                                )
                                st.session_state["dn_study_rpt_bytes"] = _dn_buf.getvalue()
                            except Exception as _re:
                                st.error(f"Report failed: {_re}")
                    with _dn_rpt_col2:
                        if st.session_state.get("dn_study_rpt_bytes"):
                            st.download_button(
                                f"Download  {_dn_p_lbl}_vs_{_dn_a_lbl}.docx",
                                data=st.session_state["dn_study_rpt_bytes"],
                                file_name=f"dn_study_{_dn_p_lbl}_vs_{_dn_a_lbl}.docx",
                                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                                width='stretch', key="dn_study_dl_rpt",
                            )


else:  # Engineering Tools
    tab_fanno, tab_ro, tab_psv, tab_cv, tab_dg, tab_pump, tab_ls = st.tabs(
        ["Fanno Flow", "RO", "PSV", "Control Valve", "Dissolved Gas Flash", "Pump", "Line Size"]
    )
    import dissolution_engine as dg

    # =============================================================================
    # Tab: Fanno Flow
    # =============================================================================
    with tab_fanno:
        from plotly.subplots import make_subplots

        st.markdown(
            "Adiabatic compressible flow in a constant-area duct with friction. "
            "Predicts whether friction chokes the flow and how far you are from the sonic limit."
        )

        with st.expander("When to use Fanno flow", expanded=False):
            _fc1, _fc2 = st.columns(2)
            with _fc1:
                st.markdown(
                    """
    **Use Fanno flow when:**
    - Single-phase **gas only** — no liquid present
    - Flow is **adiabatic** (short or well-insulated duct)
    - **Constant cross-section** — no fittings, valves, or reducers
    - Inlet **Ma > 0.1** — at lower Ma use Darcy-Weisbach gas-only mode
    - You need to check whether friction can **choke** the flow or how close you are
    - Applications: high-velocity gas headers, relief headers, pneumatic conveying, instrument gas

    **Typical services:** H₂, N₂, Air, O₂, natural gas at elevated pressure and high velocity (> 30 m/s).
    """
                )
            with _fc2:
                st.markdown(
                    """
    **Use Darcy-Weisbach (tabs A / B) instead when:**
    - Ma < 0.1 — compressibility effects < 1 %
    - **Two-phase** gas–liquid flow
    - Piping with **fittings, valves, reducers** — use equivalent-length D-W
    - **Non-ideal gas** at high pressure (Pr > 0.5) — ideal-gas assumption breaks down
    - Very long pipelines where **isothermal** flow is more representative

    **Fanno vs. isothermal D-W:** agree within ~5 % for Ma < 0.3; Fanno required above Ma = 0.3.
    """
                )

        st.divider()

        _fl, _fr = st.columns([1, 1.6], gap="large")

        with _fl:
            st.subheader("Inputs")

            st.markdown("**Gas**")
            _fn_species = st.selectbox("Species", list(fanno.GASES.keys()), index=0, key="fn_species")
            _fn_mw0, _fn_g0, _ = fanno.GASES[_fn_species]
            if _fn_species == "Custom":
                _fnc1, _fnc2 = st.columns(2)
                _fn_mw_gmol = _fnc1.number_input("MW (g/mol)", value=28.0, min_value=1.0, max_value=300.0, step=0.5, key="fn_mw")
                _fn_gamma   = _fnc2.number_input("γ (Cp/Cv)", value=1.40, min_value=1.01, max_value=1.80, step=0.01, format="%.3f", key="fn_gamma")
                _fn_mw = _fn_mw_gmol / 1000.0
            else:
                _fn_mw    = _fn_mw0
                _fn_gamma = _fn_g0
                st.caption(f"MW = {_fn_mw*1000:.3f} g/mol  |  γ = {_fn_gamma:.3f}")

            st.markdown("**Inlet conditions (static)**")
            _fnc1, _fnc2 = st.columns(2)
            _fn_P1_bara = _fnc1.number_input("Pressure (bara)", value=5.0, min_value=0.1, max_value=500.0, step=0.5, format="%.2f", key="fn_P1")
            _fn_T1_C    = _fnc2.number_input("Temperature (°C)", value=20.0, min_value=-50.0, max_value=500.0, step=1.0, format="%.1f", key="fn_T1")

            st.markdown("**Flow rate**")
            _fn_mdot_kgh = st.number_input("Mass flow rate (kg/h)", value=100.0, min_value=0.01, max_value=1e6, step=10.0, format="%.2f", key="fn_mdot")

            st.markdown("**Pipe**")
            _fnc1, _fnc2 = st.columns(2)
            _fn_dn = _fnc1.selectbox("DN", list(engine.PIPE_DATABASE.keys()), index=4, key="fn_dn")
            _fn_pn = _fnc2.selectbox("PN", ["PN20", "PN25", "PN40"], index=0, key="fn_pn")
            _fn_D_m = engine.PIPE_DATABASE[_fn_dn][_fn_pn]

            _fnc1, _fnc2 = st.columns(2)
            _fn_mat   = _fnc1.selectbox("Material", list(engine.MATERIAL_ROUGHNESS.keys()), index=0, key="fn_mat")
            _fn_liner = _fnc2.selectbox("Liner", ["None"] + list(engine.LINER_ROUGHNESS.keys()), index=0, key="fn_liner")
            _fn_rough = engine.LINER_ROUGHNESS[_fn_liner] if _fn_liner != "None" else engine.MATERIAL_ROUGHNESS[_fn_mat]

            _fn_L_m = st.number_input("Pipe length (m)", value=50.0, min_value=0.1, max_value=10000.0, step=5.0, format="%.1f", key="fn_L")

            _fn_run = st.button("Calculate", type="primary", width='stretch', key="fn_run")

        with _fr:
            st.subheader("Results")

            if not _fn_run:
                st.info("Set inputs and press **Calculate**.")
            else:
                try:
                    _fn_res = fanno.fanno_solve(
                        P1_Pa      = _fn_P1_bara * 1e5,
                        T1_K       = _fn_T1_C + 273.15,
                        mdot_kgs   = _fn_mdot_kgh / 3600.0,
                        D_m        = _fn_D_m,
                        L_m        = _fn_L_m,
                        roughness_m= _fn_rough,
                        MW_kgmol   = _fn_mw,
                        gamma      = _fn_gamma,
                        species    = _fn_species,
                    )
                except Exception as _fn_exc:
                    st.error(f"Solver error: {_fn_exc}")
                    st.stop()

                _fn_Ma1 = _fn_res["Ma1"]
                _fn_Ma2 = _fn_res["Ma2"]

                # Warnings
                if _fn_res["choked"]:
                    st.error(
                        f"**Choked flow** — pipe length ({_fn_L_m:.1f} m) exceeds critical length "
                        f"L\* = {_fn_res['L_star_m']:.1f} m. Throughput is limited. "
                        "Reduce pipe length, increase diameter, or reduce mass flow rate."
                    )
                elif _fn_res["margin_pct"] < 20.0:
                    st.warning(
                        f"Near-choked: {_fn_res['margin_pct']:.1f} % margin to choking "
                        f"(L\* = {_fn_res['L_star_m']:.1f} m)."
                    )

                if _fn_Ma1 < 0.1:
                    st.warning(
                        f"Inlet Ma = {_fn_Ma1:.4f} < 0.1. Use the gas-only Darcy-Weisbach mode in tabs A / B."
                    )
                elif _fn_Ma1 < 0.3:
                    st.info(f"Inlet Ma = {_fn_Ma1:.3f} (0.1–0.3). Fanno and isothermal D-W agree within ~5 %.")

                if _fn_Ma1 > 1.0:
                    st.info(
                        f"Supersonic inlet (Ma = {_fn_Ma1:.3f}). Friction decelerates toward Ma = 1. "
                        "Verify that the inlet is truly supersonic."
                    )

                # Key metrics
                _fm1, _fm2, _fm3, _fm4, _fm5 = st.columns(5)
                _fm1.metric("Ma₁ (inlet)",    f"{_fn_Ma1:.4f}")
                _fm2.metric("Ma₂ (exit)",     f"{_fn_Ma2:.4f}")
                _fm3.metric("ΔP static",      f"{_fn_res['dP_static_kPa']:.2f} kPa")
                _fm4.metric("ΔP stagnation",  f"{_fn_res['dP_stag_kPa']:.2f} kPa")
                _fm5.metric("L / L*",
                    f"{_fn_L_m / _fn_res['L_star_m'] * 100:.1f} %" if _fn_res['L_star_m'] > 0 else "∞")

                # Condition tables
                _ft1, _ft2 = st.columns(2)
                with _ft1:
                    st.markdown("**Inlet**")
                    st.table({
                        "Property": ["P static (bara)", "T static (°C)", "P₀ (bara)", "V (m/s)", "a (m/s)", "Ma", "Re", "f (Churchill)"],
                        "Value": [
                            f"{_fn_res['P1_bara']:.4f}",
                            f"{_fn_res['T1_K'] - 273.15:.2f}",
                            f"{_fn_res['P01_bara']:.4f}",
                            f"{_fn_res['V1_ms']:.2f}",
                            f"{_fn_res['a1_ms']:.2f}",
                            f"{_fn_Ma1:.5f}",
                            f"{_fn_res['Re1']:.2e}",
                            f"{_fn_res['f1']:.5f}",
                        ],
                    })
                with _ft2:
                    st.markdown("**Exit**")
                    st.table({
                        "Property": ["P static (bara)", "T static (°C)", "P₀ (bara)", "V (m/s)", "a (m/s)", "Ma", "L* (m)", "Margin to choke"],
                        "Value": [
                            f"{_fn_res['P2_bara']:.4f}",
                            f"{_fn_res['T2_K'] - 273.15:.2f}",
                            f"{_fn_res['P02_bara']:.4f}",
                            f"{_fn_res['V2_ms']:.2f}",
                            f"{_fn_res['a2_ms']:.2f}",
                            f"{_fn_Ma2:.5f}",
                            f"{_fn_res['L_star_m']:.2f}",
                            f"{_fn_res['margin_pct']:.1f} %" if not _fn_res['choked'] else "CHOKED",
                        ],
                    })

                # Charts
                _fn_x  = _fn_res["x_arr"]
                _fn_Ma = _fn_res["Ma_arr"]
                _fn_P  = _fn_res["P_arr_bara"]
                _fn_T  = _fn_res["T_arr_K"] - 273.15
                _fn_P0 = _fn_res["P0_arr_bara"]
                _fn_V  = _fn_res["V_arr_ms"]

                _fig1 = make_subplots(specs=[[{"secondary_y": True}]])
                _fig1.add_trace(go.Scatter(x=_fn_x, y=_fn_P,  name="P static (bara)",       line=dict(color="#2563eb", width=2)),         secondary_y=False)
                _fig1.add_trace(go.Scatter(x=_fn_x, y=_fn_P0, name="P₀ stagnation (bara)",  line=dict(color="#2563eb", width=1.5, dash="dash")), secondary_y=False)
                _fig1.add_trace(go.Scatter(x=_fn_x, y=_fn_T,  name="T static (°C)",         line=dict(color="#dc2626", width=2)),         secondary_y=True)
                _fig1.update_xaxes(title_text="Distance from inlet (m)")
                _fig1.update_yaxes(title_text="Pressure (bara)", secondary_y=False)
                _fig1.update_yaxes(title_text="Temperature (°C)", secondary_y=True)
                _fig1.update_layout(title="Pressure and Temperature Profile",
                                    legend=dict(orientation="h", yanchor="bottom", y=1.02),
                                    margin=dict(t=60, b=40), height=320)
                st.plotly_chart(_fig1, width='stretch')

                _fig2 = go.Figure()
                _fig2.add_trace(go.Scatter(x=_fn_x, y=_fn_Ma, name="Ma", line=dict(color="#16a34a", width=2.5)))
                _fig2.add_hline(y=1.0, line_dash="dot", line_color="gray", annotation_text="Ma = 1 (sonic)")
                if not _fn_res["choked"]:
                    _fig2.add_vline(x=_fn_res["L_star_m"], line_dash="dash", line_color="orange",
                                    annotation_text=f"L* = {_fn_res['L_star_m']:.1f} m",
                                    annotation_position="top right")
                _fig2.update_xaxes(title_text="Distance from inlet (m)")
                _fig2.update_yaxes(title_text="Mach number (−)")
                _fig2.update_layout(title="Mach Number Profile", margin=dict(t=60, b=40), height=280)
                st.plotly_chart(_fig2, width='stretch')

                _fig3 = go.Figure()
                _fig3.add_trace(go.Scatter(x=_fn_x, y=_fn_V, name="Velocity (m/s)", line=dict(color="#7c3aed", width=2)))
                _fig3.update_xaxes(title_text="Distance from inlet (m)")
                _fig3.update_yaxes(title_text="Velocity (m/s)")
                _fig3.update_layout(title="Velocity Profile", margin=dict(t=60, b=40), height=240)
                st.plotly_chart(_fig3, width='stretch')

                with st.expander("Assumptions and limitations"):
                    st.markdown(f"""
    - **Ideal gas**: ρ = P·MW/(Rᵤ·T). Non-ideal effects (Z ≠ 1) ignored; accuracy degrades above Pr ≈ 0.5.
    - **Adiabatic duct**: no heat transfer. Heat exchange shifts the critical length vs. this prediction.
    - **Constant cross-section**: fittings, reducers, and valves not modelled.
    - **Friction factor** by Churchill (1977) at inlet Re = {_fn_res['Re1']:.2e}, held constant along duct.
    - **γ = {_fn_gamma:.3f}, MW = {_fn_mw*1000:.3f} g/mol** constant (calorically perfect gas).
    - Stagnation enthalpy conserved (adiabatic); stagnation temperature T₀ is constant.
    - Supersonic inlet results assume the inlet is truly supersonic; upstream shock structure not checked.
    """)

                st.divider()
                _fn_rpt = report_generator.generate_calculator_report(
                    tool_name="Fanno Flow",
                    subtitle=f"{_fn_species}  ·  {_fn_dn}/{_fn_pn}  ·  {_fn_L_m:.1f} m",
                    method_text=(
                        "Adiabatic compressible duct flow (Fanno line) for a constant-area pipe "
                        "with wall friction. The inlet Mach number is derived from the specified "
                        "mass flow, pipe bore, and thermodynamic state. Friction causes subsonic "
                        "flow to accelerate toward Ma = 1 (sonic limit). The critical length L* "
                        "is the maximum pipe length at which the flow remains unchoked for the "
                        "given inlet conditions. Friction factor from Churchill (1977). "
                        "Ideal-gas law; calorically perfect gas (constant γ and MW)."
                    ),
                    inputs_rows=[
                        ("Species",          _fn_species),
                        ("MW (g/mol)",        f"{_fn_mw*1000:.3f}"),
                        ("γ (Cp/Cv)",         f"{_fn_gamma:.4f}"),
                        ("Inlet pressure P₁", f"{_fn_P1_bara:.3f} bara"),
                        ("Inlet temperature", f"{_fn_T1_C:.1f} °C"),
                        ("Mass flow",         f"{_fn_mdot_kgh:.3f} kg/h"),
                        ("Pipe",              f"{_fn_dn} / {_fn_pn}  (ID = {_fn_D_m*1000:.2f} mm)"),
                        ("Material",          _fn_mat if _fn_liner == "None" else f"{_fn_mat} + {_fn_liner} liner"),
                        ("Roughness ε",       f"{_fn_rough:.2e} m"),
                        ("Pipe length L",     f"{_fn_L_m:.2f} m"),
                    ],
                    results_rows=[
                        ("Inlet Ma₁",              f"{_fn_res['Ma1']:.5f}"),
                        ("Exit Ma₂",               f"{_fn_res['Ma2']:.5f}"),
                        ("P₁ static (bara)",        f"{_fn_res['P1_bara']:.4f}"),
                        ("T₁ static (°C)",          f"{_fn_res['T1_K']-273.15:.2f}"),
                        ("P₀₁ stagnation (bara)",   f"{_fn_res['P01_bara']:.4f}"),
                        ("V₁ (m/s)",               f"{_fn_res['V1_ms']:.2f}"),
                        ("a₁ speed of sound (m/s)", f"{_fn_res['a1_ms']:.2f}"),
                        ("Re₁",                    f"{_fn_res['Re1']:.3e}"),
                        ("f (Churchill)",           f"{_fn_res['f1']:.5f}"),
                        ("P₂ static (bara)",        f"{_fn_res['P2_bara']:.4f}"),
                        ("T₂ static (°C)",          f"{_fn_res['T2_K']-273.15:.2f}"),
                        ("P₀₂ stagnation (bara)",   f"{_fn_res['P02_bara']:.4f}"),
                        ("V₂ (m/s)",               f"{_fn_res['V2_ms']:.2f}"),
                        ("ΔP static (kPa)",         f"{_fn_res['dP_static_kPa']:.3f}"),
                        ("ΔP stagnation (kPa)",     f"{_fn_res['dP_stag_kPa']:.3f}"),
                        ("Critical length L* (m)",  f"{_fn_res['L_star_m']:.2f}"),
                        ("L / L* (%)",              f"{_fn_L_m/_fn_res['L_star_m']*100:.1f} %"
                                                    if _fn_res['L_star_m'] > 0 else "∞"),
                        ("Margin to choke",         f"{_fn_res['margin_pct']:.1f} %"
                                                    if not _fn_res['choked'] else "CHOKED"),
                    ],
                    fig=_fig1,
                    fig_caption_text="Pressure and temperature profiles along the duct.",
                    fig_height=340,
                )
                st.download_button(
                    "Export Word (.docx)",
                    _fn_rpt,
                    file_name=f"fanno_{_fn_species}_{_fn_dn}.docx",
                    mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                    key="fn_dl",
                )

    # =============================================================================
    # Tab: Restriction Orifice (RO)
    # =============================================================================
    with tab_ro:
        st.markdown(
            "Single-phase restriction orifice sizing and rating — "
            "gas (compressible, choke detection) and liquid (cavitation check). "
            "ISO 5167-2 discharge coefficient · API 520 critical flow · IEC 60534 cavitation index."
        )

        # ── Top controls ──────────────────────────────────────────────────────────
        _ro_c1, _ro_c2 = st.columns(2)
        _ro_mode = _ro_c1.radio("Fluid", ["Gas", "Liquid"], horizontal=True, key="ro_mode")
        _ro_dir  = _ro_c2.radio("Calculation", ["Size bore  (flow → bore)", "Rate bore  (bore → flow)"], horizontal=True, key="ro_dir")
        _ro_size_mode = _ro_dir.startswith("Size")

        st.divider()
        _ro_left, _ro_right = st.columns([1, 1.6], gap="large")

        # ── Inputs ────────────────────────────────────────────────────────────────
        with _ro_left:
            st.subheader("Inputs")

            # Pressures (common to both modes)
            _rc1, _rc2 = st.columns(2)
            _ro_P1 = _rc1.number_input("P₁ upstream (bara)", value=10.0, min_value=0.01, max_value=1000.0, step=0.5, format="%.2f", key="ro_P1")
            _ro_P2 = _rc2.number_input("P₂ downstream (bara)", value=5.0, min_value=0.01, max_value=1000.0, step=0.5, format="%.2f", key="ro_P2")

            if _ro_P2 >= _ro_P1:
                st.warning("P₂ must be less than P₁.")

            # Pipe selector (common)
            st.markdown("**Pipe (upstream)**")
            _rc1, _rc2 = st.columns(2)
            _ro_dn  = _rc1.selectbox("DN", list(engine.PIPE_DATABASE.keys()), index=4, key="ro_dn")
            _ro_pn  = _rc2.selectbox("PN", ["PN20", "PN25", "PN40"], index=0, key="ro_pn")
            _ro_D_m = engine.PIPE_DATABASE[_ro_dn][_ro_pn]
            st.caption(f"Pipe ID = {_ro_D_m*1000:.1f} mm")

            # ── Gas-specific inputs ────────────────────────────────────────────
            if _ro_mode == "Gas":
                st.markdown("**Gas**")
                _ro_species = st.selectbox("Species", list(ro.GASES.keys()), index=0, key="ro_species")
                _ro_mw0, _ro_g0, _ = ro.GASES[_ro_species]
                if _ro_species == "Custom":
                    _rc1, _rc2 = st.columns(2)
                    _ro_mw_gmol = _rc1.number_input("MW (g/mol)", value=28.0, min_value=1.0, max_value=300.0, key="ro_mw")
                    _ro_gamma   = _rc2.number_input("γ", value=1.40, min_value=1.01, max_value=1.80, format="%.3f", key="ro_gamma")
                    _ro_mw = _ro_mw_gmol / 1000.0
                else:
                    _ro_mw, _ro_gamma = _ro_mw0, _ro_g0
                    st.caption(f"MW = {_ro_mw*1000:.3f} g/mol  |  γ = {_ro_gamma:.3f}")

                _ro_T1 = st.number_input("T₁ (°C)", value=20.0, min_value=-50.0, max_value=500.0, step=1.0, format="%.1f", key="ro_T1")

                if _ro_size_mode:
                    _ro_mdot_kgh = st.number_input("Mass flow (kg/h)", value=500.0, min_value=0.01, max_value=1e6, step=10.0, format="%.1f", key="ro_mdot")
                else:
                    _ro_d_mm = st.number_input("Orifice bore d (mm)", value=15.0, min_value=1.0, max_value=_ro_D_m*990, step=0.5, format="%.1f", key="ro_d")

            # ── Liquid-specific inputs ─────────────────────────────────────────
            else:
                st.markdown("**Liquid**")
                _ro_cp_fluids = list(engine.LIQUID_COOLPROP_ID.keys())
                _ro_liq_opts  = _ro_cp_fluids + ["KOH solution"] + ["Custom"]
                _ro_liq_type  = st.selectbox("Fluid", _ro_liq_opts, index=0, key="ro_liq_type")

                if _ro_liq_type == "KOH solution":
                    _ro_T_liq = st.number_input("Temperature (°C)", value=40.0,
                                                 min_value=10.0, max_value=90.0,
                                                 step=1.0, format="%.1f", key="ro_T_liq_koh")
                    _ro_koh_conc = st.slider("Concentration (wt%)", min_value=5, max_value=40,
                                              value=30, step=5, key="ro_koh_conc",
                                              help="Gilliam et al. 2007 correlation; valid 0–40 wt%, 10–90 °C.")
                    _ro_lp  = ro.koh_liquid_properties(_ro_T_liq, float(_ro_koh_conc))
                    _ro_rho = _ro_lp["rho_kgm3"]
                    _ro_mu  = _ro_lp["mu_pas"]
                    _ro_Pv  = _ro_lp["Pv_Pa"]
                    st.caption(
                        f"ρ = {_ro_rho:.1f} kg/m³  |  μ = {_ro_mu*1e3:.3f} cP  |  "
                        f"Pᵥ = {_ro_Pv/1e5:.5f} bara  (water activity {1-ro._KOH_ACTIVITY_SLOPE*_ro_koh_conc:.3f})"
                    )
                elif _ro_liq_type != "Custom":
                    _ro_cp_id  = engine.LIQUID_COOLPROP_ID[_ro_liq_type]
                    _ro_T_liq  = st.number_input("Temperature (°C)", value=20.0,
                                                  min_value=-200.0, max_value=500.0,
                                                  step=1.0, format="%.1f", key="ro_T_liq")
                    try:
                        _ro_lp = ro.liquid_properties(_ro_cp_id, _ro_T_liq, _ro_P1 * 1e5)
                        _ro_rho = _ro_lp["rho_kgm3"]
                        _ro_mu  = _ro_lp["mu_pas"]
                        _ro_Pv  = _ro_lp["Pv_Pa"]
                        _pv_str = f"{_ro_Pv/1e5:.4f}" if _ro_Pv > 0 else "N/A (above Tᶜ)"
                        st.caption(f"ρ = {_ro_rho:.1f} kg/m³  |  μ = {_ro_mu*1e3:.4f} cP  |  Pᵥ = {_pv_str} bara")
                        if _ro_Pv > 0 and _ro_Pv > _ro_P1 * 1e5:
                            st.warning(
                                f"Pᵥ ({_ro_Pv/1e5:.2f} bara) > P₁ ({_ro_P1:.2f} bara) — "
                                f"**{_ro_liq_type} is a vapour at these conditions.** "
                                "Raise P₁ above Pᵥ, or lower the temperature."
                            )
                    except Exception as _lp_err:
                        st.error(f"CoolProp error: {_lp_err}")
                        _ro_rho, _ro_mu, _ro_Pv = 1000.0, 1e-3, 0.0
                else:
                    _rc1, _rc2, _rc3 = st.columns(3)
                    _ro_rho = _rc1.number_input("ρ (kg/m³)",  value=998.0, min_value=1.0,  max_value=3000.0, step=1.0,   format="%.1f",  key="ro_rho")
                    _ro_mu  = _rc2.number_input("μ (cP)",     value=1.0,   min_value=0.001, max_value=1e5,   step=0.1,   format="%.3f",  key="ro_mu") * 1e-3
                    _ro_Pv  = _rc3.number_input("Pᵥ (bara)",  value=0.023, min_value=0.0,   max_value=_ro_P1, step=0.001, format="%.4f", key="ro_Pv") * 1e5

                if _ro_size_mode:
                    _ro_mdot_kgh = st.number_input("Mass flow (kg/h)", value=50.0, min_value=0.01, max_value=1e6, step=1.0, format="%.2f", key="ro_mdot_liq")
                else:
                    _ro_d_mm = st.number_input("Orifice bore d (mm)", value=15.0, min_value=1.0, max_value=_ro_D_m*990, step=0.5, format="%.1f", key="ro_d_liq")

            _ro_run = st.button("Calculate", type="primary", width='stretch', key="ro_run")

        # ── Results ───────────────────────────────────────────────────────────────
        with _ro_right:
            st.subheader("Results")

            if not _ro_run or _ro_P2 >= _ro_P1:
                st.info("Set inputs and press **Calculate**.")
            else:
                try:
                    if _ro_mode == "Gas":
                        if _ro_size_mode:
                            _ro_res = ro.ro_gas_size(
                                _ro_P1*1e5, _ro_P2*1e5, _ro_T1+273.15,
                                _ro_mw, _ro_gamma, _ro_species,
                                _ro_mdot_kgh/3600.0, _ro_D_m,
                            )
                            _ro_d_result = _ro_res["d_m"] * 1000
                            _ro_flow_kgh = _ro_mdot_kgh
                        else:
                            _ro_res = ro.ro_gas_rate(
                                _ro_P1*1e5, _ro_P2*1e5, _ro_T1+273.15,
                                _ro_mw, _ro_gamma, _ro_species,
                                _ro_d_mm/1000.0, _ro_D_m,
                            )
                            _ro_d_result = _ro_d_mm
                            _ro_flow_kgh = _ro_res["mdot_kgs"] * 3600.0
                    else:
                        if _ro_size_mode:
                            _ro_res = ro.ro_liquid_size(
                                _ro_P1*1e5, _ro_P2*1e5,
                                _ro_rho, _ro_mu, _ro_Pv,
                                _ro_mdot_kgh/3600.0, _ro_D_m,
                            )
                            _ro_d_result = _ro_res["d_m"] * 1000
                            _ro_flow_kgh = _ro_mdot_kgh
                        else:
                            _ro_res = ro.ro_liquid_rate(
                                _ro_P1*1e5, _ro_P2*1e5,
                                _ro_rho, _ro_mu, _ro_Pv,
                                _ro_d_mm/1000.0, _ro_D_m,
                            )
                            _ro_d_result = _ro_d_mm
                            _ro_flow_kgh = _ro_res["mdot_kgs"] * 3600.0

                except Exception as _ro_exc:
                    st.error(f"Solver error: {_ro_exc}")
                    st.stop()

                # ── Warnings ──────────────────────────────────────────────────
                if _ro_mode == "Gas":
                    _ro_r_c = _ro_res["r_c"]
                    if _ro_res["choked"]:
                        st.error(
                            f"**Single-stage choked** — P₂/P₁ = {_ro_P2/_ro_P1:.3f} ≤ r_c = {_ro_r_c:.3f}. "
                            "The single orifice operates at critical (sonic) flow. "
                            "See multi-stage recommendation below."
                        )
                    elif _ro_res["Ma_throat"] > 0.85:
                        st.warning(
                            f"High throat Mach number (Ma = {_ro_res['Ma_throat']:.3f}). "
                            "Choking margin is narrow; noise and erosion risk."
                        )
                    if _ro_res["beta"] < 0.10:
                        st.warning(f"β = {_ro_res['beta']:.3f} < 0.10 — outside ISO 5167 validity. Cd uncertainty ↑.")
                    elif _ro_res["beta"] > 0.75:
                        st.warning(f"β = {_ro_res['beta']:.3f} > 0.75 — outside ISO 5167 validity. Consider larger pipe.")
                else:
                    _ro_Kc = _ro_res["Kc"]
                    if _ro_res["choked"]:
                        st.error("**Liquid choke** — P₂ ≤ Pᵥ. Fluid will flash. Use multi-stage.")
                    elif _ro_res.get("cavitation_severe"):
                        st.error(
                            f"**Severe cavitation** — Kc = {_ro_Kc:.2f} ≥ {ro.KC_SEVERE}. "
                            "Heavy erosion risk. Multi-stage strongly recommended."
                        )
                    elif _ro_res["cavitating"]:
                        st.warning(
                            f"**Incipient cavitation** — Kc = {_ro_Kc:.2f} ≥ {ro.KC_INCIPIENT}. "
                            "Noise and surface erosion risk. Consider multi-stage."
                        )

                # ── Key metrics ────────────────────────────────────────────────
                _rm1, _rm2, _rm3, _rm4 = st.columns(4)
                _rm1.metric("Bore d", f"{_ro_d_result:.2f} mm")
                _rm2.metric("β = d/D", f"{_ro_res['beta']:.4f}")
                _rm3.metric("C (ISO 5167)", f"{_ro_res['C']:.4f}")
                _rm4.metric("Flow", f"{_ro_flow_kgh:.1f} kg/h")

                if _ro_mode == "Gas":
                    _rm5, _rm6, _rm7, _rm8 = st.columns(4)
                    _rm5.metric("Ma throat", f"{_ro_res['Ma_throat']:.4f}")
                    _rm6.metric("ε (expansion)", f"{_ro_res['eps']:.4f}" if not _ro_res['choked'] else "N/A (choked)")
                    _rm7.metric("ρ₁ (kg/m³)", f"{_ro_res['rho1_kgm3']:.3f}")
                    _rm8.metric("Re_D", f"{_ro_res['Re_D']:.2e}")
                else:
                    _rm5, _rm6 = st.columns(2)
                    _rm5.metric("Kc (cavitation)", f"{_ro_res['Kc']:.3f}" if _ro_res['Kc'] != float('inf') else "∞")
                    _rm6.metric("Re_D", f"{_ro_res['Re_D']:.2e}")

                # ── Multi-stage section ────────────────────────────────────────
                st.divider()
                st.markdown("#### Multi-stage analysis")

                if _ro_mode == "Gas":
                    _ms = ro.multistage_gas(
                        _ro_P1*1e5, _ro_P2*1e5, _ro_T1+273.15,
                        _ro_mw, _ro_gamma, _ro_species,
                        _ro_flow_kgh/3600.0, _ro_D_m,
                    )
                    if _ms["N"] == 1:
                        st.success(
                            f"Single stage is adequate — P₂/P₁ = {_ro_P2/_ro_P1:.3f} > "
                            f"r_c × 0.90 = {_ms['r_c_design']:.3f}."
                        )
                    else:
                        st.info(
                            f"**{_ms['N']} stages** recommended — each stage ratio "
                            f"{_ms['r_stage']:.3f} > r_c×0.90 = {_ms['r_c_design']:.3f}."
                        )

                    import pandas as pd
                    _ms_df = pd.DataFrame([{
                        "Stage": s["stage"],
                        "P_in (bara)":  f"{s['P_in_bara']:.3f}",
                        "P_out (bara)": f"{s['P_out_bara']:.3f}",
                        "ΔP (bar)":     f"{s['dP_bar']:.3f}",
                        "Bore (mm)":    f"{s['d_mm']:.2f}",
                        "β":            f"{s['beta']:.4f}",
                        "C":            f"{s['C']:.4f}",
                        "ε":            f"{s['eps']:.4f}",
                        "Ma throat":    f"{s['Ma_throat']:.4f}",
                    } for s in _ms["stages"]])
                    st.dataframe(_ms_df, width='stretch', hide_index=True)

                else:
                    _ms = ro.multistage_liquid(
                        _ro_P1*1e5, _ro_P2*1e5,
                        _ro_rho, _ro_mu, _ro_Pv,
                        _ro_flow_kgh/3600.0, _ro_D_m,
                    )
                    if _ms["N"] == 1:
                        st.success(
                            f"Single stage is adequate — Kc = {_ms['Kc_single']:.3f} < "
                            f"limit {_ms['kc_limit']:.2f}."
                        )
                    else:
                        st.info(
                            f"**{_ms['N']} stages** recommended — single-stage "
                            f"Kc = {_ms['Kc_single']:.2f} ≥ limit {_ms['kc_limit']:.2f}."
                        )

                    import pandas as pd
                    _ms_df = pd.DataFrame([{
                        "Stage":        s["stage"],
                        "P_in (bara)":  f"{s['P_in_bara']:.3f}",
                        "P_out (bara)": f"{s['P_out_bara']:.3f}",
                        "ΔP (bar)":     f"{s['dP_bar']:.3f}",
                        "Bore (mm)":    f"{s['d_mm']:.2f}",
                        "β":            f"{s['beta']:.4f}",
                        "C":            f"{s['C']:.4f}",
                        "Kc":           f"{s['Kc']:.4f}",
                    } for s in _ms["stages"]])
                    st.dataframe(_ms_df, width='stretch', hide_index=True)

                # ── Pressure profile chart ─────────────────────────────────────
                _P_sched = _ms["P_schedule_bara"]
                _stages_x = []
                _stages_y = []
                for _i, (_pin, _pout) in enumerate(zip(_P_sched[:-1], _P_sched[1:])):
                    _stages_x += [_i, _i, _i + 1, _i + 1]
                    _stages_y += [None, _pin, _pout, None]

                _pfig = go.Figure()
                _pfig.add_trace(go.Scatter(
                    x=list(range(len(_P_sched))),
                    y=_P_sched,
                    mode="markers+lines",
                    marker=dict(size=10, color="#2563eb"),
                    line=dict(color="#2563eb", width=2, dash="dot"),
                    name="Stage boundary pressure",
                ))
                # Horizontal spans per stage
                for _i, _stg in enumerate(_ms["stages"]):
                    _pfig.add_shape(
                        type="rect",
                        x0=_i, x1=_i + 1,
                        y0=_stg["P_out_bara"], y1=_stg["P_in_bara"],
                        fillcolor="#DBEAFE", opacity=0.3, line_width=0,
                    )
                if _ro_P2 / _ro_P1 <= ro.critical_pressure_ratio(_ro_gamma if _ro_mode == "Gas" else 1.4):
                    _pfig.add_hline(y=_ro_P1 * ro.critical_pressure_ratio(_ro_gamma if _ro_mode == "Gas" else 1.4),
                                    line_dash="dot", line_color="red",
                                    annotation_text="choke limit (single stage)")
                _pfig.update_xaxes(
                    title_text="Stage number",
                    tickvals=list(range(len(_P_sched))),
                    ticktext=[f"Inlet"] + [f"After S{k+1}" for k in range(len(_P_sched)-1)],
                )
                _pfig.update_yaxes(title_text="Pressure (bara)")
                _pfig.update_layout(
                    title="Pressure Drop Distribution Across Stages",
                    margin=dict(t=50, b=40), height=320,
                    showlegend=False,
                )
                st.plotly_chart(_pfig, width='stretch')

                with st.expander("Assumptions and references"):
                    st.markdown(f"""
    **Discharge coefficient C** — ISO 5167-2:2022, Reader-Harris/Gallagher equation, corner taps.
    Valid for β ∈ [0.10, 0.75], Re_D ≥ 5 000. Outside this range: C = 0.60 (conservative default).

    **Expansion factor ε** — ISO 5167-2:2022, corner-tap formula.
    Not applicable above the critical pressure ratio (choked gas).

    **Choked gas flow** — API RP 520 Pt I / isentropic critical-flow formula.
    r_c = (2/(γ+1))^(γ/(γ−1)) = **{ro.critical_pressure_ratio(_ro_gamma if _ro_mode == "Gas" else 1.4):.4f}** for the selected gas.
    Multi-stage gas: equal pressure-ratio per stage with 10 % choking margin.

    **Cavitation index** — IEC 60534-2 convention: Kc = ΔP / (P₁ − Pᵥ).
    Incipient: Kc ≥ {ro.KC_INCIPIENT:.2f} | Severe: Kc ≥ {ro.KC_SEVERE:.2f} | Choke: P₂ ≤ Pᵥ.
    Multi-stage liquid: equal ΔP per stage; minimum N so last-stage Kc < {ro.KC_INCIPIENT:.2f}.

    **Bore tolerance** — standard orifice bore tolerance per ASME B16.36 / ISO 5167 is typically ±0.05 mm.
    Round calculated bore to nearest 0.5 mm (or preferred drill size) before ordering.
    """)

                # ── Word export ────────────────────────────────────────────────
                st.divider()
                _ro_fluid_lbl = (
                    f"{_ro_species}  (MW={_ro_mw*1000:.1f} g/mol, γ={_ro_gamma:.3f})"
                    if _ro_mode == "Gas"
                    else getattr(_ro_liq_type, "__str__", lambda: str(_ro_liq_type))()
                )
                _ro_dir_lbl = "Size (flow → bore)" if _ro_size_mode else "Rate (bore → flow)"
                _ro_inp = [
                    ("Fluid",             f"{_ro_mode} — {_ro_fluid_lbl}"),
                    ("Calculation",       _ro_dir_lbl),
                    ("P₁ upstream",       f"{_ro_P1:.3f} bara"),
                    ("P₂ downstream",     f"{_ro_P2:.3f} bara"),
                    ("ΔP",                f"{(_ro_P1-_ro_P2)*100:.2f} kPa"),
                    ("Pipe",              f"{_ro_dn}/{_ro_pn}  (ID = {_ro_D_m*1000:.1f} mm)"),
                ]
                if _ro_mode == "Gas":
                    _ro_inp += [
                        ("Temperature T₁", f"{_ro_T1:.1f} °C"),
                        ("Flow / bore",    f"{_ro_mdot_kgh:.2f} kg/h" if _ro_size_mode
                                           else f"{_ro_d_mm:.2f} mm"),
                    ]
                else:
                    _ro_inp += [
                        ("Density ρ",   f"{_ro_rho:.2f} kg/m³"),
                        ("Viscosity μ", f"{_ro_mu*1e3:.4f} mPa·s"),
                        ("Flow / bore", f"{_ro_mdot_kgh:.2f} kg/h" if _ro_size_mode
                                        else f"{_ro_d_mm:.2f} mm"),
                    ]
                _ro_res_rows = [
                    ("Orifice bore d",  f"{_ro_d_result:.3f} mm"),
                    ("β = d/D",         f"{_ro_res['beta']:.5f}"),
                    ("C (ISO 5167)",    f"{_ro_res['C']:.5f}"),
                    ("Flow",           f"{_ro_flow_kgh:.3f} kg/h"),
                    ("Re_D",           f"{_ro_res['Re_D']:.3e}"),
                ]
                if _ro_mode == "Gas":
                    _ro_res_rows += [
                        ("Ma throat",        f"{_ro_res['Ma_throat']:.5f}"),
                        ("ε (expansion)",    f"{_ro_res['eps']:.5f}"
                                             if not _ro_res['choked'] else "N/A (choked)"),
                        ("Choked?",          "Yes" if _ro_res['choked'] else "No"),
                        ("r_c (critical)",   f"{_ro_res['r_c']:.5f}"),
                    ]
                else:
                    _kc_v = _ro_res['Kc']
                    _ro_res_rows += [
                        ("Kc (cavitation)", f"{_kc_v:.4f}" if _kc_v != float('inf') else "∞"),
                        ("Cavitating?",     "Yes" if _ro_res['cavitating'] else "No"),
                        ("Choked?",         "Yes" if _ro_res['choked'] else "No"),
                    ]
                _ro_ms_data = []
                if _ro_mode == "Gas":
                    _ro_ms_hdrs = ["Stage","P_in (bara)","P_out (bara)","ΔP (bar)","Bore (mm)","β","C","ε","Ma throat"]
                    for _s in _ms["stages"]:
                        _ro_ms_data.append([
                            str(_s["stage"]),
                            f"{_s['P_in_bara']:.3f}", f"{_s['P_out_bara']:.3f}",
                            f"{_s['dP_bar']:.3f}", f"{_s['d_mm']:.2f}",
                            f"{_s['beta']:.4f}", f"{_s['C']:.4f}",
                            f"{_s['eps']:.4f}", f"{_s['Ma_throat']:.4f}",
                        ])
                    _ro_ms_cw = [0.4, 0.75, 0.75, 0.65, 0.65, 0.55, 0.55, 0.55, 0.65]
                else:
                    _ro_ms_hdrs = ["Stage","P_in (bara)","P_out (bara)","ΔP (bar)","Bore (mm)","β","C","Kc"]
                    for _s in _ms["stages"]:
                        _ro_ms_data.append([
                            str(_s["stage"]),
                            f"{_s['P_in_bara']:.3f}", f"{_s['P_out_bara']:.3f}",
                            f"{_s['dP_bar']:.3f}", f"{_s['d_mm']:.2f}",
                            f"{_s['beta']:.4f}", f"{_s['C']:.4f}",
                            f"{_s['Kc']:.4f}",
                        ])
                    _ro_ms_cw = [0.4, 0.75, 0.75, 0.65, 0.65, 0.55, 0.55, 0.65]
                _ro_rpt = report_generator.generate_calculator_report(
                    tool_name="Restriction Orifice",
                    subtitle=f"{_ro_mode}  ·  {_ro_dir_lbl}  ·  {_ro_dn}/{_ro_pn}",
                    method_text=(
                        "ISO 5167-2:2022 orifice plate sizing with the Reader-Harris/Gallagher "
                        "discharge coefficient. For gas service, the expansion factor ε accounts "
                        "for compressibility; critical (choked) flow is detected when the "
                        "pressure ratio P₂/P₁ falls below the isentropic critical ratio r_c. "
                        "For liquid service, the cavitation index Kc = ΔP/(P₁−Pᵥ) is evaluated "
                        "against ISO 5167 / IEC 60534 thresholds. "
                        f"Multi-stage analysis: {_ms['N']} stage(s) recommended."
                    ),
                    inputs_rows=_ro_inp,
                    results_rows=_ro_res_rows,
                    extra_tables=[{
                        "title": "Multi-stage Analysis",
                        "headers": _ro_ms_hdrs,
                        "data":    _ro_ms_data,
                        "col_widths": _ro_ms_cw,
                    }],
                    fig=_pfig,
                    fig_caption_text="Pressure schedule across stages.",
                    fig_height=340,
                )
                st.download_button(
                    "Export Word (.docx)",
                    _ro_rpt,
                    file_name=f"ro_{_ro_mode.lower()}_{_ro_dn}.docx",
                    mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                    key="ro_dl",
                )

    # =============================================================================
    # Tab: PSV — API 520 Part I / API 526
    # =============================================================================
    with tab_psv:
        st.markdown(
            "Pressure Safety Valve sizing per **API 520 Part I** (SI). "
            "Standard orifice selection per **API 526**."
        )

        _psv_c1, _psv_c2 = st.columns([1, 2], gap="large")

        with _psv_c1:
            st.markdown("**SERVICE**")
            _psv_service = st.selectbox(
                "Service type", ["Gas / Vapour", "Steam", "Liquid"], key="psv_service"
            )
            _psv_type = st.selectbox(
                "PSV type",
                ["Conventional (spring-loaded)", "Balanced bellows", "Pilot-operated"],
                key="psv_type",
            )

            st.markdown("**INLET CONDITIONS**")
            _psv_Pset_barg = st.number_input(
                "Set pressure (barg)", value=10.0, min_value=0.1, step=0.5, key="psv_Pset"
            )
            _psv_op_pct = st.number_input(
                "Allowable overpressure (%)", value=10.0, min_value=1.0, max_value=21.0,
                step=1.0, key="psv_op_pct",
                help="10 % normal fire case; 21 % for fire + ASME Sec VIII.",
            )
            _psv_Patm_bara = st.number_input(
                "Atmospheric pressure (bara)", value=1.01325, step=0.005, key="psv_Patm",
                format="%.4f",
            )
            _psv_Pback_barg = st.number_input(
                "Back pressure (barg)", value=0.0, min_value=0.0, step=0.5, key="psv_Pback"
            )
            _psv_T_C = st.number_input(
                "Relieving temperature (°C)", value=50.0, step=5.0, key="psv_T_C"
            )

            # Derived pressures
            _psv_P1_bara = _psv_Pset_barg * (1.0 + _psv_op_pct / 100.0) + _psv_Patm_bara
            _psv_P1_kPa  = _psv_P1_bara * 100.0
            _psv_Pback_kPa = _psv_Pback_barg * 100.0 + _psv_Patm_bara * 100.0
            _psv_T1_K    = _psv_T_C + 273.15

            st.caption(
                f"Relieving pressure P₁ = {_psv_P1_bara:.3f} bara = {_psv_P1_kPa:.1f} kPa abs"
            )

            st.markdown("**REQUIRED RELIEF RATE**")
            if _psv_service != "Liquid":
                _psv_W_kgh = st.number_input(
                    "Mass flow (kg/h)", value=1000.0, min_value=0.1, step=100.0, key="psv_W"
                )
                _psv_Q_m3h = None
            else:
                _psv_Q_m3h = st.number_input(
                    "Volumetric flow (m³/h)", value=10.0, min_value=0.001, step=1.0, key="psv_Q"
                )
                _psv_W_kgh = None

            st.markdown("**CORRECTIONS**")
            _psv_rupture_disc = st.checkbox(
                "Rupture disc in series upstream",
                value=False, key="psv_disc",
                help="Applies Kc = 0.9 per API 520 §4.7",
            )
            _psv_Kc = psv.KC_DISC if _psv_rupture_disc else psv.KC_NONE

        # ── Fluid properties (right side top) ──────────────────────────────────
        with _psv_c2:
            if _psv_service == "Gas / Vapour":
                st.markdown("**GAS PROPERTIES**")
                _psv_species = st.selectbox(
                    "Species", list(fanno.GASES.keys()), key="psv_species"
                )
                _psv_cp_name   = fanno.GASES[_psv_species][2]   # CoolProp ID or None
                _psv_MW_kgmol  = fanno.GASES[_psv_species][0]   # kg/mol or None
                _psv_gam_table = fanno.GASES[_psv_species][1]   # table γ or None

                if _psv_cp_name is not None:
                    # Known species: fetch γ and Z from CoolProp at relieving conditions
                    _psv_MW    = (_psv_MW_kgmol * 1000.0) if _psv_MW_kgmol else 28.97
                    _psv_gamma = _psv_gam_table or 1.40
                    _psv_Z     = 1.0
                    try:
                        import CoolProp.CoolProp as CP
                        _P_Pa_psv = _psv_P1_kPa * 1000.0
                        _Cp_v = CP.PropsSI("C", "T", _psv_T1_K, "P", _P_Pa_psv, _psv_cp_name)
                        _Cv_v = CP.PropsSI("O", "T", _psv_T1_K, "P", _P_Pa_psv, _psv_cp_name)
                        if _Cv_v > 0:
                            _psv_gamma = _Cp_v / _Cv_v
                        _rho_v = CP.PropsSI("D", "T", _psv_T1_K, "P", _P_Pa_psv, _psv_cp_name)
                        _psv_Z = _P_Pa_psv * _psv_MW / (_rho_v * 8314.46 * _psv_T1_K)
                        st.caption(
                            f"CoolProp at {_psv_T1_K-273.15:.0f} °C, {_psv_P1_kPa/100:.3f} bara — "
                            f"γ = {_psv_gamma:.4f}  |  Z = {_psv_Z:.4f}  |  MW = {_psv_MW:.3f} kg/kmol"
                        )
                    except Exception as _cp_err:
                        st.caption(
                            f"CoolProp unavailable ({_cp_err}) — using table: "
                            f"γ = {_psv_gamma:.4f}, Z = 1.0"
                        )
                else:
                    # Custom gas: manual inputs
                    _cg1, _cg2, _cg3 = st.columns(3)
                    _psv_MW    = _cg1.number_input("MW (kg/kmol)", value=28.97, min_value=1.0, step=0.5, key="psv_MW")
                    _psv_gamma = _cg2.number_input("γ (Cp/Cv)", value=1.40, min_value=1.01, max_value=2.0, step=0.01, key="psv_gamma")
                    _psv_Z     = _cg3.number_input("Z", value=1.0, min_value=0.1, max_value=2.0, step=0.01, key="psv_Z")

                # Backpressure correction
                if "Conventional" in _psv_type:
                    _psv_Kb = psv.kb_conventional(_psv_Pback_kPa, _psv_P1_kPa)
                elif "Balanced" in _psv_type:
                    _psv_Kb = psv.kb_balanced_bellows(_psv_Pback_kPa, _psv_P1_kPa, _psv_gamma)
                else:
                    _psv_Kb = psv.kb_pilot(_psv_Pback_kPa, _psv_P1_kPa)

                _psv_res = psv.psv_gas_size(
                    _psv_W_kgh, _psv_P1_kPa, _psv_T1_K, _psv_MW, _psv_gamma,
                    Kb=_psv_Kb, Kc=_psv_Kc, Z=_psv_Z,
                )

            elif _psv_service == "Steam":
                st.markdown("**STEAM PROPERTIES** — via CoolProp")
                st.caption("γ and Z are computed from CoolProp at the relieving conditions.")
                _psv_Z = 1.0

                if "Conventional" in _psv_type:
                    _psv_Kb = psv.kb_conventional(_psv_Pback_kPa, _psv_P1_kPa)
                elif "Balanced" in _psv_type:
                    _psv_Kb = psv.kb_balanced_bellows(_psv_Pback_kPa, _psv_P1_kPa, 1.33)
                else:
                    _psv_Kb = psv.kb_pilot(_psv_Pback_kPa, _psv_P1_kPa)

                _psv_res = psv.psv_steam_size(
                    _psv_W_kgh, _psv_P1_kPa, _psv_T1_K,
                    Kb=_psv_Kb, Kc=_psv_Kc,
                )
                for _note in _psv_res.get("notes", []):
                    if _note.startswith("UNPHYSICAL"):
                        st.error(_note)
                    else:
                        st.caption(_note)
                _psv_gamma = _psv_res.get("gamma", 1.33)
                _psv_MW    = _psv_res.get("MW", 18.015)

            else:  # Liquid
                st.markdown("**LIQUID PROPERTIES**")
                _liq_opts = list(engine.LIQUID_COOLPROP_ID.keys()) + ["KOH solution"] + ["Custom"]
                _liq_pick = st.selectbox("Fluid", _liq_opts, key="psv_liq_fluid")

                _ll1, _ll2, _ll3 = st.columns(3)
                if _liq_pick == "KOH solution":
                    _psv_koh_conc = st.slider("Concentration (wt%)", min_value=5, max_value=40,
                                               value=30, step=5, key="psv_koh_conc",
                                               help="Valid 0–40 wt%, 10–90 °C (Gilliam et al. 2007)")
                    _liq_props = ro.koh_liquid_properties(_psv_T_C, float(_psv_koh_conc))
                    _rho_def   = _liq_props["rho_kgm3"]
                    _mu_cP_def = _liq_props["mu_pas"] * 1e3
                    st.caption(
                        f"ρ = {_rho_def:.1f} kg/m³  |  μ = {_mu_cP_def:.3f} cP  |  "
                        f"Pᵥ = {_liq_props['Pv_Pa']/1e5:.5f} bara"
                    )
                elif _liq_pick != "Custom":
                    _liq_cp_id = engine.LIQUID_COOLPROP_ID[_liq_pick]
                    try:
                        _liq_props = ro.liquid_properties(_liq_cp_id, _psv_T_C, _psv_P1_bara * 1e5)
                        _rho_def   = _liq_props["rho_kgm3"]
                        _mu_cP_def = _liq_props["mu_pas"] * 1e3
                    except Exception:
                        _rho_def, _mu_cP_def = 1000.0, 1.0
                else:
                    _rho_def, _mu_cP_def = 1000.0, 1.0

                with _ll1:
                    _psv_rho = st.number_input(
                        "Density (kg/m³)", value=round(_rho_def, 2), min_value=1.0,
                        step=10.0, key="psv_rho"
                    )
                with _ll2:
                    _psv_mu_cP = st.number_input(
                        "Viscosity (cP)", value=round(_mu_cP_def, 3), min_value=0.001,
                        step=0.1, key="psv_mu"
                    )
                with _ll3:
                    _psv_Kw = 1.0 if "Conventional" in _psv_type else st.number_input(
                        "Kw (back-pressure)", value=1.0, min_value=0.5, max_value=1.0,
                        step=0.01, key="psv_Kw"
                    )

                _psv_P2_liq_kPa = _psv_Pback_kPa

                _psv_res = psv.psv_liquid_size(
                    _psv_Q_m3h, _psv_P1_kPa, _psv_P2_liq_kPa,
                    _psv_rho, _psv_mu_cP, Kw=_psv_Kw, Kc=_psv_Kc,
                )
                _psv_gamma = None
                _psv_MW    = None
                _psv_Kb    = _psv_Kw

            # ── Warnings ────────────────────────────────────────────────────────
            _bp_pct = psv.backpressure_pct(_psv_Pback_kPa, _psv_P1_kPa)
            if "Conventional" in _psv_type and _bp_pct > 10.0:
                st.warning(
                    f"Back pressure = {_bp_pct:.1f}% of relieving pressure P₁. "
                    "API 520 recommends consulting the valve vendor above 10% for conventional valves."
                )
            if "Balanced" in _psv_type and _bp_pct > 30.0:
                st.warning(
                    f"Back pressure = {_bp_pct:.1f}% of P₁. Kb correction applied. "
                    "Verify with valve vendor for back pressure > 30%."
                )
            if _psv_res["orifice_letter"] == "T+":
                st.error(
                    "Required area exceeds the largest API 526 standard orifice (T = 16 774 mm²). "
                    "Consider multiple PSVs in parallel or a non-standard design."
                )

            # Back-pressure subcritical warning for gas/steam
            if _psv_service in ("Gas / Vapour", "Steam") and _psv_gamma:
                _r_c = (2.0 / (_psv_gamma + 1.0)) ** (_psv_gamma / (_psv_gamma - 1.0))
                if _psv_Pback_kPa / _psv_P1_kPa > _r_c:
                    st.warning(
                        f"Back pressure ratio {_psv_Pback_kPa/_psv_P1_kPa:.3f} > critical ratio "
                        f"{_r_c:.4f} — flow may be subcritical at the nozzle. "
                        "Contact valve vendor for subcritical-flow rated capacity."
                    )

            # ── Results ─────────────────────────────────────────────────────────
            st.divider()
            st.subheader("Sizing Results")

            _rc1, _rc2, _rc3, _rc4 = st.columns(4)
            _rc1.metric("Required area", f"{_psv_res['A_req_mm2']:.1f} mm²")
            _rc2.metric("API 526 orifice", _psv_res["orifice_letter"])
            _rc3.metric("Orifice area", f"{_psv_res['orifice_area_mm2']:.1f} mm²"
                        if _psv_res["orifice_area_mm2"] < math.inf else "N/A")
            _area_margin = (
                (_psv_res["orifice_area_mm2"] / _psv_res["A_req_mm2"] - 1.0) * 100.0
                if _psv_res["orifice_area_mm2"] < math.inf and _psv_res["A_req_mm2"] > 0
                else 0.0
            )
            _rc4.metric("Area margin", f"+{_area_margin:.1f}%")

            # ── Flange row ───────────────────────────────────────────────────────
            _flange = psv.flange_nps(_psv_res["orifice_letter"])
            if _flange:
                _in_nps, _out_nps = _flange
                _in_dn,  _out_dn  = psv.nps_to_dn(_in_nps), psv.nps_to_dn(_out_nps)
                _fl_class = psv.min_flange_class(_psv_P1_bara)
                _rfl1, _rfl2, _rfl3, _rfl4 = st.columns(4)
                _rfl1.metric("Inlet flange",
                             f"NPS {_in_nps:g}\" / DN {_in_dn}")
                _rfl2.metric("Outlet flange",
                             f"NPS {_out_nps:g}\" / DN {_out_dn}")
                _rfl3.metric("Min. flange class",
                             f"ASME {_fl_class} lb",
                             help="ASME B16.5 Group 1.1 CS at 38 °C. Class 300 is the "
                                  "industry minimum for PSV inlets per API 526.")
                _rfl4.metric("Orifice designation",
                             f"{_in_nps:g}\" × {_out_nps:g}\" – {_psv_res['orifice_letter']}")

            # Rated capacity row
            if _psv_service in ("Gas / Vapour", "Steam") and "capacity_selected_kgh" in _psv_res:
                _rcap1, _rcap2, _rcap3, _rcap4 = st.columns(4)
                _rcap1.metric("Rated capacity", f"{_psv_res['capacity_selected_kgh']:.0f} kg/h")
                _rcap2.metric("Required flow",  f"{_psv_W_kgh:.0f} kg/h")
                _rcap3.metric("C coefficient",  f"{_psv_res['C_coeff']:.5f}")
                _rcap4.metric("Kb",             f"{_psv_res['Kb']:.3f}")
            elif _psv_service == "Liquid" and "capacity_selected_m3h" in _psv_res:
                _rcap1, _rcap2, _rcap3, _rcap4 = st.columns(4)
                _rcap1.metric("Rated capacity", f"{_psv_res['capacity_selected_m3h']:.2f} m³/h")
                _rcap2.metric("Required flow",  f"{_psv_Q_m3h:.2f} m³/h")
                _rcap3.metric("Kv (viscosity)", f"{_psv_res['Kv']:.4f}")
                _rcap4.metric("Re_v",           f"{_psv_res['Re_v']:.0f}")

            # Correction factors table
            st.markdown("**Correction factors applied**")
            _kfactors = {
                "Factor": ["Kd (discharge)", "Kb (back pressure)", "Kc (comb. disc)"],
                "Value":  [
                    f"{_psv_res['Kd']:.3f}",
                    f"{_psv_res.get('Kb', _psv_res.get('Kw', 1.0)):.3f}",
                    f"{_psv_res['Kc']:.3f}",
                ],
                "Source": [
                    "API 520 Table 5",
                    "API 520 §3.3 / PSV type",
                    "API 520 §4.7",
                ],
            }
            if _psv_service == "Liquid":
                _kfactors["Factor"].append("Kw (back pressure)")
                _kfactors["Value"].append(f"{_psv_res['Kw']:.3f}")
                _kfactors["Source"].append("API 520 §3.5")
            st.dataframe(pd.DataFrame(_kfactors), width='stretch', hide_index=True)

            # Standard orifice table with flange sizes
            st.markdown("**API 526 standard orifice and flange sizes**")
            _orifice_rows = []
            for _letter, _area in psv.API526_ORIFICES.items():
                _fn = psv.flange_nps(_letter)
                _in_str  = f"NPS {_fn[0]:g}\" (DN {psv.nps_to_dn(_fn[0])})" if _fn else "—"
                _out_str = f"NPS {_fn[1]:g}\" (DN {psv.nps_to_dn(_fn[1])})" if _fn else "—"
                _orifice_rows.append({
                    "Letter": _letter,
                    "Eff. area (mm²)": f"{_area:.1f}",
                    "Inlet flange": _in_str,
                    "Outlet flange": _out_str,
                    "": "✓" if _letter == _psv_res["orifice_letter"] else "",
                })
            st.dataframe(pd.DataFrame(_orifice_rows), width='stretch', hide_index=True)

            with st.expander("Assumptions and references"):
                st.markdown(f"""
    **API 520 Part I** (SI, 10th Ed.) — PSV sizing for gas/vapour, steam, and liquid service.

    **API 526** (7th Ed.) — standard effective orifice areas (D through T). Effective area is
    derated from the curtain area; actual installed orifice will be larger.

    **Gas / vapour formula** — API 520 Eq. (3):
    A = W / (C · Kd · P₁ · Kb · Kc) · √(T·Z/M)
    with C = 0.03948·√(γ·(2/(γ+1))^((γ+1)/(γ−1))) = **{psv.c_gas(_psv_gamma if _psv_gamma else 1.4):.5f}** for the selected gas.

    **Liquid formula** — derived from Bernoulli (API 520 Eq. (12)):
    A = (Q/3600) / (Kd · Kw · Kc · Kv · √(2·ρ·ΔP))

    **Viscosity correction Kv** — API 520 Figure 13:
    Kv = 1/(0.9935 + 2.878/R^0.5 + 342.75/R^1.5), R = 17 900·W/(μ·√A).
    Kv ≈ 1.0 for R ≥ 10 000 (typical light-to-medium liquids).

    **Backpressure Kb (conventional)** — Kb = 1.0 (nozzle remains choked).
    Limit: back pressure ≤ 10% of set pressure before contacting vendor.

    **Backpressure Kb (balanced bellows)** — API 520 Figure 31 conservative curve-fit.
    Kb = 1.0 when back pressure ≤ critical ratio; linear decay above.

    **Overpressure allowance** — 10% for non-fire; 21% for fire case (API 520 §3.1 / ASME Sec VIII-1).
    """)

        # ── Word export ───────────────────────────────────────────────────────
        st.divider()
        _psv_flow_str = (f"{_psv_W_kgh:.1f} kg/h" if _psv_service != "Liquid"
                         else f"{_psv_Q_m3h:.3f} m³/h")
        _psv_inp = [
            ("Service",              _psv_service),
            ("PSV type",             _psv_type),
            ("Set pressure",         f"{_psv_Pset_barg:.2f} barg"),
            ("Allowable overpressure", f"{_psv_op_pct:.1f} %"),
            ("Relieving pressure P₁", f"{_psv_P1_bara:.3f} bara"),
            ("Back pressure",        f"{_psv_Pback_barg:.2f} barg"),
            ("Relieving temperature", f"{_psv_T_C:.1f} °C"),
            ("Relief flow",          _psv_flow_str),
            ("Rupture disc upstream", "Yes (Kc = 0.90)" if _psv_rupture_disc else "No (Kc = 1.00)"),
        ]
        if _psv_service == "Gas / Vapour":
            _psv_inp += [
                ("Species",   _psv_species),
                ("MW",        f"{_psv_MW:.3f} kg/kmol"),
                ("γ (Cp/Cv)", f"{_psv_gamma:.4f}"),
                ("Z",         f"{_psv_Z:.4f}"),
            ]
        elif _psv_service == "Liquid":
            _psv_inp += [
                ("Density ρ",   f"{_psv_rho:.2f} kg/m³"),
                ("Viscosity μ", f"{_psv_mu_cP:.4f} cP"),
            ]
        _psv_flange = psv.flange_nps(_psv_res["orifice_letter"])
        _psv_flange_str = (
            f"NPS {_psv_flange[0]:g}\" × NPS {_psv_flange[1]:g}\""
            if _psv_flange else "—"
        )
        _psv_res_rows = [
            ("Required area A_req",    f"{_psv_res['A_req_mm2']:.2f} mm²"),
            ("API 526 orifice letter",  _psv_res["orifice_letter"]),
            ("Orifice effective area",  f"{_psv_res['orifice_area_mm2']:.1f} mm²"
                                        if _psv_res["orifice_area_mm2"] < math.inf else "Exceeds T"),
            ("Area margin",            f"+{_area_margin:.1f} %"),
            ("Flange (inlet × outlet)", _psv_flange_str),
            ("Min. flange class (inlet)", f"ASME {psv.min_flange_class(_psv_P1_bara)} lb"),
            ("Kd (discharge)",         f"{_psv_res['Kd']:.3f}"),
            ("Kb / Kw (back pressure)", f"{_psv_res.get('Kb', _psv_res.get('Kw', 1.0)):.3f}"),
            ("Kc (rupture disc)",       f"{_psv_res['Kc']:.3f}"),
        ]
        if _psv_service != "Liquid" and "capacity_selected_kgh" in _psv_res:
            _psv_res_rows += [
                ("Rated capacity (selected orifice)", f"{_psv_res['capacity_selected_kgh']:.0f} kg/h"),
                ("C coefficient",                     f"{_psv_res['C_coeff']:.5f}"),
            ]
        elif _psv_service == "Liquid" and "capacity_selected_m3h" in _psv_res:
            _psv_res_rows += [
                ("Rated capacity (selected orifice)", f"{_psv_res['capacity_selected_m3h']:.3f} m³/h"),
                ("Kv (viscosity correction)",         f"{_psv_res['Kv']:.5f}"),
            ]
        _psv_orifice_data = []
        for _let, _ar in psv.API526_ORIFICES.items():
            _fn = psv.flange_nps(_let)
            _psv_orifice_data.append([
                _let,
                f"{_ar:.1f}",
                f"NPS {_fn[0]:g}\"" if _fn else "—",
                f"NPS {_fn[1]:g}\"" if _fn else "—",
                "✓" if _let == _psv_res["orifice_letter"] else "",
            ])
        _psv_rpt = report_generator.generate_calculator_report(
            tool_name="Pressure Safety Valve",
            subtitle=f"{_psv_service}  ·  {_psv_type}  ·  {_psv_res['orifice_letter']} orifice",
            method_text=[
                "API 520 Part I (SI, 10th Ed.) orifice area sizing. "
                "The required effective orifice area is calculated from the specified relief "
                "flow, relieving conditions, and correction factors (Kd, Kb, Kc). "
                "The smallest API 526 standard orifice that meets or exceeds the required area "
                "is selected. The standard effective orifice areas are from API 526 Table 2.",
                f"Gas/steam: A = W/(C·Kd·P₁·Kb·Kc) · √(T·Z/M).  "
                f"Liquid: A = (Q/3600)/(Kd·Kw·Kc·Kv·√(2·ρ·ΔP)).",
            ],
            inputs_rows=_psv_inp,
            results_rows=_psv_res_rows,
            extra_tables=[{
                "title": "API 526 Standard Orifice Table",
                "headers": ["Letter", "Eff. area (mm²)", "Inlet flange", "Outlet flange", "Selected"],
                "data": _psv_orifice_data,
                "col_widths": [0.55, 1.0, 1.3, 1.3, 0.65],
            }],
        )
        st.download_button(
            "Export Word (.docx)",
            _psv_rpt,
            file_name=f"psv_{_psv_service.replace(' / ','_').lower()}_{_psv_res['orifice_letter']}.docx",
            mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            key="psv_dl",
        )

    # =========================================================================
    # Tab: Control Valve — IEC 60534-2-1
    # =========================================================================
    with tab_cv:
        st.markdown(
            "Control valve sizing per **IEC 60534-2-1** (≡ ISA 75.01.01). "
            "Calculates the required flow coefficient Kv / Cv and recommends a valve body size."
        )

        _cv_col_in, _cv_col_out = st.columns([1, 2], gap="large")

        with _cv_col_in:
            st.markdown("**SERVICE**")
            _cv_service = st.selectbox(
                "Service type", ["Liquid", "Gas / Vapour", "Steam"], key="cv_service"
            )

            st.markdown("**VALVE TYPE**")
            _cv_vtype = st.selectbox(
                "Valve type", list(cv.VALVE_TYPES.keys()), key="cv_vtype"
            )
            _vt = cv.VALVE_TYPES[_cv_vtype]
            _vt_col1, _vt_col2 = st.columns(2)
            with _vt_col1:
                _cv_FL = st.number_input(
                    "FL (pressure recovery)",
                    value=float(_vt["FL"]) if _vt["FL"] else 0.90,
                    min_value=0.1, max_value=1.0, step=0.01, format="%.2f",
                    key="cv_FL",
                    help="Liquid pressure recovery factor — vendor datasheet preferred.",
                )
            with _vt_col2:
                _cv_xT = st.number_input(
                    "xT (choked ΔP ratio)",
                    value=float(_vt["xT"]) if _vt["xT"] else 0.72,
                    min_value=0.1, max_value=1.0, step=0.01, format="%.2f",
                    key="cv_xT",
                    help="Gas pressure drop ratio at choked flow — vendor datasheet preferred.",
                )

            st.markdown("**CONDITIONS**")
            _cv_P1 = st.number_input(
                "Upstream pressure P₁ (bara)", value=10.0, min_value=0.1,
                step=0.5, key="cv_P1"
            )
            _cv_P2 = st.number_input(
                "Downstream pressure P₂ (bara)", value=8.0, min_value=0.0,
                max_value=_cv_P1, step=0.5, key="cv_P2"
            )
            _cv_T_C = st.number_input(
                "Temperature (°C)", value=20.0, step=5.0, key="cv_T_C"
            )
            _cv_T1_K = _cv_T_C + 273.15
            _cv_dP   = _cv_P1 - _cv_P2
            st.caption(f"ΔP = {_cv_dP:.3f} bar  |  ΔP/P₁ = {_cv_dP/_cv_P1:.3f}")

            st.markdown("**PIPING CORRECTION**")
            _cv_FP = st.number_input(
                "FP (piping geometry factor)", value=1.0,
                min_value=0.5, max_value=1.0, step=0.01, format="%.3f",
                key="cv_FP",
                help="1.0 for no reducers. Compute per IEC 60534-2-3 when pipe ≠ valve size.",
            )

        # ── Fluid and flow rate (right column) ───────────────────────────────
        with _cv_col_out:
            if _cv_service == "Liquid":
                st.markdown("**LIQUID PROPERTIES**")
                _cv_liq_opts = list(engine.LIQUID_COOLPROP_ID.keys()) + ["KOH solution", "Custom"]
                _cv_liq = st.selectbox("Fluid", _cv_liq_opts, key="cv_liq")

                _cv_T_liq = _cv_T_C   # shared temperature

                if _cv_liq == "KOH solution":
                    _cv_koh_conc = st.slider(
                        "Concentration (wt%)", 5, 40, 30, 5, key="cv_koh_conc"
                    )
                    _lp = ro.koh_liquid_properties(_cv_T_liq, float(_cv_koh_conc))
                    _cv_rho = _lp["rho_kgm3"]
                    _cv_mu  = _lp["mu_pas"]
                    _cv_Pv  = _lp["Pv_Pa"]
                    _cv_Pc  = 0.0  # no critical pressure for mixture
                    st.caption(f"ρ = {_cv_rho:.1f} kg/m³  |  μ = {_cv_mu*1e3:.3f} cP  |  Pᵥ = {_cv_Pv/1e5:.5f} bara")
                elif _cv_liq != "Custom":
                    _cv_cp_id = engine.LIQUID_COOLPROP_ID[_cv_liq]
                    try:
                        _lp = ro.liquid_properties(_cv_cp_id, _cv_T_liq, _cv_P1 * 1e5)
                        _cv_rho = _lp["rho_kgm3"]
                        _cv_mu  = _lp["mu_pas"]
                        _cv_Pv  = _lp["Pv_Pa"]
                        _cv_Pc  = cv.critical_pressure(_cv_cp_id)
                        st.caption(
                            f"ρ = {_cv_rho:.1f} kg/m³  |  μ = {_cv_mu*1e3:.4f} cP  |  "
                            f"Pᵥ = {_cv_Pv/1e5:.5f} bara  |  Pc = {_cv_Pc/1e5:.1f} bara"
                        )
                    except Exception as _e:
                        st.error(f"CoolProp: {_e}")
                        _cv_rho, _cv_mu, _cv_Pv, _cv_Pc = 1000.0, 1e-3, 0.0, 0.0
                else:
                    _cvc1, _cvc2, _cvc3 = st.columns(3)
                    _cv_rho = _cvc1.number_input("ρ (kg/m³)",  value=1000.0, min_value=1.0,  step=10.0,  key="cv_rho")
                    _cv_mu  = _cvc2.number_input("μ (cP)",     value=1.0,    min_value=0.001, step=0.1,  key="cv_mu") * 1e-3
                    _cv_Pv  = _cvc3.number_input("Pᵥ (bara)",  value=0.023,  min_value=0.0,  step=0.001, key="cv_Pv",  format="%.4f") * 1e5
                    _cv_Pc  = st.number_input("Pc (bara, crit.)", value=220.6, min_value=0.0, step=1.0, key="cv_Pc") * 1e5

                st.markdown("**FLOW RATE**")
                _cv_liq_flow_mode = st.radio(
                    "Input as", ["Volumetric (m³/h)", "Mass (kg/h)"],
                    horizontal=True, key="cv_liq_flow_mode"
                )
                if _cv_liq_flow_mode == "Volumetric (m³/h)":
                    _cv_Q_m3h = st.number_input(
                        "Flow rate (m³/h)", value=10.0, min_value=0.0001, step=1.0, key="cv_Q_liq"
                    )
                else:
                    _cv_W_kgh_liq = st.number_input(
                        "Mass flow (kg/h)", value=10_000.0, min_value=0.001, step=100.0, key="cv_W_liq"
                    )
                    _cv_Q_m3h = _cv_W_kgh_liq / _cv_rho if _cv_rho > 0 else 0.0

                _cv_res = cv.cv_liquid_size(
                    _cv_Q_m3h, _cv_P1 * 1e5, _cv_P2 * 1e5,
                    _cv_rho, _cv_Pv, _cv_Pc, _cv_FL, _cv_FP,
                )

            elif _cv_service == "Gas / Vapour":
                st.markdown("**GAS PROPERTIES**")
                _gp1, _gp2 = st.columns(2)
                with _gp1:
                    _cv_species = st.selectbox(
                        "Species", list(fanno.GASES.keys()), key="cv_gas_species"
                    )
                _cv_MW_def  = fanno.GASES[_cv_species][0]
                _cv_gam_def = fanno.GASES[_cv_species][1]
                with _gp2:
                    _cv_Z = st.number_input(
                        "Compressibility Z", value=1.0, min_value=0.1, max_value=2.0,
                        step=0.01, key="cv_Z"
                    )
                _gp3, _gp4 = st.columns(2)
                with _gp3:
                    _cv_MW = st.number_input(
                        "MW (kg/kmol)",
                        value=float(_cv_MW_def) * 1000.0 if _cv_MW_def else 28.97,
                        min_value=1.0, step=0.5, key="cv_MW"
                    )
                with _gp4:
                    _cv_gamma = st.number_input(
                        "γ (Cp/Cv)",
                        value=float(_cv_gam_def) if _cv_gam_def else 1.40,
                        min_value=1.01, max_value=2.0, step=0.01, key="cv_gamma"
                    )

                st.markdown("**FLOW RATE**")
                _cv_gas_flow_mode = st.radio(
                    "Input as", ["Mass flow (kg/h)", "Normal volumetric (Nm³/h)"],
                    horizontal=True, key="cv_gas_flow_mode"
                )
                if _cv_gas_flow_mode == "Mass flow (kg/h)":
                    _cv_W_kgh = st.number_input(
                        "Mass flow (kg/h)", value=1000.0, min_value=0.001, step=100.0, key="cv_W_gas"
                    )
                else:
                    _cv_Qn = st.number_input(
                        "Normal flow (Nm³/h)", value=5000.0, min_value=0.001, step=100.0, key="cv_Qn"
                    )
                    _cv_W_kgh = cv.nm3h_to_kgh(_cv_Qn, _cv_MW, _cv_Z)
                    st.caption(f"≡ {_cv_W_kgh:.1f} kg/h at MW = {_cv_MW:.2f} kg/kmol")

                _cv_res = cv.cv_gas_size(
                    _cv_W_kgh, _cv_P1 * 1e5, _cv_P2 * 1e5, _cv_T1_K,
                    _cv_MW, _cv_gamma, _cv_Z, _cv_xT, _cv_FP,
                )

            else:  # Steam
                st.markdown("**STEAM PROPERTIES** — from CoolProp at relieving conditions")

                st.markdown("**FLOW RATE**")
                _cv_W_steam = st.number_input(
                    "Mass flow (kg/h)", value=1000.0, min_value=0.001, step=100.0, key="cv_W_steam"
                )

                _cv_res = cv.cv_steam_size(
                    _cv_W_steam, _cv_P1 * 1e5, _cv_P2 * 1e5, _cv_T1_K, _cv_xT, _cv_FP,
                )
                for _note in _cv_res.get("notes", []):
                    if _note.startswith("UNPHYSICAL"):
                        st.error(_note)
                    else:
                        st.caption(_note)

            # ── Results ───────────────────────────────────────────────────────
            st.divider()
            st.subheader("Sizing Results")

            _kv  = _cv_res["Kv_req"]
            _cv_ = _cv_res["Cv_req"]

            _r1, _r2, _r3, _r4 = st.columns(4)
            _r1.metric("Kv required", f"{_kv:.2f} m³/h/√bar")
            _r2.metric("Cv required", f"{_cv_:.2f} USgpm/√psi")
            _r3.metric("ΔP actual",   f"{_cv_res['dP_bar']:.3f} bar")

            if _cv_service == "Liquid":
                _dp_chok = _cv_res["dP_choked_bar"]
                _r4.metric(
                    "ΔP choked",
                    f"{_dp_chok:.3f} bar" if _dp_chok < math.inf else "∞ (no Pᵥ data)"
                )
            else:
                _r4.metric("x / xT", f"{_cv_res['x']:.3f} / {_cv_res['x_choked']:.3f}")

            # Flow status row
            _s1, _s2, _s3, _s4 = st.columns(4)
            if _cv_service == "Liquid":
                _s1.metric("Ff (crit. press. ratio)", f"{_cv_res['Ff']:.4f}")
                _s2.metric("FL", f"{_cv_res['FL']:.3f}")
                _sig = _cv_res["sigma"]
                _s3.metric("σ (cavitation index)", f"{_sig:.3f}" if _sig < 1e6 else "∞")
                _s4.metric("FP (piping)", f"{_cv_res['FP']:.3f}")
            else:
                _s1.metric("Y (expansion factor)", f"{_cv_res['Y']:.4f}")
                _s2.metric("Fγ", f"{_cv_res['Fgamma']:.3f}")
                _s3.metric("ρ₁", f"{_cv_res['rho1_kgm3']:.3f} kg/m³")
                _s4.metric("FP (piping)", f"{_cv_res['FP']:.3f}")

            # Warnings
            if _cv_service == "Liquid":
                if _cv_res["choked"]:
                    st.warning(
                        f"Choked liquid flow — ΔP ({_cv_res['dP_bar']:.3f} bar) exceeds "
                        f"ΔP_choked ({_cv_res['dP_choked_bar']:.3f} bar). "
                        "Kv is sized to the choked ΔP; actual flow will not increase with further ΔP increase."
                    )
                if _cv_res["cavitating"]:
                    st.warning(
                        f"Cavitation risk — σ = {_cv_res['sigma']:.3f} < 1/FL² = "
                        f"{1/_cv_FL**2:.3f}. Consider a higher-FL valve, multi-stage letdown, "
                        "or an anti-cavitation trim."
                    )
            else:
                if _cv_res["choked"]:
                    st.warning(
                        f"Choked gas flow — x ({_cv_res['x']:.3f}) ≥ Fγ·xT ({_cv_res['x_choked']:.3f}). "
                        "Flow is at the sonic limit; further lowering P₂ does not increase flow."
                    )

            # Valve body recommendation
            st.markdown("**Valve body size recommendation**")
            _sz = cv.suggest_valve_size(_kv, target_opening=0.80)
            _b1, _b2, _b3 = st.columns(3)
            _b1.metric("Minimum Kv_100 needed", f"{_sz['Kv_min']:.2f}")
            _b2.metric("Suggested body (indicative)", _sz["size_label"])
            _b3.metric("Opening at design flow", f"{_sz['opening_pct']:.0f}%")

            if _sz["oversized"]:
                st.error("Required Kv exceeds the largest standard body in the table. Use parallel valves or a custom design.")
            elif _sz["opening_pct"] < 20.0:
                st.warning(
                    f"Valve is only {_sz['opening_pct']:.0f}% open at design flow — "
                    "consider a smaller body for better controllability."
                )
            elif _sz["opening_pct"] > 90.0:
                st.warning(
                    f"Valve is {_sz['opening_pct']:.0f}% open — tight controllability margin. "
                    "Verify with vendor Kv_100 before ordering."
                )

            # Indicative body size table
            st.markdown("**Indicative Kv_100 by body size** (globe valve — verify with vendor)")
            _kv_rows = []
            for _lbl, _k100 in cv.INDICATIVE_KV.items():
                _kv_rows.append({
                    "Body size":     _lbl,
                    "Kv_100":        f"{_k100:.1f}",
                    "Cv_100":        f"{_k100 * cv.KV_TO_CV:.1f}",
                    "At design flow": f"{_kv/_k100*100:.0f}%" if _k100 >= _kv * 0.2 else "—",
                    "":              "✓" if _lbl == _sz["size_label"] else "",
                })
            st.dataframe(pd.DataFrame(_kv_rows), width='stretch', hide_index=True)

            with st.expander("Equations and references"):
                st.markdown(f"""
**Standard**: IEC 60534-2-1:2011 (≡ ISA 75.01.01) — Inherent flow characteristics and rangeability.

**Liquid Kv** (no choke):
Kv = Q [m³/h] / √(ΔP [bar] / G_f)  where G_f = ρ / 1000 (specific gravity)

**Choked liquid ΔP** (IEC 60534-2-1 §5.2):
ΔP_choked = FL² · (P₁ − Ff · Pv)
Ff = 0.96 − 0.28 · √(Pv / Pc) = **{_cv_res.get('Ff', 0.96):.4f}**
FL = **{_cv_FL:.2f}**  (liquid pressure recovery factor — valve specific)

**Gas / vapour Kv** (IEC 60534-2-1 §5.3):
W [kg/h] = 31.6 · FP · Kv · Y · √(x · P₁ [bar] · ρ₁ [kg/m³])
Y = 1 − x / (3 · Fγ · xT) = **{_cv_res.get('Y', 1.0):.4f}** (expansion factor; 0.667 at choke)
Fγ = γ / 1.4,  xT = **{_cv_xT:.2f}**  (pressure drop ratio at choke — valve specific)
Choked when x = ΔP/P₁ ≥ Fγ · xT

**Kv → Cv**: Cv = Kv / 0.865  (Kv: m³/h at 1 bar ΔP → Cv: USgpm at 1 psi ΔP)

**Valve body table**: indicative globe-trim Kv_100; ball and butterfly valves are
typically 2–5× higher for the same body size. Always confirm with vendor.

**Cavitation**: σ = (P₁ − Pv) / ΔP. Incipient cavitation when σ < 1/FL².
Anti-cavitation trims can operate at σ down to ~0.5 FL².
""")

        # ── Word export ───────────────────────────────────────────────────────
        st.divider()
        _cv_inp = [
            ("Service",        _cv_service),
            ("Valve type",     _cv_vtype),
            ("FL",             f"{_cv_FL:.3f}"),
            ("xT",             f"{_cv_xT:.3f}"),
            ("FP (piping)",    f"{_cv_FP:.3f}"),
            ("P₁ upstream",    f"{_cv_P1:.3f} bara"),
            ("P₂ downstream",  f"{_cv_P2:.3f} bara"),
            ("ΔP",             f"{_cv_dP:.4f} bar"),
            ("Temperature",    f"{_cv_T_C:.1f} °C"),
        ]
        if _cv_service == "Liquid":
            _cv_lbl = getattr(_cv_liq, "__str__", lambda: str(_cv_liq))()
            _cv_inp += [
                ("Fluid",        _cv_lbl),
                ("Density ρ",    f"{_cv_rho:.2f} kg/m³"),
                ("Viscosity μ",  f"{_cv_mu*1e3:.4f} mPa·s"),
                ("Flow rate",    f"{_cv_Q_m3h:.4f} m³/h"),
            ]
        elif _cv_service == "Gas / Vapour":
            _cv_inp += [
                ("Species",    _cv_species),
                ("MW",         f"{_cv_MW:.3f} kg/kmol"),
                ("γ (Cp/Cv)",  f"{_cv_gamma:.4f}"),
                ("Z",          f"{_cv_Z:.4f}"),
                ("Flow rate",  f"{_cv_W_kgh:.2f} kg/h"),
            ]
        else:
            _cv_inp.append(("Flow rate", f"{_cv_W_steam:.2f} kg/h"))

        _cv_res_rows = [
            ("Kv required",  f"{_cv_res['Kv_req']:.3f} m³/h/√bar"),
            ("Cv required",  f"{_cv_res['Cv_req']:.3f} USgpm/√psi"),
            ("ΔP actual",    f"{_cv_res['dP_bar']:.4f} bar"),
            ("FP (piping)",  f"{_cv_res['FP']:.4f}"),
            ("Choked?",      "Yes" if _cv_res['choked'] else "No"),
        ]
        if _cv_service == "Liquid":
            _cv_res_rows += [
                ("ΔP choked",    f"{_cv_res['dP_choked_bar']:.4f} bar"
                                 if _cv_res['dP_choked_bar'] < math.inf else "∞"),
                ("Ff",           f"{_cv_res['Ff']:.5f}"),
                ("FL",           f"{_cv_res['FL']:.4f}"),
                ("σ (cavitation)", f"{_cv_res['sigma']:.4f}"
                                   if _cv_res['sigma'] < 1e6 else "∞"),
                ("Cavitating?",  "Yes" if _cv_res['cavitating'] else "No"),
            ]
        else:
            _cv_res_rows += [
                ("x (ΔP/P₁)",        f"{_cv_res['x']:.5f}"),
                ("xT·Fγ (choke)",    f"{_cv_res['x_choked']:.5f}"),
                ("Y (expansion)",    f"{_cv_res.get('Y', 1.0):.5f}"),
                ("Fγ",               f"{_cv_res.get('Fgamma', 1.0):.5f}"),
                ("ρ₁ (kg/m³)",       f"{_cv_res.get('rho1_kgm3', 0.0):.4f}"),
            ]
        _cv_res_rows += [
            ("Suggested body size",  _sz["size_label"]),
            ("Kv_100 min needed",    f"{_sz['Kv_min']:.2f}"),
            ("Opening at design",    f"{_sz['opening_pct']:.0f} %"),
        ]
        _cv_kv_data = [
            [_lbl, f"{_k100:.1f}", f"{_k100*cv.KV_TO_CV:.1f}",
             f"{_kv/_k100*100:.0f}%" if _k100 >= _kv * 0.2 else "—",
             "✓" if _lbl == _sz["size_label"] else ""]
            for _lbl, _k100 in cv.INDICATIVE_KV.items()
        ]
        _cv_rpt = report_generator.generate_calculator_report(
            tool_name="Control Valve",
            subtitle=f"{_cv_service}  ·  {_cv_vtype}  ·  Kv = {_cv_res['Kv_req']:.2f} m³/h/√bar",
            method_text=(
                "IEC 60534-2-1:2011 (≡ ISA 75.01.01) flow coefficient sizing. "
                "For liquid service, the choked ΔP limit is computed from FL² · (P₁ − Ff · Pv); "
                "the cavitation index σ = (P₁ − Pv)/ΔP is checked against 1/FL². "
                "For gas/steam service, the expansion factor Y = 1 − x/(3·Fγ·xT) accounts for "
                "gas expansion; choked flow occurs when x ≥ Fγ·xT. "
                "Body size recommendation is based on indicative globe-trim Kv_100 values at "
                "80 % opening target."
            ),
            inputs_rows=_cv_inp,
            results_rows=_cv_res_rows,
            extra_tables=[{
                "title": "Indicative Kv_100 by Body Size",
                "headers": ["Body size", "Kv_100", "Cv_100", "Opening", ""],
                "data": _cv_kv_data,
                "col_widths": [1.1, 0.8, 0.8, 0.9, 0.4],
            }],
        )
        st.download_button(
            "Export Word (.docx)",
            _cv_rpt,
            file_name=f"cv_{_cv_service.replace(' / ','_').lower()}.docx",
            mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            key="cv_dl",
        )

    # =========================================================================
    # Tab: Dissolved Gas Flash
    # =========================================================================
    with tab_dg:
        st.markdown("## Dissolved Gas Flash")
        st.markdown(
            "Calculates gas released (or absorbed) when a liquid **saturated** with a gas "
            "at upstream conditions flashes to lower pressure and/or temperature. "
            "Uses tabulated Henry's Law constants (Battino et al.) with a Sechenov correction "
            "for KOH salting-out."
        )

        _dg_c1, _dg_c2 = st.columns(2)

        with _dg_c1:
            st.markdown("**Gas & Solvent**")
            _dg_gas = st.selectbox(
                "Dissolved gas",
                list(dg.GAS_LABELS.keys()),
                format_func=lambda g: dg.GAS_LABELS[g],
                key="dg_gas",
            )
            _dg_solvent = st.selectbox(
                "Solvent",
                ["KOH solution", "Pure water"],
                key="dg_solvent",
            )
            if _dg_solvent == "KOH solution":
                _dg_koh_wt = st.slider(
                    "KOH concentration (wt%)", 5.0, 50.0, 30.0, 1.0, key="dg_koh_wt"
                )
            else:
                _dg_koh_wt = 0.0

        with _dg_c2:
            st.markdown("**Upstream conditions (saturated)**")
            _dg_T1 = st.number_input("T₁ (°C)", -10.0, 200.0, 80.0, 1.0, key="dg_T1")
            _dg_P1 = st.number_input("P₁ (bar a)", 0.5, 300.0, 30.0, 0.5, key="dg_P1")

            st.markdown("**Downstream conditions**")
            _dg_T2 = st.number_input("T₂ (°C)", -10.0, 200.0, 70.0, 1.0, key="dg_T2")
            _dg_P2 = st.number_input("P₂ (bar a)", 0.1, 300.0, 29.0, 0.5, key="dg_P2")

        st.divider()

        # ── Compute ──────────────────────────────────────────────────────────
        try:
            _dg_res = dg.flash_dissolved_gas(
                gas=_dg_gas,
                T1_C=_dg_T1,
                P1_bar=_dg_P1,
                T2_C=_dg_T2,
                P2_bar=_dg_P2,
                wt_pct_koh=_dg_koh_wt,
                y_gas=1.0,
            )

            _dg_released = _dg_res["dC_combined_mol_L"] > 0
            _dg_vol_pct  = _dg_res["vol_pct_outlet"]

            # ── KPIs ─────────────────────────────────────────────────────────
            st.markdown("**Flash result — combined (ΔP + ΔT)**")
            _dg_ka, _dg_kb, _dg_kc, _dg_kd, _dg_ke = st.columns(5)
            _dg_ka.metric(
                "Net Δ concentration",
                f"{_dg_res['dC_combined_mol_L']*1e3:+.3f} mmol/L",
                help="Positive = gas released; negative = more gas can dissolve (liquid stays undersaturated).",
            )
            _dg_kb.metric(
                "Released volume",
                f"{_dg_res['released_mL_per_L']:.2f} mL/L" if _dg_released else "0 (absorbed)",
                help="mL of gas at STP (0 °C, 1 atm) per litre of liquid.",
            )
            _dg_kc.metric(
                "Nm³/m³ liquid",
                f"{_dg_res['released_Nm3_per_m3']:.4f}" if _dg_released else "0",
            )
            _dg_kd.metric(
                "Mass released",
                f"{_dg_res['released_g_per_L']*1e3:.3f} mg/L" if _dg_released else "0",
            )
            _dg_pump_note = (
                "⚠ Pump risk (>4%)" if _dg_vol_pct > 4.0
                else ("△ Monitor (2–4%)" if _dg_vol_pct > 2.0 else "✓ OK (<2%)")
            ) if _dg_released else "—"
            _dg_ke.metric(
                "Vol% gas at outlet",
                f"{_dg_vol_pct:.2f} %" if _dg_released else "0 %",
                delta=_dg_pump_note,
                delta_color="inverse" if _dg_vol_pct > 4.0 else ("off" if _dg_vol_pct > 2.0 else "normal"),
                help="Actual gas volume fraction in the liquid+gas mixture at T₂, P₂ (ideal gas law). "
                     "Centrifugal pumps typically tolerate <2 vol%; above 4 vol% performance and stability degrade.",
            )

            if not _dg_released:
                st.info(
                    f"The outlet conditions can hold **more** {_dg_gas} than at inlet. "
                    "The liquid is **undersaturated** after the HX — no gas is released."
                )

            # ── Decomposition table ──────────────────────────────────────────
            st.markdown("**Effect decomposition**")
            _dg_dec = {
                "Effect": ["Pressure drop only (ΔP, same T₁)", "Cooling only (ΔT, same P₁)", "Combined (ΔP + ΔT)"],
                "ΔC (mmol/L)": [
                    f"{_dg_res['dC_pressure_mol_L']*1e3:+.3f}",
                    f"{_dg_res['dC_temp_mol_L']*1e3:+.3f}",
                    f"{_dg_res['dC_combined_mol_L']*1e3:+.3f}",
                ],
                "Released (mL/L STP)": [
                    f"{_dg_res['pressure_effect']['V_mL_per_L']:.3f}" if _dg_res['dC_pressure_mol_L'] > 0 else "0 (absorbed)",
                    f"{_dg_res['temp_effect']['V_mL_per_L']:.3f}"    if _dg_res['dC_temp_mol_L'] > 0    else "0 (absorbed)",
                    f"{_dg_res['combined_effect']['V_mL_per_L']:.3f}" if _dg_res['dC_combined_mol_L'] > 0 else "0 (absorbed)",
                ],
                "Nm³/m³": [
                    f"{_dg_res['pressure_effect']['V_Nm3_per_m3']:.4f}" if _dg_res['dC_pressure_mol_L'] > 0 else "—",
                    f"{_dg_res['temp_effect']['V_Nm3_per_m3']:.4f}"    if _dg_res['dC_temp_mol_L'] > 0    else "—",
                    f"{_dg_res['combined_effect']['V_Nm3_per_m3']:.4f}" if _dg_res['dC_combined_mol_L'] > 0 else "—",
                ],
            }
            st.dataframe(pd.DataFrame(_dg_dec), hide_index=True, width='stretch')

            # ── Henry constants table ────────────────────────────────────────
            st.markdown("**Henry's Law constants**")
            _dg_solvent_label = f"KOH {_dg_koh_wt:.0f} wt%" if _dg_koh_wt > 0 else "Pure water"
            _dg_htab = {
                "State": ["Upstream (T₁, P₁)", "Downstream (T₂, P₂)"],
                "T (°C)": [f"{_dg_T1:.1f}", f"{_dg_T2:.1f}"],
                "P gas (bar)": [
                    f"{_dg_P1:.3f}",
                    f"{_dg_P2:.3f}",
                ],
                "K_H water (mol/L/bar)": [
                    f"{_dg_res['K_H1_water']:.4e}",
                    f"{_dg_res['K_H2_water']:.4e}",
                ],
                f"K_H {_dg_solvent_label} (mol/L/bar)": [
                    f"{_dg_res['K_H1_soln']:.4e}",
                    f"{_dg_res['K_H2_soln']:.4e}",
                ],
                "K_s (L/mol)": [
                    f"{_dg_res['K_s1']:.4f}",
                    f"{_dg_res['K_s2']:.4f}",
                ],
                "C (mol/L)": [
                    f"{_dg_res['C1_mol_L']:.4e}",
                    f"{_dg_res['C2_mol_L']:.4e}",
                ],
            }
            st.dataframe(pd.DataFrame(_dg_htab), hide_index=True, width='stretch')

            if _dg_koh_wt > 0:
                _dg_c_koh = _dg_res["c_koh1_mol_L"]
                st.caption(
                    f"KOH {_dg_koh_wt:.0f} wt% → {_dg_c_koh:.2f} mol/L at T₁.  "
                    f"Sechenov factor at T₁: 10^(−{_dg_res['K_s1']:.3f} × {_dg_c_koh:.2f}) "
                    f"= {10**(-_dg_res['K_s1']*_dg_c_koh):.3f}  "
                    f"(solubility reduced to {10**(-_dg_res['K_s1']*_dg_c_koh)*100:.0f}% of pure-water value)"
                )

            # ── K_H vs T chart ────────────────────────────────────────────────
            st.markdown("**K_H vs Temperature (solubility curve)**")
            _dg_temps = [t for t in range(-5, 101, 5)]
            _dg_kh_w = [dg._interp_kh_water(t, _dg_gas) * 1e4 for t in _dg_temps]
            _dg_kh_s = [dg.kh_solution(t, _dg_gas, _dg_koh_wt) * 1e4 for t in _dg_temps] if _dg_koh_wt > 0 else None

            _dg_fig = go.Figure()
            _dg_fig.add_trace(go.Scatter(
                x=_dg_temps, y=_dg_kh_w, mode="lines",
                name="Pure water", line=dict(color="#3b82f6", width=2),
            ))
            if _dg_kh_s:
                _dg_fig.add_trace(go.Scatter(
                    x=_dg_temps, y=_dg_kh_s, mode="lines",
                    name=f"KOH {_dg_koh_wt:.0f} wt%", line=dict(color="#f59e0b", width=2, dash="dash"),
                ))
            # Mark operating points
            _dg_kh1_plot = dg.kh_solution(_dg_T1, _dg_gas, _dg_koh_wt) * 1e4
            _dg_kh2_plot = dg.kh_solution(_dg_T2, _dg_gas, _dg_koh_wt) * 1e4
            _dg_fig.add_trace(go.Scatter(
                x=[_dg_T1, _dg_T2], y=[_dg_kh1_plot, _dg_kh2_plot],
                mode="markers",
                marker=dict(size=10, color=["#16a34a", "#dc2626"], symbol="circle"),
                name="Operating points (T₁/T₂)",
            ))
            _dg_fig.update_layout(
                xaxis_title="Temperature (°C)",
                yaxis_title="K_H (×10⁻⁴ mol/L/bar)",
                height=320,
                margin=dict(t=20, b=40, l=60, r=20),
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
            )
            st.plotly_chart(_dg_fig, width='stretch')

            # ── Word export ───────────────────────────────────────────────────
            st.divider()
            _dg_solvent_lbl = (f"KOH {_dg_koh_wt:.0f} wt%" if _dg_koh_wt > 0
                               else "Pure water")
            _dg_rpt = report_generator.generate_calculator_report(
                tool_name="Dissolved Gas Flash",
                subtitle=f"{dg.GAS_LABELS[_dg_gas]}  in  {_dg_solvent_lbl}",
                method_text=(
                    "Henry's Law dissolution: C [mol/L] = K_H(T) × P_gas [bar]. "
                    "K_H is interpolated from Battino et al. tabular data for pure water. "
                    "For KOH solutions, solubility is reduced by the Sechenov equation: "
                    "log₁₀(K_H_water / K_H_KOH) = K_s × c_KOH. "
                    "Gas released on depressurisation is the difference in equilibrium "
                    "dissolved concentration between upstream and downstream conditions. "
                    "The flash is split into a pressure-only component (ΔP at constant T₁) "
                    "and a temperature-only component (ΔT at constant P₁); the combined "
                    "result is the net change between (T₁, P₁) and (T₂, P₂)."
                ),
                inputs_rows=[
                    ("Dissolved gas",         dg.GAS_LABELS[_dg_gas]),
                    ("Solvent",               _dg_solvent_lbl),
                    ("Upstream T₁",           f"{_dg_T1:.1f} °C"),
                    ("Upstream P₁",           f"{_dg_P1:.3f} bara"),
                    ("Downstream T₂",         f"{_dg_T2:.1f} °C"),
                    ("Downstream P₂",         f"{_dg_P2:.3f} bara"),
                ],
                results_rows=[
                    ("Upstream concentration C₁",  f"{_dg_res['C1_mol_L']:.4e} mol/L"),
                    ("Downstream concentration C₂", f"{_dg_res['C2_mol_L']:.4e} mol/L"),
                    ("Net ΔC (combined)",           f"{_dg_res['dC_combined_mol_L']*1e3:+.4f} mmol/L"),
                    ("Released volume",             f"{_dg_res['released_mL_per_L']:.4f} mL/L STP"
                                                    if _dg_released else "0 (absorbed)"),
                    ("Nm³/m³ liquid",               f"{_dg_res['released_Nm3_per_m3']:.6f}"
                                                    if _dg_released else "0"),
                    ("Mass released",               f"{_dg_res['released_g_per_L']*1e3:.4f} mg/L"
                                                    if _dg_released else "0"),
                    ("Vol% gas at outlet",          f"{_dg_vol_pct:.3f} %"),
                    ("ΔC — pressure effect only",   f"{_dg_res['dC_pressure_mol_L']*1e3:+.4f} mmol/L"),
                    ("ΔC — temperature effect only", f"{_dg_res['dC_temp_mol_L']*1e3:+.4f} mmol/L"),
                    ("K_H water at T₁",             f"{_dg_res['K_H1_water']:.4e} mol/L/bar"),
                    ("K_H water at T₂",             f"{_dg_res['K_H2_water']:.4e} mol/L/bar"),
                    (f"K_H {_dg_solvent_lbl} at T₁", f"{_dg_res['K_H1_soln']:.4e} mol/L/bar"),
                    (f"K_H {_dg_solvent_lbl} at T₂", f"{_dg_res['K_H2_soln']:.4e} mol/L/bar"),
                    ("Sechenov K_s at T₁",          f"{_dg_res['K_s1']:.4f} L/mol"),
                    ("Sechenov K_s at T₂",          f"{_dg_res['K_s2']:.4f} L/mol"),
                ],
                fig=_dg_fig,
                fig_caption_text="Henry's Law constant K_H vs temperature with operating points.",
                fig_height=340,
            )
            st.download_button(
                "Export Word (.docx)",
                _dg_rpt,
                file_name=f"dissolved_gas_{_dg_gas.lower()}_{_dg_solvent_lbl.replace(' ','_')}.docx",
                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                key="dg_dl",
            )

        except Exception as _dg_err:
            st.error(f"Calculation error: {_dg_err}")

        with st.expander("Theory & References"):
            st.markdown("""
**Henry's Law**

At equilibrium: C [mol/L] = K_H(T) × P_gas [bar]

K_H is strongly temperature-dependent and shows a **minimum near 45–55 °C for H₂** — above
this temperature, H₂ becomes *more* soluble with increasing temperature (unusual behaviour).
This means cooling a hot, pressurised H₂ stream can *increase* solubility and partially or
fully offset the effect of a pressure drop.

**Sechenov equation (KOH salting-out)**

log₁₀(K_H_water / K_H_KOH) = K_s × c_KOH

where K_s ≈ 0.069 L/mol for H₂ in KOH (Tremosa et al. 2019).
At 30 wt% KOH (~5.4 mol/L) and 25 °C, this gives a solubility reduction to ~45% of the pure-water value.

K_s has a mild temperature dependence approximated as K_s(T) ∝ (298.15/T)^0.3.

**Flash convention**
- Positive ΔC → gas released from solution
- Negative ΔC → liquid is *undersaturated* after the step; no gas is released

**References**

- Battino R., Rettich T.R., Tominaga T. (1984). *The solubility of nitrogen and air in liquids.* JPCRD **13**, 563.
  (H₂ and O₂ tables used for interpolation)
- Tremosa J. et al. (2019). *Geochemical characterization and modelling of hydrogen gas solubility in KOH solutions.* Applied Geochemistry.
  (K_s = 0.069 L/mol for H₂ in KOH)
- Sander R. (2015). *Compilation of Henry's law constants.* ACP **15**, 4399.
  (O₂ Sechenov constants)
""")

    # =========================================================================
    # Tab: Pump Hydraulics
    # =========================================================================
    with tab_pump:
        import pump_engine as pe
        import plotly.graph_objects as go
        import numpy as np

        st.markdown(
            "Centrifugal and positive-displacement pump sizing — hydraulic head, "
            "NPSH, power, and design pressure for downstream piping per API 610 / "
            "NORSOK P-001 / ASME B31.3."
        )

        _pu_c1, _pu_c2 = st.columns([1, 1.4], gap="large")

        # ── LEFT COLUMN — INPUTS ─────────────────────────────────────────────
        with _pu_c1:

            # ── Pump type ────────────────────────────────────────────────────
            st.markdown("**PUMP TYPE**")
            _pu_type = st.radio(
                "Pump type", ["Centrifugal", "Positive Displacement (PD)"],
                horizontal=True, key="pu_type", label_visibility="collapsed",
            )
            _pu_is_pd = (_pu_type == "Positive Displacement (PD)")

            st.divider()

            # ── Fluid ────────────────────────────────────────────────────────
            st.markdown("**FLUID**")
            _pu_fluid_type = st.radio(
                "Fluid type", ["KOH solution", "CoolProp liquid"],
                horizontal=True, key="pu_fluid_type", label_visibility="collapsed",
            )
            _pu_T_C = st.number_input(
                "Temperature (°C)", value=float(st.session_state.get("pu_T_C", 25.0)),
                min_value=-50.0, max_value=200.0, step=5.0, key="pu_T_C",
            )
            _pu_P_bara = st.number_input(
                "Suction pressure (bara)", value=float(st.session_state.get("pu_P_bara", 2.0)),
                min_value=0.1, max_value=500.0, step=0.5, key="pu_P_bara",
            )

            if _pu_fluid_type == "KOH solution":
                _pu_koh_conc = st.slider(
                    "KOH concentration (wt%)", min_value=20, max_value=40,
                    value=int(st.session_state.get("pu_koh_conc", 30)),
                    step=1, key="pu_koh_conc",
                    help="Typical alkaline electrolyser: 25–32 wt%.",
                )
                _pu_rho, _pu_mu, _pu_Pv = pe.koh_properties(_pu_T_C, _pu_koh_conc)
            else:
                _pu_cp_opts = list(engine.LIQUID_COOLPROP_ID.keys())
                _pu_fluid_name = st.selectbox(
                    "Fluid", _pu_cp_opts,
                    index=_pu_cp_opts.index(st.session_state.get("pu_fluid_name", "Water"))
                          if st.session_state.get("pu_fluid_name", "Water") in _pu_cp_opts else 0,
                    key="pu_fluid_name",
                )
                try:
                    _pu_rho, _pu_mu, _pu_Pv = pe.coolprop_liquid_properties(
                        _pu_fluid_name, _pu_T_C, _pu_P_bara)
                except Exception as _e:
                    st.error(f"CoolProp error: {_e}")
                    _pu_rho, _pu_mu, _pu_Pv = 1000.0, 1e-3, 0.023

            _fm1, _fm2, _fm3 = st.columns(3)
            _fm1.metric("ρ", f"{_pu_rho:.1f} kg/m³")
            _fm2.metric("μ", f"{_pu_mu*1e3:.3f} mPa·s")
            _fm3.metric("Pv", f"{_pu_Pv:.4f} bara",
                        help="Vapour pressure at pump inlet temperature. "
                             "KOH Pv accounts for ionic activity depression.")

            st.divider()

            # ── H-Q curve (centrifugal only) ──────────────────────────────
            if not _pu_is_pd:
                st.markdown("**PUMP CURVE  (H-Q)**")
                _pu_hq_mode = st.radio(
                    "Input mode", ["3-point parametric", "Tabular (up to 10 points)"],
                    horizontal=True, key="pu_hq_mode", label_visibility="collapsed",
                )

                if _pu_hq_mode == "3-point parametric":
                    _pu_H0  = st.number_input("Shut-off head H₀ (m)", value=float(st.session_state.get("pu_H0", 80.0)),
                                              min_value=1.0, max_value=5000.0, step=5.0, key="pu_H0")
                    _hc1, _hc2 = st.columns(2)
                    _pu_Hbep = _hc1.number_input("BEP head (m)", value=float(st.session_state.get("pu_Hbep", 60.0)),
                                                  min_value=0.1, max_value=4000.0, step=5.0, key="pu_Hbep")
                    _pu_Qbep = _hc2.number_input("BEP flow (m³/h)", value=float(st.session_state.get("pu_Qbep", 50.0)),
                                                  min_value=0.1, max_value=100000.0, step=5.0, key="pu_Qbep")
                    _pu_use_runout = st.checkbox("Add runout point", key="pu_use_runout")
                    if _pu_use_runout:
                        _ro1, _ro2 = st.columns(2)
                        _pu_Hro = _ro1.number_input("Runout head (m)", value=float(st.session_state.get("pu_Hro", 40.0)),
                                                     min_value=0.0, max_value=4000.0, step=5.0, key="pu_Hro")
                        _pu_Qro = _ro2.number_input("Runout flow (m³/h)", value=float(st.session_state.get("pu_Qro", 80.0)),
                                                     min_value=0.1, max_value=100000.0, step=5.0, key="pu_Qro")
                    else:
                        _pu_Hro = _pu_Qro = None

                    try:
                        _pu_hq_coeffs = pe.fit_hq_3point(
                            _pu_H0, _pu_Hbep, _pu_Qbep, _pu_Hro, _pu_Qro)
                    except ValueError as _e:
                        st.error(f"H-Q curve error: {_e}")
                        _pu_hq_coeffs = None

                else:  # Tabular
                    st.caption("Enter Q (m³/h) and H (m) pairs — minimum 3 rows including Q=0.")
                    _pu_tab_default = [
                        {"Q (m³/h)": 0.0,  "H (m)": 80.0},
                        {"Q (m³/h)": 25.0, "H (m)": 72.0},
                        {"Q (m³/h)": 50.0, "H (m)": 60.0},
                        {"Q (m³/h)": 75.0, "H (m)": 42.0},
                        {"Q (m³/h)": 90.0, "H (m)": 28.0},
                    ]
                    _pu_tab_data = st.data_editor(
                        _pu_tab_default, num_rows="dynamic", key="pu_hq_table",
                        column_config={
                            "Q (m³/h)": st.column_config.NumberColumn(min_value=0.0),
                            "H (m)":    st.column_config.NumberColumn(min_value=0.0),
                        },
                        hide_index=True,
                    )
                    _pu_Qs = [r["Q (m³/h)"] for r in _pu_tab_data if r["Q (m³/h)"] is not None]
                    _pu_Hs = [r["H (m)"]    for r in _pu_tab_data if r["H (m)"]    is not None]
                    if len(_pu_Qs) >= 3:
                        try:
                            _pu_hq_coeffs = pe.fit_hq_tabular(_pu_Qs, _pu_Hs)
                            _pu_H0  = pe.hq_shutoff(_pu_hq_coeffs)
                            _pu_Qbep = max(_pu_Qs) * 0.65
                        except Exception as _e:
                            st.error(f"Curve fit error: {_e}")
                            _pu_hq_coeffs = None
                    else:
                        st.warning("Enter at least 3 Q/H points.")
                        _pu_hq_coeffs = None

                # Speed / VSD
                st.markdown("**SPEED**")
                _sc1, _sc2 = st.columns(2)
                _pu_speed_rpm = _sc1.number_input(
                    "Rated speed (rpm)", value=int(st.session_state.get("pu_speed_rpm", 1450)),
                    min_value=100, max_value=10000, step=50, key="pu_speed_rpm",
                )
                _pu_vsd = _sc2.checkbox("VSD fitted", key="pu_vsd")
                if _pu_vsd:
                    _pu_n_max_pct = st.slider(
                        "Max VSD speed (% of rated)", min_value=50, max_value=120,
                        value=int(st.session_state.get("pu_n_max_pct", 105)),
                        step=1, key="pu_n_max_pct",
                    )
                    _pu_n_ratio = _pu_n_max_pct / 100.0
                else:
                    _pu_n_ratio = 1.0

                # Efficiency
                st.markdown("**EFFICIENCY**")
                _ec1, _ec2 = st.columns(2)
                _pu_eta_bep = _ec1.number_input(
                    "Pump η at BEP (%)", value=float(st.session_state.get("pu_eta_bep", 72.0)),
                    min_value=10.0, max_value=95.0, step=1.0, key="pu_eta_bep",
                )
                _pu_eta_motor = _ec2.number_input(
                    "Motor η (%)", value=float(st.session_state.get("pu_eta_motor", 93.0)),
                    min_value=50.0, max_value=99.0, step=0.5, key="pu_eta_motor",
                )
                if _pu_hq_coeffs:
                    _pu_Q_max = pe.hq_max_flow(_pu_hq_coeffs)
                    _pu_eta_params = pe.fit_eta_parabolic(
                        _pu_eta_bep,
                        _pu_Qbep if _pu_hq_mode == "3-point parametric" else _pu_Q_max * 0.65,
                        Q_runout=_pu_Q_max,
                    )

                st.divider()

            # ── System curve ─────────────────────────────────────────────────
            st.markdown("**SYSTEM CURVE**")
            _pu_H_static = st.number_input(
                "Static head H_static (m)",
                value=float(st.session_state.get("pu_H_static", 20.0)),
                min_value=-500.0, max_value=5000.0, step=1.0, key="pu_H_static",
                help=(
                    "Total head the pump must overcome at zero flow — elevation + pressure difference.  \n"
                    "H_static = Δz + (P_discharge − P_suction) × 1e5 / (ρ × g)  \n"
                    "Negative if pumping downhill or suction vessel is at higher pressure."
                ),
            )
            with st.expander("H_static calculator"):
                _hsc1, _hsc2 = st.columns(2)
                _pu_dz_calc   = _hsc1.number_input("Elevation Δz (m)", value=5.0,
                                                    min_value=-500.0, max_value=500.0, step=1.0,
                                                    key="pu_dz_calc",
                                                    help="z_discharge − z_suction.")
                _pu_dP_calc   = _hsc2.number_input("Pressure difference ΔP (bar)",
                                                    value=5.0, step=0.5, key="pu_dP_calc",
                                                    help="P_discharge − P_suction in bar. Negative if suction vessel is at higher pressure.")
                _pu_H_static_calc = _pu_dz_calc + _pu_dP_calc * 1e5 / (_pu_rho * pe.g)
                st.metric("Calculated H_static", f"{_pu_H_static_calc:.2f} m",
                          help="Copy this value into H_static above.")

            _sy1, _sy2 = st.columns(2)
            _pu_Q_ref = _sy1.number_input(
                "Reference flow (m³/h)",
                value=float(st.session_state.get("pu_Q_ref", 50.0)),
                min_value=0.0, step=5.0, key="pu_Q_ref",
                help="Flow at which the friction head below is estimated.",
            )
            _pu_H_fric_ref = _sy2.number_input(
                "Friction head at ref. flow (m)",
                value=float(st.session_state.get("pu_H_fric_ref", 8.0)),
                min_value=0.0, step=1.0, key="pu_H_fric_ref",
                help="Pipeline friction losses expressed as metres of head at the reference flow.  \n"
                     "H_friction = ΔP_friction [bar] × 1e5 / (ρ × g)",
            )
            _pu_k_fric = pe.k_from_reference(_pu_Q_ref, _pu_H_fric_ref) if _pu_Q_ref > 0 else 0.0

            st.divider()

            # ── NPSH ─────────────────────────────────────────────────────────
            if not _pu_is_pd:
                st.markdown("**NPSH**")
                _nc1, _nc2 = st.columns(2)
                _pu_z_suc = _nc1.number_input(
                    "Liquid level above pump CL (m)",
                    value=float(st.session_state.get("pu_z_suc", 2.0)),
                    min_value=-20.0, max_value=50.0, step=0.5, key="pu_z_suc",
                    help="Positive if liquid surface is above pump centreline (flooded suction). "
                         "Negative for suction lift.",
                )
                _pu_h_suc_loss = _nc2.number_input(
                    "Suction pipe loss (m head)",
                    value=float(st.session_state.get("pu_h_suc_loss", 1.0)),
                    min_value=0.0, max_value=50.0, step=0.5, key="pu_h_suc_loss",
                    help="Friction + fitting losses in suction line, expressed as metres of head.",
                )
                _pu_NPSH_R = st.number_input(
                    "NPSH required — vendor value (m)",
                    value=float(st.session_state.get("pu_NPSH_R", 3.0)),
                    min_value=0.0, max_value=50.0, step=0.5, key="pu_NPSH_R",
                    help="From pump datasheet at operating flow. "
                         "Use the value at BEP flow if operating point is unknown.",
                )
                st.divider()

            # ── Design pressure ───────────────────────────────────────────────
            st.markdown("**DESIGN PRESSURE — DOWNSTREAM PIPING**")

            if _pu_is_pd:
                _pu_PSV_set = st.number_input(
                    "PSV set pressure (barg)", value=float(st.session_state.get("pu_PSV_set", 12.0)),
                    min_value=0.1, max_value=500.0, step=0.5, key="pu_PSV_set",
                    help="Mandatory for PD pumps. Set at or below MAWP of weakest downstream element.",
                )
                _pu_pd_accum = st.selectbox(
                    "Accumulation scenario",
                    ["Standard process — 10 %", "Fire case — 21 %"],
                    key="pu_pd_accum_pd",
                )
                _pu_pd_acc_pct = 10.0 if "10" in _pu_pd_accum else 21.0
            else:
                _pu_dp_method = st.selectbox(
                    "Design pressure method",
                    options=list(pe.DESIGN_PRESSURE_METHODS.keys()),
                    format_func=lambda x: f"{x}. {pe.DESIGN_PRESSURE_METHODS[x]}",
                    key="pu_dp_method",
                )

                _pu_needs_PSV = _pu_dp_method in (4, 5, 6)

                if _pu_dp_method in (1, 2, 3):
                    if _pu_dp_method == 3:
                        st.caption("Enter the upstream PSV accumulation pressure as suction basis.")
                    _pu_P_suc_max = st.number_input(
                        "Max suction pressure (bara)" if _pu_dp_method < 3
                        else "Upstream PSV accumulation pressure (bara)",
                        value=float(st.session_state.get("pu_P_suc_max", 11.0)),
                        min_value=0.1, max_value=500.0, step=0.5, key="pu_P_suc_max",
                        help="Conservative basis: upstream vessel design pressure. "
                             "Method 3: upstream PSV set × 1.10.",
                    )
                else:
                    _pu_P_suc_max = _pu_P_bara  # not used for PSV methods

                if _pu_needs_PSV:
                    _pu_PSV_set = st.number_input(
                        "Discharge PSV set pressure (barg)",
                        value=float(st.session_state.get("pu_PSV_set", 12.0)),
                        min_value=0.1, max_value=500.0, step=0.5, key="pu_PSV_set",
                        help="PSV must be installed on the pump discharge side of the first isolation valve.",
                    )
                else:
                    _pu_PSV_set = None

                if _pu_dp_method == 6:
                    st.info(
                        "Method 6 requires a SIL-assessed SIS per IEC 61511. "
                        "The design pressure calculation is the same as Method 4 — "
                        "the SIL documentation provides the engineering justification "
                        "for using PSV accumulation as the sole design basis.",
                        icon="ℹ️",
                    )

            _pu_mat_group = st.selectbox(
                "Pipe material group (ASME B16.5)",
                pe.MATERIAL_GROUPS,
                index=pe.MATERIAL_GROUPS.index(
                    st.session_state.get("pu_mat_group", pe.MATERIAL_GROUPS[1])
                ) if st.session_state.get("pu_mat_group") in pe.MATERIAL_GROUPS else 1,
                key="pu_mat_group",
            )

        # ── RIGHT COLUMN — RESULTS ────────────────────────────────────────────
        with _pu_c2:

            # ── Pump computation (pure Python, no Streamlit) ──────────────────
            _pu_Qbep_used = (
                _pu_Qbep
                if (not _pu_is_pd and _pu_hq_mode == "3-point parametric")
                else (pe.hq_max_flow(_pu_hq_coeffs) * 0.65
                      if (not _pu_is_pd and _pu_hq_coeffs) else 0.0)
            )
            try:
                _pu = compute_pump_case(
                    is_pd      = _pu_is_pd,
                    rho        = _pu_rho,
                    Pv         = _pu_Pv,
                    P_bara     = _pu_P_bara,
                    hq_coeffs  = _pu_hq_coeffs  if not _pu_is_pd else None,
                    eta_params = _pu_eta_params  if (not _pu_is_pd and _pu_hq_coeffs) else None,
                    n_ratio    = _pu_n_ratio     if not _pu_is_pd else 1.0,
                    Qbep_used  = _pu_Qbep_used,
                    H_static   = _pu_H_static,
                    k_fric     = _pu_k_fric,
                    eta_motor  = _pu_eta_motor   if not _pu_is_pd else 93.0,
                    z_suc      = _pu_z_suc       if not _pu_is_pd else 0.0,
                    h_suc_loss = _pu_h_suc_loss  if not _pu_is_pd else 0.0,
                    NPSH_R     = _pu_NPSH_R      if not _pu_is_pd else 0.0,
                    dp_method  = _pu_dp_method   if not _pu_is_pd else 1,
                    P_suc_max  = _pu_P_suc_max   if not _pu_is_pd else _pu_P_bara,
                    PSV_set    = _pu_PSV_set,
                    mat_group  = _pu_mat_group,
                    pd_acc_pct = _pu_pd_acc_pct  if _pu_is_pd else 10.0,
                )
            except ValueError as _pu_e:
                st.warning(str(_pu_e))
                st.stop()
            except Exception as _pu_e:
                st.error(f"Pump calculation error: {_pu_e}")
                st.stop()

            # ── KPIs ──────────────────────────────────────────────────────────
            if not _pu_is_pd and _pu["op_ok"]:
                st.markdown("**OPERATING POINT**")
                _k1, _k2, _k3, _k4 = st.columns(4)
                _k1.metric("Flow", f"{_pu['Q_op']:.1f} m³/h")
                _k2.metric("Head", f"{_pu['H_op']:.1f} m",
                           delta=f"{_pu['Hop_bar']:.2f} bar", delta_color="off")
                _k3.metric("η pump", f"{_pu['eta_op']:.1f} %")
                _k4.metric("P shaft", f"{_pu['P_shaft']:.1f} kW")

                _k5, _k6, _k7, _k8 = st.columns(4)
                _k5.metric("P motor (input)", f"{_pu['P_motor']:.1f} kW",
                           help=f"Next IEC frame: {_pu['P_frame']:.0f} kW")
                _k6.metric("NPSH available", f"{_pu['NPSH_A']:.2f} m")
                _k7.metric("NPSH required", f"{_pu_NPSH_R:.2f} m")
                _npsh_delta = f"{_pu['npsh_margin']:+.2f} m  ({_pu['npsh_status']})"
                _k8.metric("NPSH margin", f"{_pu['npsh_margin']:.2f} m",
                           delta=_npsh_delta,
                           delta_color="normal" if _pu["npsh_color"] == "green"
                                       else ("off" if _pu["npsh_color"] == "orange" else "inverse"))

                # BEP deviation warning
                if _pu["Q_op"] < 0.70 * _pu_Qbep_used:
                    st.warning(
                        f"Operating flow ({_pu['Q_op']:.1f} m³/h) is below 70 % of BEP "
                        f"({_pu_Qbep_used:.1f} m³/h). Risk of internal recirculation, "
                        "vibration, and premature seal/bearing wear."
                    )
                elif _pu["Q_op"] > 1.10 * _pu_Qbep_used:
                    st.warning(
                        f"Operating flow ({_pu['Q_op']:.1f} m³/h) exceeds 110 % of BEP. "
                        "Risk of cavitation, motor overload, and reduced seal life."
                    )

                st.divider()

                # ── H-Q + System curve chart ──────────────────────────────────
                st.markdown("**PUMP & SYSTEM CURVES**")
                _fig_hq = go.Figure()
                _fig_hq.add_trace(go.Scatter(
                    x=_pu["Q_plot"], y=_pu["H_plot"],
                    name="Pump H-Q", line=dict(color="#2563EB", width=2.5),
                ))
                _fig_hq.add_trace(go.Scatter(
                    x=_pu["Q_plot"], y=_pu["Hs_plot"],
                    name="System curve", line=dict(color="#D97706", width=2.5, dash="dash"),
                ))
                _fig_hq.add_trace(go.Scatter(
                    x=[_pu["Q_op"]], y=[_pu["H_op"]],
                    name="Operating point",
                    mode="markers",
                    marker=dict(size=12, color="#16A34A", symbol="circle",
                                line=dict(color="white", width=2)),
                ))
                _pu_Hbep_on_curve = pe.eval_hq(_pu_hq_coeffs, _pu_Qbep_used)
                _fig_hq.add_trace(go.Scatter(
                    x=[_pu_Qbep_used], y=[_pu_Hbep_on_curve],
                    name="BEP",
                    mode="markers",
                    marker=dict(size=10, color="#7C3AED", symbol="diamond",
                                line=dict(color="white", width=2)),
                ))
                _fig_hq.update_layout(
                    xaxis_title="Flow (m³/h)", yaxis_title="Head (m)",
                    height=300, margin=dict(l=10, r=10, t=10, b=10),
                    legend=dict(orientation="h", y=-0.25),
                )
                st.plotly_chart(_fig_hq, use_container_width=True)

                # η-Q chart
                _fig_eta = go.Figure()
                _fig_eta.add_trace(go.Scatter(
                    x=_pu["Q_plot"], y=_pu["eta_plot"],
                    name="Efficiency", line=dict(color="#7C3AED", width=2.5),
                    fill="tozeroy", fillcolor="rgba(124,58,237,0.08)",
                ))
                _fig_eta.add_vline(x=_pu["Q_op"], line=dict(color="#16A34A", dash="dot", width=1.5),
                                   annotation_text=f"Q_op={_pu['Q_op']:.1f}", annotation_position="top right")
                _fig_eta.update_layout(
                    xaxis_title="Flow (m³/h)", yaxis_title="Efficiency (%)",
                    yaxis_range=[0, 100],
                    height=200, margin=dict(l=10, r=10, t=10, b=10),
                    showlegend=False,
                )
                st.plotly_chart(_fig_eta, use_container_width=True)

                st.divider()

            # ── Design pressure results ───────────────────────────────────────
            st.markdown("**DESIGN PRESSURE — DOWNSTREAM PIPING**")
            if not _pu_is_pd and _pu_hq_coeffs:
                st.caption(
                    f"Shut-off head (at {'max VSD' if _pu_vsd else 'rated'} speed): "
                    f"**{_pu['H0_max']:.1f} m**  =  **{_pu['H0_bar']:.2f} bar**  "
                    f"(ρ = {_pu_rho:.0f} kg/m³)"
                )
            _dp1, _dp2, _dp3 = st.columns(3)
            _dp1.metric("P design", f"{_pu['P_design_bara']:.2f} bara")
            _dp2.metric("P design", f"{_pu['P_design_barg']:.2f} barg")
            _dp3.metric("ANSI class", _pu["ansi"]["class_label"],
                        delta=f"Rated {_pu['ansi']['rated_barg']:.1f} barg",
                        delta_color="normal" if _pu["ansi"]["adequate"] else "inverse")

            if not _pu["ansi"]["adequate"]:
                st.error(
                    f"Design pressure ({_pu['P_design_barg']:.2f} barg) exceeds the maximum "
                    f"ANSI 2500 rating ({_pu['ansi']['rated_barg']:.1f} barg) for the selected "
                    "material group. Review design basis or use a higher-pressure standard."
                )

            st.caption(_pu["dp_res"]["notes"])

            # ── ANSI class table ──────────────────────────────────────────────
            with st.expander("ANSI B16.5 pressure class ratings — all classes"):
                _ansi_df = pd.DataFrame(_pu["ansi"]["all_classes"])
                st.dataframe(_ansi_df, hide_index=True, use_container_width=True)

            # ── Design pressure method comparison table ────────────────────────
            if not _pu_is_pd:
                with st.expander("Design pressure method comparison"):
                    st.markdown(
                        "All six methods evaluated at current inputs. "
                        "Methods 4–6 require a PSV on the pump discharge."
                    )
                    _cmp_rows = []
                    for _row in _pu["method_comparison"]:
                        _sel = " ◄ selected" if _row["#"] == _pu_dp_method else ""
                        _p_val = _row["P_design_barg"]
                        _cmp_rows.append({
                            "#":               _row["#"],
                            "Method":          _row["method_label"],
                            "P design (barg)": _p_val if _p_val is not None else "—  (PSV set required)",
                            "ANSI class":      (_row["ansi_label"] + _sel) if _p_val is not None else "—",
                        })
                    st.dataframe(
                        pd.DataFrame(_cmp_rows),
                        hide_index=True, use_container_width=True,
                        column_config={"P design (barg)": st.column_config.NumberColumn(format="%.2f")},
                    )

            # ── Word export ───────────────────────────────────────────────────
            st.divider()
            _pu_fluid_lbl = (
                f"KOH {_pu_koh_conc} wt%"
                if _pu_fluid_type == "KOH solution"
                else str(_pu_fluid_name)
            )
            _pu_inp = [
                ("Pump type",          _pu_type),
                ("Fluid",              _pu_fluid_lbl),
                ("Temperature",        f"{_pu_T_C:.1f} °C"),
                ("Suction pressure",   f"{_pu_P_bara:.3f} bara"),
                ("Density ρ",          f"{_pu_rho:.2f} kg/m³"),
                ("Vapour pressure Pv", f"{_pu_Pv:.5f} bara"),
                ("H_static (m)",       f"{_pu_H_static:.2f}"),
                ("k_fric (m/(m³/h)²)", f"{_pu_k_fric:.6f}"),
            ]
            if not _pu_is_pd:
                _pu_inp += [
                    ("Motor efficiency",   f"{_pu_eta_motor:.1f} %"),
                    ("Suction elevation",  f"{_pu_z_suc:.2f} m"),
                    ("Suction head loss",  f"{_pu_h_suc_loss:.2f} m"),
                    ("NPSH required",      f"{_pu_NPSH_R:.2f} m"),
                    ("BEP flow",           f"{_pu_Qbep_used:.2f} m³/h"),
                    ("Design P method",    f"#{_pu_dp_method}"),
                    ("Material group",     _pu_mat_group),
                ]
            _pu_res_rows = [
                ("P design (bara)",    f"{_pu['P_design_bara']:.3f}"),
                ("P design (barg)",    f"{_pu['P_design_barg']:.3f}"),
                ("ANSI class",         _pu["ansi"]["class_label"]),
                ("ANSI rated (barg)",  f"{_pu['ansi']['rated_barg']:.1f}"),
            ]
            if not _pu_is_pd and _pu["op_ok"]:
                _pu_res_rows = [
                    ("Operating flow Q",   f"{_pu['Q_op']:.2f} m³/h"),
                    ("Operating head H",   f"{_pu['H_op']:.2f} m  ({_pu['Hop_bar']:.3f} bar)"),
                    ("Pump efficiency η",  f"{_pu['eta_op']:.1f} %"),
                    ("Shaft power",        f"{_pu['P_shaft']:.2f} kW"),
                    ("Motor power (input)", f"{_pu['P_motor']:.2f} kW"),
                    ("NPSH available",     f"{_pu['NPSH_A']:.3f} m"),
                    ("NPSH required",      f"{_pu_NPSH_R:.2f} m"),
                    ("NPSH margin",        f"{_pu['npsh_margin']:+.3f} m  ({_pu['npsh_status']})"),
                ] + _pu_res_rows
            _pu_ansi_data = [
                [r["Class"], r["Rated (barg)"], r["Adequate"]]
                for r in _pu["ansi"]["all_classes"]
            ]
            _pu_rpt = report_generator.generate_calculator_report(
                tool_name="Pump",
                subtitle=f"{_pu_type}  ·  {_pu_fluid_lbl}  ·  P_design = {_pu['P_design_barg']:.2f} barg",
                method_text=(
                    "Centrifugal pump: H-Q curve fitted to 3-point or tabular data "
                    "(polynomial H = a + b·Q + c·Q²). Operating point by intersection of "
                    "the pump curve and the system curve (H_static + k_fric·Q²). "
                    "Efficiency η by parabolic fit. Shaft power = ρ·g·Q·H/η. "
                    "NPSH available = (P_suc − Pv)/(ρg) + z_suc − h_suc_loss. "
                    "Design pressure per selected API 610 / NORSOK method; ANSI B16.5 "
                    "class from the design pressure and material group."
                    if not _pu_is_pd else
                    "Positive displacement pump: design pressure = suction pressure + "
                    "differential pressure (accumulation margin applied)."
                ),
                inputs_rows=_pu_inp,
                results_rows=_pu_res_rows,
                extra_tables=[{
                    "title": "ANSI B16.5 Pressure Class Ratings",
                    "headers": ["Class", "Rated (barg)", "Adequate"],
                    "data": _pu_ansi_data,
                    "col_widths": [1.2, 1.5, 0.9],
                }],
                fig=_fig_hq if (not _pu_is_pd and _pu["op_ok"]) else None,
                fig_caption_text="Pump H-Q curve and system curve with operating point.",
                fig_height=330,
            )
            st.download_button(
                "Export Word (.docx)",
                _pu_rpt,
                file_name=f"pump_{_pu_type.split()[0].lower()}_{_pu_fluid_lbl.replace(' ','_')}.docx",
                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                key="pu_dl",
            )

    # =========================================================================
    # Tab: Line Size
    # =========================================================================
    with tab_ls:
        import math as _ls_math

        st.markdown(
            "Quick pipe size selection — find the minimum DN that meets velocity "
            "and pressure-drop-per-100-m criteria for a given flow and fluid."
        )

        # ── Service presets ──────────────────────────────────────────────────
        _LS_PRESETS = {
            "General process liquid":     dict(v_min=1.0,  v_max=3.0,  dp=50.0),
            "Pump suction line":          dict(v_min=0.3,  v_max=1.5,  dp=20.0),
            "Pump discharge line":        dict(v_min=1.5,  v_max=3.0,  dp=100.0),
            "Cooling water":              dict(v_min=1.0,  v_max=3.0,  dp=50.0),
            "Slurry / solids-bearing":    dict(v_min=1.5,  v_max=3.5,  dp=100.0),
            "Low-pressure gas (< 5 bar)": dict(v_min=5.0,  v_max=15.0, dp=50.0),
            "High-pressure gas":          dict(v_min=10.0, v_max=25.0, dp=20.0),
            "Steam (low pressure)":       dict(v_min=20.0, v_max=40.0, dp=100.0),
            "Custom":                     dict(v_min=None, v_max=None,  dp=None),
        }

        _ls_c1, _ls_c2 = st.columns([1, 1.3], gap="large")

        with _ls_c1:
            # ── Phase & fluid ────────────────────────────────────────────────
            st.markdown("**FLUID**")
            _ls_phase = st.radio(
                "Phase", ["Liquid", "Gas"], horizontal=True,
                key="ls_phase", label_visibility="collapsed",
            )
            _ls_T = st.number_input(
                "Temperature (°C)", value=25.0,
                min_value=-50.0, max_value=300.0, step=5.0, key="ls_T",
            )
            _ls_P = st.number_input(
                "Pressure (bara)", value=5.0,
                min_value=0.1, max_value=500.0, step=0.5, key="ls_P",
            )

            if _ls_phase == "Liquid":
                _ls_liq_src = st.radio(
                    "Source", ["KOH solution", "CoolProp liquid"],
                    horizontal=True, key="ls_liq_src", label_visibility="collapsed",
                )
                if _ls_liq_src == "KOH solution":
                    _ls_koh_c = st.slider(
                        "KOH concentration (wt%)", 20, 40, 30, key="ls_koh_c",
                    )
                    _ls_rho, _ls_mu, _ = engine.koh_properties(_ls_T, _ls_koh_c)
                else:
                    _ls_cp_opts = list(engine.LIQUID_COOLPROP_ID.keys())
                    _ls_cp_fluid = st.selectbox(
                        "Fluid", _ls_cp_opts,
                        index=_ls_cp_opts.index(
                            st.session_state.get("ls_cp_fluid", "Water")
                        ) if st.session_state.get("ls_cp_fluid", "Water") in _ls_cp_opts else 0,
                        key="ls_cp_fluid",
                    )
                    import CoolProp.CoolProp as _ls_CP
                    _ls_cpid = engine.LIQUID_COOLPROP_ID[_ls_cp_fluid]
                    try:
                        _ls_rho = _ls_CP.PropsSI("D", "T", _ls_T+273.15, "P", _ls_P*1e5, _ls_cpid)
                        _ls_mu  = _ls_CP.PropsSI("V", "T", _ls_T+273.15, "P", _ls_P*1e5, _ls_cpid)
                    except Exception:
                        _ls_rho, _ls_mu = 1000.0, 1e-3
            else:
                _ls_gas_opts = [k for k in engine.GAS_SPECIES if k != "Custom"]
                _ls_gas = st.selectbox("Gas species", _ls_gas_opts, key="ls_gas")
                _ls_MW  = engine.GAS_SPECIES[_ls_gas]["MW"]   # kg/mol
                _ls_rho = _ls_P * 1e5 * _ls_MW / (8.31446 * (_ls_T + 273.15))
                import CoolProp.CoolProp as _ls_CP
                _ls_cpid_g = engine.GAS_SPECIES[_ls_gas].get("coolprop_id")
                try:
                    _ls_mu = _ls_CP.PropsSI("V", "T", _ls_T+273.15, "P", _ls_P*1e5, _ls_cpid_g) if _ls_cpid_g else engine.GAS_SPECIES[_ls_gas].get("mu_ref", 1.8e-5)
                except Exception:
                    _ls_mu = engine.GAS_SPECIES[_ls_gas].get("mu_ref", 1.8e-5)

            _lf1, _lf2 = st.columns(2)
            if _ls_phase == "Liquid":
                _lf1.metric("ρ", f"{_ls_rho:.1f} kg/m³")
                _lf2.metric("μ", f"{_ls_mu*1e3:.3f} mPa·s")
            else:
                _lf1.metric("ρ", f"{_ls_rho:.4f} kg/m³")
                _lf2.metric("μ", f"{_ls_mu*1e6:.2f} μPa·s")
                st.caption("Density at inlet conditions (incompressible approximation). "
                           "Accurate when ΔP < 10 % of inlet pressure.")

            st.divider()

            # ── Flow rate ────────────────────────────────────────────────────
            st.markdown("**FLOW RATE**")
            _ls_fu = st.radio(
                "Unit", ["kg/h", "m³/h"], horizontal=True,
                key="ls_fu", label_visibility="collapsed",
            )
            _ls_fv = st.number_input(
                "Flow", value=10000.0, min_value=0.001, step=100.0,
                format="%.3f", key="ls_fv",
            )
            if _ls_fu == "kg/h":
                _ls_m_kgh  = _ls_fv
                _ls_Q_m3h  = _ls_m_kgh / max(_ls_rho, 1e-6)
            else:
                _ls_Q_m3h  = _ls_fv
                _ls_m_kgh  = _ls_Q_m3h * _ls_rho
            st.caption(f"= {_ls_Q_m3h:.3f} m³/h  ·  {_ls_m_kgh:.1f} kg/h")

            st.divider()

            # ── Pipe ─────────────────────────────────────────────────────────
            st.markdown("**PIPE**")
            _ls_pn_opts  = ["PN20", "PN25", "PN40"]
            _ls_mat_opts = list(engine.MATERIAL_ROUGHNESS.keys())
            _ls_dn_opts  = list(engine.PIPE_DATABASE.keys())

            _lp1, _lp2 = st.columns(2)
            _ls_pn  = _lp1.selectbox("PN rating", _ls_pn_opts, index=2, key="ls_pn")
            _ls_mat = _lp2.selectbox("Material", _ls_mat_opts, key="ls_mat")

            _ls_dn_min = st.selectbox(
                "Min DN", _ls_dn_opts, index=0, key="ls_dn_min",
            )
            _ls_dn_max = st.selectbox(
                "Max DN", _ls_dn_opts,
                index=len(_ls_dn_opts) - 1, key="ls_dn_max",
            )

            st.divider()

            # ── Sizing criteria ───────────────────────────────────────────────
            st.markdown("**SIZING CRITERIA**")
            _ls_preset = st.selectbox(
                "Service preset", list(_LS_PRESETS.keys()), key="ls_preset",
            )
            _pre = _LS_PRESETS[_ls_preset]
            if _pre["v_min"] is not None:
                st.caption(
                    f"Guidance for *{_ls_preset}*: "
                    f"v = {_pre['v_min']}–{_pre['v_max']} m/s, "
                    f"ΔP/100m ≤ {_pre['dp']} kPa"
                )

            _sc1, _sc2 = st.columns(2)
            _ls_v_min = _sc1.number_input(
                "Min velocity (m/s)",
                value=float(_pre["v_min"] if _pre["v_min"] is not None else 1.0),
                min_value=0.0, step=0.5, key="ls_v_min",
            )
            _ls_v_max = _sc2.number_input(
                "Max velocity (m/s)",
                value=float(_pre["v_max"] if _pre["v_max"] is not None else 3.0),
                min_value=0.0, step=0.5, key="ls_v_max",
            )
            _ls_dp_max = st.number_input(
                "Max ΔP/100m (kPa)",
                value=float(_pre["dp"] if _pre["dp"] is not None else 50.0),
                min_value=0.0, step=5.0, key="ls_dp_max",
            )

        # ── RIGHT COLUMN — RESULTS ────────────────────────────────────────────
        with _ls_c2:

            _ls_eps    = engine.MATERIAL_ROUGHNESS.get(_ls_mat, 1.5e-5)
            _ls_Q_m3s  = _ls_Q_m3h / 3600.0
            _ls_dn_i   = _ls_dn_opts.index(_ls_dn_min)
            _ls_dn_j   = _ls_dn_opts.index(_ls_dn_max)
            _ls_dns    = _ls_dn_opts[_ls_dn_i: _ls_dn_j + 1]

            _ls_rows         = []
            _ls_recommended  = None

            for _dn in _ls_dns:
                _pndb = engine.PIPE_DATABASE.get(_dn, {})
                if _ls_pn not in _pndb:
                    continue
                _D   = _pndb[_ls_pn]            # bore in metres
                _A   = _ls_math.pi / 4.0 * _D ** 2
                _v   = _ls_Q_m3s / _A
                _Re  = _ls_rho * _v * _D / max(_ls_mu, 1e-12)
                _eD  = _ls_eps / _D
                _f   = churchill_f(_Re, _eD)
                _dp_pa_m     = _f * (_ls_rho * _v ** 2 / 2.0) / _D
                _dp_kpa_100m = _dp_pa_m * 100.0 / 1000.0

                _regime = ("Laminar" if _Re < 2300
                           else ("Transitional" if _Re < 4000 else "Turbulent"))
                _v_ok  = _ls_v_min <= _v <= _ls_v_max
                _dp_ok = _dp_kpa_100m <= _ls_dp_max
                _ok    = _v_ok and _dp_ok

                if _ok and _ls_recommended is None:
                    _ls_recommended = _dn

                _ls_rows.append({
                    "DN":               _dn,
                    "ID (mm)":          round(_D * 1000, 1),
                    "v (m/s)":          round(_v, 3),
                    "Re":               int(_Re),
                    "Regime":           _regime,
                    "ΔP/100m (kPa)":    round(_dp_kpa_100m, 2),
                    "v ✓":              "✓" if _v_ok  else "✗",
                    "ΔP ✓":             "✓" if _dp_ok else "✗",
                    "Adequate":         "✓" if _ok    else "—",
                })

            if not _ls_rows:
                st.warning("No DN entries found for the selected PN rating in this range.")
            else:
                if _ls_recommended:
                    st.success(
                        f"**Recommended: {_ls_recommended}** — "
                        f"smallest DN meeting both velocity and ΔP/100m criteria."
                    )
                else:
                    st.warning(
                        "No DN in the selected range meets all criteria. "
                        "Try a larger DN range or relax the criteria."
                    )

                _ls_df = pd.DataFrame(_ls_rows)
                st.dataframe(
                    _ls_df, hide_index=True, use_container_width=True,
                    column_config={
                        "v (m/s)":       st.column_config.NumberColumn(format="%.3f"),
                        "Re":            st.column_config.NumberColumn(format="%d"),
                        "ΔP/100m (kPa)": st.column_config.NumberColumn(format="%.2f"),
                    },
                )

                # ── Velocity and ΔP/100m vs DN chart ──────────────────────────
                _ls_dns_plot  = [r["DN"]            for r in _ls_rows]
                _ls_v_plot    = [r["v (m/s)"]       for r in _ls_rows]
                _ls_dp_plot   = [r["ΔP/100m (kPa)"] for r in _ls_rows]

                _fig_ls = go.Figure()

                # Velocity trace (primary y)
                _fig_ls.add_trace(go.Scatter(
                    x=_ls_dns_plot, y=_ls_v_plot, name="Velocity (m/s)",
                    mode="lines+markers",
                    line=dict(color="#2563EB", width=2.5),
                    marker=dict(size=7),
                    yaxis="y1",
                ))
                # Velocity band shading
                _fig_ls.add_hrect(
                    y0=_ls_v_min, y1=_ls_v_max,
                    fillcolor="rgba(37,99,235,0.08)", line_width=0,
                    annotation_text="v target", annotation_position="top left",
                    yref="y1",
                )

                # ΔP/100m trace (secondary y)
                _fig_ls.add_trace(go.Scatter(
                    x=_ls_dns_plot, y=_ls_dp_plot, name="ΔP/100m (kPa)",
                    mode="lines+markers",
                    line=dict(color="#D97706", width=2.5, dash="dash"),
                    marker=dict(size=7),
                    yaxis="y2",
                ))
                # ΔP limit line
                _fig_ls.add_hline(
                    y=_ls_dp_max, line=dict(color="#D97706", dash="dot", width=1.5),
                    annotation_text=f"ΔP limit {_ls_dp_max} kPa",
                    annotation_position="bottom right",
                    yref="y2",
                )

                # Recommended DN vertical marker (add_vline doesn't support categorical axes)
                if _ls_recommended:
                    _fig_ls.add_shape(
                        type="line",
                        x0=_ls_recommended, x1=_ls_recommended,
                        y0=0, y1=1, yref="paper",
                        line=dict(color="#16A34A", dash="dot", width=2),
                    )
                    _fig_ls.add_annotation(
                        x=_ls_recommended, y=1.02, yref="paper",
                        text=f"▼ {_ls_recommended}",
                        showarrow=False, yanchor="bottom",
                        font=dict(color="#16A34A", size=12),
                    )

                _fig_ls.update_layout(
                    xaxis_title="Pipe size",
                    yaxis=dict(title=dict(text="Velocity (m/s)", font=dict(color="#2563EB"))),
                    yaxis2=dict(
                        title=dict(text="ΔP/100m (kPa)", font=dict(color="#D97706")),
                        overlaying="y", side="right",
                    ),
                    legend=dict(orientation="h", y=-0.25),
                    height=320, margin=dict(l=10, r=10, t=20, b=10),
                )
                st.plotly_chart(_fig_ls, use_container_width=True)

                # ── Export buttons ────────────────────────────────────────────
                st.divider()
                _ls_fluid_lbl = (
                    f"KOH {_ls_koh_c} wt%"
                    if _ls_phase == "Liquid" and _ls_liq_src == "KOH solution"
                    else (_ls_cp_fluid if _ls_phase == "Liquid" else _ls_gas)
                )
                _ls_criteria_str = (
                    f"v = {_ls_v_min}–{_ls_v_max} m/s  |  ΔP/100m ≤ {_ls_dp_max} kPa  |  {_ls_preset}"
                )
                _ls_inp_rows = [
                    ("Phase",              _ls_phase),
                    ("Fluid",              _ls_fluid_lbl),
                    ("Temperature",        f"{_ls_T:.1f} °C"),
                    ("Pressure",           f"{_ls_P:.3f} bara"),
                    ("Density ρ",          f"{_ls_rho:.4f} kg/m³"),
                    ("Viscosity μ",        f"{_ls_mu*1e6:.2f} μPa·s" if _ls_phase == "Gas"
                                           else f"{_ls_mu*1e3:.4f} mPa·s"),
                    ("Flow rate",          f"{_ls_m_kgh:.2f} kg/h  ({_ls_Q_m3h:.4f} m³/h)"),
                    ("PN rating",          _ls_pn),
                    ("Material",           _ls_mat),
                    ("DN range",           f"{_ls_dn_min} – {_ls_dn_max}"),
                    ("Service preset",     _ls_preset),
                    ("Min velocity",       f"{_ls_v_min:.2f} m/s"),
                    ("Max velocity",       f"{_ls_v_max:.2f} m/s"),
                    ("Max ΔP/100m",        f"{_ls_dp_max:.2f} kPa"),
                ]
                _ls_res_rows = [
                    ("Recommended DN",     _ls_recommended if _ls_recommended else "None in range"),
                    ("Sizing criteria",    _ls_criteria_str),
                ]
                _ls_table_data = [
                    [r["DN"], str(r["ID (mm)"]), f"{r['v (m/s)']:.3f}",
                     f"{r['Re']:,}", r["Regime"], f"{r['ΔP/100m (kPa)']:.2f}",
                     r["v ✓"], r["ΔP ✓"], r["Adequate"]]
                    for r in _ls_rows
                ]
                _ls_exp_c1, _ls_exp_c2 = st.columns(2)
                with _ls_exp_c1:
                    _ls_rpt = report_generator.generate_calculator_report(
                        tool_name="Line Size",
                        subtitle=f"{_ls_phase}  ·  {_ls_fluid_lbl}  ·  {_ls_Q_m3h:.3f} m³/h",
                        method_text=(
                            "Darcy-Weisbach friction pressure drop per 100 m of straight pipe, "
                            "evaluated for each DN in the selected range. "
                            "Friction factor from Churchill (1977) covering laminar, transitional, "
                            "and turbulent regimes. Velocity computed from volumetric flow rate "
                            "and pipe bore area. The recommended DN is the smallest that meets "
                            "both the velocity band and the ΔP/100m limit simultaneously."
                        ),
                        inputs_rows=_ls_inp_rows,
                        results_rows=_ls_res_rows,
                        extra_tables=[{
                            "title": "DN Sizing Table",
                            "headers": ["DN","ID (mm)","v (m/s)","Re","Regime",
                                        "ΔP/100m (kPa)","v ✓","ΔP ✓","Adequate"],
                            "data": _ls_table_data,
                            "col_widths": [0.55, 0.6, 0.65, 0.7, 0.75, 0.85, 0.4, 0.4, 0.55],
                        }],
                        fig=_fig_ls,
                        fig_caption_text="Velocity and ΔP/100m vs pipe size.",
                        fig_height=340,
                    )
                    st.download_button(
                        "Export Word (.docx)",
                        _ls_rpt,
                        file_name=f"linesize_{_ls_fluid_lbl.replace(' ','_')}_{_ls_pn}.docx",
                        mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                        use_container_width=True,
                        key="ls_dl_w",
                    )
                with _ls_exp_c2:
                    import io as _ls_io
                    _ls_xl_buf = _ls_io.BytesIO()
                    with pd.ExcelWriter(_ls_xl_buf, engine="openpyxl") as _ls_xw:
                        _ls_info = pd.DataFrame([
                            ["Phase",        _ls_phase],
                            ["Fluid",        _ls_fluid_lbl],
                            ["Temperature",  f"{_ls_T:.1f} °C"],
                            ["Pressure",     f"{_ls_P:.3f} bara"],
                            ["Density",      f"{_ls_rho:.4f} kg/m³"],
                            ["Viscosity",    f"{_ls_mu*1e3:.4f} mPa·s" if _ls_phase == "Liquid"
                                             else f"{_ls_mu*1e6:.2f} μPa·s"],
                            ["Flow rate",    f"{_ls_m_kgh:.2f} kg/h"],
                            ["Flow rate",    f"{_ls_Q_m3h:.4f} m³/h"],
                            ["PN rating",    _ls_pn],
                            ["Material",     _ls_mat],
                            ["Service",      _ls_preset],
                            ["Min velocity", f"{_ls_v_min:.2f} m/s"],
                            ["Max velocity", f"{_ls_v_max:.2f} m/s"],
                            ["Max ΔP/100m",  f"{_ls_dp_max:.2f} kPa"],
                            ["Recommended",  _ls_recommended or "None"],
                        ], columns=["Parameter", "Value"])
                        _ls_info.to_excel(_ls_xw, sheet_name="Inputs", index=False)
                        pd.DataFrame(_ls_rows).to_excel(
                            _ls_xw, sheet_name="DN Sizing Table", index=False)
                    _ls_xl_buf.seek(0)
                    st.download_button(
                        "Export Excel (.xlsx)",
                        _ls_xl_buf,
                        file_name=f"linesize_{_ls_fluid_lbl.replace(' ','_')}_{_ls_pn}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="ls_dl_x",
                    )


